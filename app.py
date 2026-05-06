#!/usr/bin/env python3
"""
Freshdesk AI Ticket Analyzer - Web Application
A web dashboard that fetches Freshdesk tickets, analyzes them with Claude AI,
and displays results for team review.
"""

import json
import logging
import os
import re
import html as html_lib
import sqlite3
import time
import threading
from datetime import datetime, timezone, timedelta
from functools import wraps

import requests
from requests.auth import HTTPBasicAuth
from anthropic import Anthropic
try:
    from ai.llm.gateway import LLMGateway
    from ai.llm.registry import MODEL_REGISTRY
except ImportError:
    LLMGateway = None
    MODEL_REGISTRY = {}
import subprocess
import tempfile
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, g, send_file
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
import io
from export_report import generate_pdf, generate_pptx
from agents import AgentOrchestrator, init_agent_tables

load_dotenv()

# ── App Setup ────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "change-me-in-production-" + os.urandom(8).hex())

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR = os.getenv("DATA_DIR", _APP_DIR)
os.makedirs(_DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(_DATA_DIR, "analyzer.db")

# Global state for background job (protected by _job_lock for thread safety)
_job_lock = threading.Lock()
job_status = {"running": False, "progress": "", "processed": 0, "total": 0, "errors": []}


def _update_job(**kwargs):
    """Thread-safe update of job_status fields."""
    with _job_lock:
        job_status.update(kwargs)


def _job_add_error(msg):
    """Thread-safe append to job_status errors."""
    with _job_lock:
        job_status["errors"].append(msg)


def _job_increment_processed():
    """Thread-safe increment of job_status processed count."""
    with _job_lock:
        job_status["processed"] += 1


def _get_job_status():
    """Thread-safe read of job_status (returns a snapshot copy)."""
    with _job_lock:
        return dict(job_status, errors=list(job_status["errors"]))


# ── Database ─────────────────────────────────────────────────────────────────

def get_db():
    """Get database connection for current request.
    Uses check_same_thread=False because agent orchestrator may share this
    connection across ThreadPoolExecutor workers (WAL mode handles concurrency)."""
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, check_same_thread=False)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def get_db_standalone():
    """Get database connection outside Flask request context.
    Uses check_same_thread=False because agent orchestrator may share this
    connection across ThreadPoolExecutor workers (WAL mode handles concurrency)."""
    db = sqlite3.connect(DB_PATH, check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    return db


def _row_get(row, key, default=""):
    """Safely get a value from a sqlite3.Row (which doesn't support .get()).
    Returns default if the key doesn't exist or the value is None."""
    try:
        val = row[key]
        return val if val is not None else default
    except (KeyError, IndexError):
        return default


def init_db():
    """Create database tables if they don't exist."""
    db = sqlite3.connect(DB_PATH)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS tickets (
            id INTEGER PRIMARY KEY,
            ticket_id INTEGER UNIQUE NOT NULL,
            ticket_url TEXT,
            subject TEXT,
            status TEXT,
            status_code INTEGER,
            priority TEXT,
            priority_code INTEGER,
            group_name TEXT,
            requester_name TEXT,
            requester_email TEXT,
            country TEXT,
            created_at TEXT,
            updated_at TEXT,
            last_analysis TEXT,
            classification TEXT,
            confidence INTEGER DEFAULT 0,
            needs_review TEXT DEFAULT 'Yes',
            summary TEXT,
            analysis TEXT,
            draft_response TEXT,
            next_steps TEXT,
            sources TEXT,
            review_status TEXT DEFAULT 'Pending',
            processing_date TEXT,
            risk_level TEXT,
            raw_description TEXT,
            raw_conversations TEXT,
            compiled_thread TEXT,
            draft_response_en TEXT,
            rice_reach REAL DEFAULT 0,
            rice_impact REAL DEFAULT 0,
            rice_confidence REAL DEFAULT 0,
            rice_effort REAL DEFAULT 0,
            rice_score REAL DEFAULT 0
        );

        CREATE INDEX IF NOT EXISTS idx_tickets_ticket_id ON tickets(ticket_id);
        CREATE INDEX IF NOT EXISTS idx_tickets_status ON tickets(review_status);
        CREATE INDEX IF NOT EXISTS idx_tickets_classification ON tickets(classification);
        CREATE INDEX IF NOT EXISTS idx_tickets_risk ON tickets(risk_level);

        CREATE TABLE IF NOT EXISTS knowledge_base (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            title TEXT NOT NULL,
            content TEXT NOT NULL DEFAULT '',
            entry_type TEXT NOT NULL DEFAULT 'text',
            file_path TEXT DEFAULT '',
            file_name TEXT DEFAULT '',
            url TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_kb_category ON knowledge_base(category);

        CREATE TABLE IF NOT EXISTS terminology (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            term_fr TEXT NOT NULL,
            term_en TEXT NOT NULL,
            definition TEXT DEFAULT '',
            category TEXT DEFAULT 'general',
            ecdf_reference TEXT DEFAULT '',
            usage_context TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_terminology_category ON terminology(category);
    """)
    db.commit()

    # Migrate: add new columns to tickets if they don't exist
    cursor_t = db.execute("PRAGMA table_info(tickets)")
    ticket_cols = {row[1] for row in cursor_t.fetchall()}
    ticket_migrations = {
        "draft_response_en": "ALTER TABLE tickets ADD COLUMN draft_response_en TEXT",
        "rice_reach": "ALTER TABLE tickets ADD COLUMN rice_reach REAL DEFAULT 0",
        "rice_impact": "ALTER TABLE tickets ADD COLUMN rice_impact REAL DEFAULT 0",
        "rice_confidence": "ALTER TABLE tickets ADD COLUMN rice_confidence REAL DEFAULT 0",
        "rice_effort": "ALTER TABLE tickets ADD COLUMN rice_effort REAL DEFAULT 0",
        "rice_score": "ALTER TABLE tickets ADD COLUMN rice_score REAL DEFAULT 0",
        "po_decision": "ALTER TABLE tickets ADD COLUMN po_decision TEXT DEFAULT 'pending'",
        "po_decision_reason": "ALTER TABLE tickets ADD COLUMN po_decision_reason TEXT DEFAULT ''",
        "po_decision_date": "ALTER TABLE tickets ADD COLUMN po_decision_date TEXT DEFAULT ''",
        "prd_content": "ALTER TABLE tickets ADD COLUMN prd_content TEXT DEFAULT ''",
        "responded": "ALTER TABLE tickets ADD COLUMN responded TEXT DEFAULT 'In Inbox'",
        "resolved_at": "ALTER TABLE tickets ADD COLUMN resolved_at TEXT DEFAULT ''",
        "first_responded_at": "ALTER TABLE tickets ADD COLUMN first_responded_at TEXT DEFAULT ''",
        "template_name": "ALTER TABLE tickets ADD COLUMN template_name TEXT DEFAULT ''",
        "workflow_name": "ALTER TABLE tickets ADD COLUMN workflow_name TEXT DEFAULT ''",
        "sla_resolution_hours": "ALTER TABLE tickets ADD COLUMN sla_resolution_hours REAL DEFAULT 0",
        "sla_first_response_hours": "ALTER TABLE tickets ADD COLUMN sla_first_response_hours REAL DEFAULT 0",
        "screenshots_json": "ALTER TABLE tickets ADD COLUMN screenshots_json TEXT DEFAULT '[]'",
        "qa_status": "ALTER TABLE tickets ADD COLUMN qa_status TEXT DEFAULT ''",
        "qa_issues": "ALTER TABLE tickets ADD COLUMN qa_issues TEXT DEFAULT '[]'",
        "analysis_raw_output": "ALTER TABLE tickets ADD COLUMN analysis_raw_output TEXT DEFAULT ''",
        "last_learned_conv_id": "ALTER TABLE tickets ADD COLUMN last_learned_conv_id INTEGER DEFAULT 0",
        "jira_ticket_key": "ALTER TABLE tickets ADD COLUMN jira_ticket_key TEXT DEFAULT ''",
        "jira_ticket_url": "ALTER TABLE tickets ADD COLUMN jira_ticket_url TEXT DEFAULT ''",
        "bso_status": "ALTER TABLE tickets ADD COLUMN bso_status TEXT DEFAULT ''",
    }
    for col, sql in ticket_migrations.items():
        if col not in ticket_cols:
            try:
                db.execute(sql)
            except sqlite3.OperationalError:
                pass

    # Normalize po_decision values to lowercase (fix historical Title-case from bulk action)
    try:
        db.execute("UPDATE tickets SET po_decision = LOWER(po_decision) WHERE po_decision != LOWER(po_decision)")
    except Exception:
        pass
    db.commit()

    # Migrate: add new columns to knowledge_base if they don't exist
    cursor = db.execute("PRAGMA table_info(knowledge_base)")
    existing_cols = {row[1] for row in cursor.fetchall()}
    migrations = {
        "entry_type": "ALTER TABLE knowledge_base ADD COLUMN entry_type TEXT NOT NULL DEFAULT 'text'",
        "file_path": "ALTER TABLE knowledge_base ADD COLUMN file_path TEXT DEFAULT ''",
        "file_name": "ALTER TABLE knowledge_base ADD COLUMN file_name TEXT DEFAULT ''",
        "url": "ALTER TABLE knowledge_base ADD COLUMN url TEXT DEFAULT ''",
    }
    for col, sql in migrations.items():
        if col not in existing_cols:
            try:
                db.execute(sql)
            except sqlite3.OperationalError:
                pass  # Column might already exist

    # Fix existing tickets with wrong status labels (one-time migration)
    # Maps status_code to correct label using the current STATUS_MAP
    try:
        db.execute("UPDATE tickets SET status = 'In Progress' WHERE status_code = 3 AND status != 'In Progress'")
        db.execute("UPDATE tickets SET status = 'Pending Approval' WHERE status_code = 18 AND status != 'Pending Approval'")
        db.execute("UPDATE tickets SET status = 'Assign to 3L' WHERE status_code = 20 AND status != 'Assign to 3L'")
        # Also fix any remaining "Custom" or "Unknown" based on status_code
        for code, label in {2: "Open", 4: "Resolved", 5: "Closed", 6: "Waiting on Customer", 7: "Waiting on Third Party"}.items():
            db.execute("UPDATE tickets SET status = ? WHERE status_code = ? AND status IN ('Custom', 'Unknown')", (label, code))
    except sqlite3.OperationalError:
        pass

    db.commit()

    # Create agent_model_config table (provider-agnostic per-agent LLM config)
    try:
        db.execute("""CREATE TABLE IF NOT EXISTS agent_model_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            agent_name TEXT UNIQUE NOT NULL,
            provider TEXT NOT NULL DEFAULT 'anthropic',
            model TEXT NOT NULL,
            temperature REAL DEFAULT 0.0,
            max_tokens INTEGER DEFAULT 2000,
            enabled INTEGER DEFAULT 1,
            fallback_provider TEXT DEFAULT '',
            fallback_model TEXT DEFAULT '',
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        db.commit()
    except Exception:
        pass

    # Seed default agent model configs if table is empty
    try:
        count = db.execute("SELECT COUNT(*) FROM agent_model_config").fetchone()[0]
        if count == 0:
            seed_agent_model_configs(db)
    except Exception:
        pass

    # Initialize terminology glossary
    init_terminology(db)
    db.close()


def init_terminology(db):
    """Pre-populate terminology table with Luxembourg accounting and legal terms."""
    # Check if terminology table is empty
    count = db.execute("SELECT COUNT(*) FROM terminology").fetchone()[0]
    if count > 0:
        return  # Already populated

    terms = [
        # Legal Entity Terms
        ("Gérant Unique", "Sole Manager", "Manager of a SARL with 1 manager", "legal_entity", "", "SARL governance"),
        ("Conseil de Gérance", "Board of Managers", "Management body of a SARL with 2+ managers", "legal_entity", "", "SARL governance"),
        ("Administrateur Unique", "Sole Director", "Director of an SA with 1 director", "legal_entity", "", "SA governance"),
        ("Conseil d'Administration", "Board of Directors", "Management body of an SA", "legal_entity", "", "SA governance"),
        ("Commissaire aux comptes", "Statutory auditor", "Auditor appointed by law", "legal_entity", "", "Audit oversight"),
        ("Réviseur d'entreprises agréé", "Approved auditor", "Auditor with special approval", "legal_entity", "", "Audit oversight"),
        ("Associé", "Shareholder (SARL)", "Shareholder in a SARL", "legal_entity", "", "SARL ownership"),
        ("Actionnaire", "Shareholder (SA)", "Shareholder in an SA", "legal_entity", "", "SA ownership"),

        # Accounting Terms
        ("Plan comptable normalisé", "Standard chart of accounts", "Luxembourg standard CoA", "accounting", "PCN", "Chart of accounts"),
        ("Comptes annuels", "Annual accounts", "Year-end financial statements", "accounting", "", "Financial statements"),
        ("Bilan", "Balance sheet", "Statement of financial position", "accounting", "", "Financial statements"),
        ("Compte de profits et pertes", "Profit and loss account", "Income statement", "accounting", "", "Financial statements"),
        ("Annexe", "Notes to the accounts", "Explanatory notes to annual accounts", "accounting", "", "Financial statements"),
        ("Immobilisations corporelles", "Tangible fixed assets", "Property, plant, equipment", "accounting", "", "Balance sheet items"),
        ("Immobilisations incorporelles", "Intangible fixed assets", "Software, patents, goodwill", "accounting", "", "Balance sheet items"),
        ("Immobilisations financières", "Financial fixed assets", "Investments, loans", "accounting", "", "Balance sheet items"),
        ("Actifs circulants", "Current assets", "Short-term assets", "accounting", "", "Balance sheet items"),
        ("Créances", "Receivables", "Money owed to company", "accounting", "", "Balance sheet items"),
        ("Dettes", "Payables", "Money owed by company", "accounting", "", "Balance sheet items"),
        ("Capitaux propres", "Equity", "Shareholders' equity", "accounting", "", "Balance sheet items"),
        ("Provisions", "Provisions", "Liabilities for uncertain obligations", "accounting", "", "Balance sheet items"),
        ("Frais de personnel", "Staff costs", "Employee wages and benefits", "accounting", "", "P&L items"),
        ("Chiffre d'affaires", "Turnover", "Total sales revenue", "accounting", "", "P&L items"),

        # Legal Documents
        ("Procès-verbal du Conseil de Gérance", "Minutes of the Board of Managers", "Record of SARL management decisions", "legal_document", "", "Corporate minutes"),
        ("Procès-verbal du Conseil d'Administration", "Minutes of the Board of Directors", "Record of SA board decisions", "legal_document", "", "Corporate minutes"),
        ("Assemblée Générale Ordinaire", "Ordinary General Meeting", "Annual shareholder meeting", "legal_document", "", "Corporate meetings"),
        ("Assemblée Générale Extraordinaire", "Extraordinary General Meeting", "Special shareholder meeting", "legal_document", "", "Corporate meetings"),
        ("Résolutions circulaires", "Written resolutions", "Decisions taken without meeting", "legal_document", "", "Corporate decisions"),
        ("Procuration", "Power of attorney", "Delegation of authority", "legal_document", "", "Corporate documents"),
        ("Convocation", "Convening notice", "Meeting notice to shareholders", "legal_document", "", "Corporate meetings"),
        ("Affectation du résultat", "Profit allocation", "Decision on profit distribution", "legal_document", "", "Profit decisions"),

        # Tax Terms
        ("Impôt sur le revenu des collectivités", "Corporate Income Tax (CIT)", "IRC - main corporate tax", "tax", "Form 500", "Corporate taxation"),
        ("Impôt commercial communal", "Municipal Business Tax (MBT)", "ICC - local business tax", "tax", "Form 506A", "Corporate taxation"),
        ("Impôt sur la fortune", "Net Wealth Tax (NWT)", "IF - wealth tax (abolished)", "tax", "", "Historical tax"),
        ("Bénéfice commercial", "Business profit", "Taxable profit before deductions", "tax", "", "Tax calculation"),
        ("Crédit d'impôt pour investissement", "Investment tax credit", "Tax credit for capital investment", "tax", "", "Tax incentives"),

        # eCDF Terms
        ("Centrale des bilans", "Central balance sheet office", "CNC - centralized filing repository", "ecdf", "eCDF", "Electronic filing"),
        ("RCSL", "Trade and Companies Register", "Public company registry", "ecdf", "eCDF", "Electronic filing"),
        ("Format eCDF", "eCDF format", "Electronic reporting format for Luxembourg", "ecdf", "eCDF", "Electronic filing"),
        ("Formulaire 500", "Form 500 (CIT)", "Corporate Income Tax return form", "ecdf", "eCDF", "Tax forms"),
        ("Formulaire 506A", "Form 506A (MBT)", "Municipal Business Tax return form", "ecdf", "eCDF", "Tax forms"),
    ]

    for term_fr, term_en, definition, category, ecdf_ref, usage in terms:
        db.execute(
            """INSERT INTO terminology (term_fr, term_en, definition, category, ecdf_reference, usage_context)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (term_fr, term_en, definition, category, ecdf_ref, usage)
        )

    db.commit()


# ── Settings Helpers ─────────────────────────────────────────────────────────

def get_setting(key, default="", db=None):
    """Get a setting value."""
    close_after = False
    if db is None:
        try:
            db = get_db()
        except RuntimeError:
            db = get_db_standalone()
            close_after = True

    row = db.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    if close_after:
        db.close()
    return row["value"] if row else default


def set_setting(key, value, db=None):
    """Set a setting value."""
    close_after = False
    if db is None:
        try:
            db = get_db()
        except RuntimeError:
            db = get_db_standalone()
            close_after = True

    db.execute(
        "INSERT OR REPLACE INTO settings (key, value, updated_at) VALUES (?, ?, ?)",
        (key, value, datetime.now(timezone.utc).isoformat()),
    )
    db.commit()
    if close_after:
        db.close()


# ── Agent Model Config ───────────────────────────────────────────────────────

AGENT_CONFIG_NAMES = [
    "kb_agent", "code_agent", "research_agent", "qa_agent", "learning_agent",
    "main_analysis_agent", "draft_response_agent", "prd_agent",
    "classification_agent", "summary_agent", "feasibility_agent",
    "batch_agent", "reply_scanner_agent", "jira_agent",
    "notification_agent", "reporting_agent",
]

_VALID_PROVIDERS = {"anthropic", "openai"}


def get_llm_config(db):
    """Return a dict with provider, api_key, model_fast, model_main, base_url from settings."""
    provider = get_setting("llm_provider", "anthropic", db=db) or "anthropic"
    api_key = get_setting("llm_api_key", "", db=db) or get_setting("anthropic_api_key", "", db=db)
    base_url = get_setting("llm_base_url", "", db=db)
    model_fast = get_setting("llm_fast_model", "", db=db)
    model_main = get_setting("llm_main_model", "", db=db)
    return {
        "provider": provider,
        "api_key": api_key,
        "base_url": base_url,
        "model_fast": model_fast,
        "model_main": model_main,
    }


def seed_agent_model_configs(db):
    """Insert default per-agent model configs if missing."""
    defaults = [
        ("kb_agent",            "anthropic", "claude-haiku-4-5-20251001", 0.0, 3000),
        ("code_agent",          "anthropic", "claude-haiku-4-5-20251001", 0.0, 3000),
        ("research_agent",      "anthropic", "claude-haiku-4-5-20251001", 0.0, 2000),
        ("qa_agent",            "anthropic", "claude-haiku-4-5-20251001", 0.0, 1500),
        ("learning_agent",      "anthropic", "claude-haiku-4-5-20251001", 0.0, 1500),
        ("main_analysis_agent", "anthropic", "claude-sonnet-4-5",         0.0, 4000),
        ("draft_response_agent","anthropic", "claude-sonnet-4-5",         0.0, 4000),
        ("prd_agent",           "anthropic", "claude-sonnet-4-5",         0.0, 4000),
        ("classification_agent","anthropic", "claude-haiku-4-5-20251001", 0.0, 1000),
        ("summary_agent",       "anthropic", "claude-haiku-4-5-20251001", 0.0, 1000),
        ("feasibility_agent",   "anthropic", "claude-haiku-4-5-20251001", 0.0, 1000),
        ("batch_agent",         "anthropic", "claude-haiku-4-5-20251001", 0.0, 2000),
        ("reply_scanner_agent", "anthropic", "claude-haiku-4-5-20251001", 0.0, 2000),
        ("jira_agent",          "anthropic", "claude-haiku-4-5-20251001", 0.0, 1000),
        ("notification_agent",  "anthropic", "claude-haiku-4-5-20251001", 0.0, 500),
        ("reporting_agent",     "anthropic", "claude-sonnet-4-5",         0.0, 3000),
    ]
    now = datetime.now(timezone.utc).isoformat()
    for name, provider, model, temp, max_tok in defaults:
        try:
            db.execute("""
                INSERT OR IGNORE INTO agent_model_config
                (agent_name, provider, model, temperature, max_tokens, enabled, updated_at)
                VALUES (?, ?, ?, ?, ?, 1, ?)
            """, (name, provider, model, temp, max_tok, now))
        except Exception:
            pass
    db.commit()


def list_agent_model_configs(db):
    """Return all rows from agent_model_config as list of dicts."""
    try:
        rows = db.execute(
            "SELECT * FROM agent_model_config ORDER BY agent_name"
        ).fetchall()
        return [dict(r) for r in rows]
    except Exception:
        return []


def get_agent_model_config(db, agent_name):
    """Return one agent config row as dict, or None."""
    try:
        row = db.execute(
            "SELECT * FROM agent_model_config WHERE agent_name = ?", (agent_name,)
        ).fetchone()
        return dict(row) if row else None
    except Exception:
        return None


def update_agent_model_config(db, agent_name, payload):
    """Validate and update an agent's model config. Raises ValueError on bad input."""
    provider = payload.get("provider", "anthropic")
    if provider not in _VALID_PROVIDERS:
        raise ValueError(f"Unknown provider '{provider}'. Must be one of: {sorted(_VALID_PROVIDERS)}")
    model = payload.get("model", "").strip()
    if not model:
        raise ValueError("model must not be empty")
    temperature = float(payload.get("temperature", 0.0))
    max_tokens = int(payload.get("max_tokens", 2000))
    enabled = int(bool(payload.get("enabled", True)))
    fallback_provider = payload.get("fallback_provider", "")
    fallback_model = payload.get("fallback_model", "")
    now = datetime.now(timezone.utc).isoformat()
    db.execute("""
        INSERT INTO agent_model_config
            (agent_name, provider, model, temperature, max_tokens, enabled,
             fallback_provider, fallback_model, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(agent_name) DO UPDATE SET
            provider=excluded.provider, model=excluded.model,
            temperature=excluded.temperature, max_tokens=excluded.max_tokens,
            enabled=excluded.enabled, fallback_provider=excluded.fallback_provider,
            fallback_model=excluded.fallback_model, updated_at=excluded.updated_at
    """, (agent_name, provider, model, temperature, max_tokens, enabled,
          fallback_provider, fallback_model, now))
    db.commit()


# ── File Upload & Text Extraction ────────────────────────────────────────────

UPLOAD_FOLDER = os.path.join(_DATA_DIR, "kb_uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {"pdf", "docx", "xlsx", "xls", "csv", "txt", "md"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def resolve_file_path(file_path):
    """Resolve a file path, handling macOS local paths in sandboxed environments."""
    if not file_path:
        return file_path
    if os.path.isfile(file_path):
        return file_path
    # Try mapping macOS paths to sandbox mount
    # e.g. /Users/x/Desktop/freshdesk-analyzer/kb_uploads/file.pdf
    #   -> /sessions/.../mnt/Desktop/freshdesk-analyzer/kb_uploads/file.pdf
    import glob as globmod
    mount_candidates = globmod.glob("/sessions/*/mnt/Desktop")
    for mount_point in mount_candidates:
        desktop_idx = file_path.lower().find("desktop")
        if desktop_idx >= 0:
            after_desktop = file_path[desktop_idx + len("desktop"):].lstrip("/\\")
            candidate = os.path.join(mount_point, after_desktop)
            if os.path.isfile(candidate):
                return candidate
    # Try just the filename in the local kb_uploads folder
    fname = os.path.basename(file_path)
    local_candidate = os.path.join(UPLOAD_FOLDER, fname)
    if os.path.isfile(local_candidate):
        return local_candidate
    return file_path  # Return original (will error, but at least logged)


def extract_text_from_file(file_path):
    """Extract text content from uploaded files (PDF, DOCX, XLSX, CSV, TXT)."""
    file_path = resolve_file_path(file_path)
    ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""

    try:
        if ext == "pdf":
            try:
                import pdfplumber
                text_parts = []
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text_parts.append(page_text)
                return "\n\n".join(text_parts) if text_parts else "(Could not extract text from PDF)"
            except ImportError:
                return "(PDF support requires pdfplumber — run: pip install pdfplumber)"

        elif ext == "docx":
            try:
                from docx import Document
                doc = Document(file_path)
                return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except ImportError:
                return "(DOCX support requires python-docx — run: pip install python-docx)"

        elif ext in ("xlsx", "xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                text_parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_parts.append(f"--- Sheet: {sheet_name} ---")
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        if any(cells):
                            text_parts.append(" | ".join(cells))
                wb.close()
                return "\n".join(text_parts) if text_parts else "(Empty spreadsheet)"
            except ImportError:
                return "(Excel support requires openpyxl — run: pip install openpyxl)"

        elif ext == "csv":
            import csv
            text_parts = []
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    text_parts.append(" | ".join(row))
            return "\n".join(text_parts[:500])  # Cap at 500 rows

        elif ext in ("txt", "md"):
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()[:50000]  # Cap at 50k chars

        else:
            return f"(Unsupported file type: .{ext})"

    except Exception as e:
        logger.error(f"Error extracting text from {file_path}: {e}")
        return f"(Error reading file: {str(e)[:200]})"


# ── Freshdesk Client ─────────────────────────────────────────────────────────

STATUS_MAP = {
    2: "Open", 3: "In Progress", 4: "Resolved", 5: "Closed",
    6: "Waiting on Customer", 7: "Waiting on Third Party",
    18: "Pending Approval", 20: "Assign to 3L",
}
PRIORITY_MAP = {1: "Low", 2: "Medium", 3: "High", 4: "Urgent"}

# ── Company / Client extraction from email ────────────────────────────────

# Known company name mappings (domain suffix → display name)
COMPANY_MAP = {
    "bdo": "BDO",
    "bakertilly": "Baker Tilly",
    "kpmg": "KPMG",
    "deloitte": "Deloitte",
    "ey": "EY",
    "pwc": "PwC",
    "mazars": "Mazars",
    "grant-thornton": "Grant Thornton",
    "grantthornton": "Grant Thornton",
    "rsm": "RSM",
    "silverfin": "Silverfin",
    "fiduciaire": "Fiduciaire",
    "fideos": "Fideos",
    "atoz": "ATOZ",
    "arendt": "Arendt",
    "sgbt": "SGBT",
    "bil": "BIL",
    "ing": "ING",
    "bnp": "BNP Paribas",
    "bnpparibas": "BNP Paribas",
    "alter-domus": "Alter Domus",
    "alterdomus": "Alter Domus",
    "intertrust": "Intertrust",
    "citco": "Citco",
    "vistra": "Vistra",
    "apex": "Apex Group",
}


def extract_company(email):
    """Extract company/client name from email address.

    E.g. john@bdo.lu → BDO, jane@bakertilly.com → Baker Tilly
    Falls back to capitalizing the domain name if not in known map.
    """
    if not email or "@" not in email:
        return "Unknown"
    domain = email.split("@")[1].lower()
    # Remove TLD parts (.lu, .com, .co.uk, etc.)
    parts = domain.split(".")
    name_part = parts[0] if parts else domain

    # Check known mappings
    if name_part in COMPANY_MAP:
        return COMPANY_MAP[name_part]

    # Also check with hyphens removed
    clean = name_part.replace("-", "")
    if clean in COMPANY_MAP:
        return COMPANY_MAP[clean]

    # Fallback: capitalize the domain name nicely
    return name_part.replace("-", " ").replace("_", " ").title()


def strip_html(text):
    """Strip HTML tags and decode entities."""
    if not text:
        return ""
    text = html_lib.unescape(text)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p>", "\n", text, flags=re.I)
    text = re.sub(r"</div>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def freshdesk_request(method, endpoint, api_key, domain, **kwargs):
    """Make Freshdesk API request."""
    url = f"https://{domain}/api/v2{endpoint}"
    auth = HTTPBasicAuth(api_key, "X")
    time.sleep(0.5)  # Rate limiting - 2 req/sec to stay well under Freshdesk's 50/min
    resp = requests.request(method, url, auth=auth, **kwargs)
    resp.raise_for_status()
    return resp.json()


def search_tickets(api_key, domain, group_id, country="", statuses="2,3,4,20"):
    """Search for tickets matching our criteria."""
    # Build status filter
    status_list = [s.strip() for s in statuses.split(",") if s.strip()]
    if status_list:
        status_clause = " OR ".join(f"status:{s}" for s in status_list)
        query = f"group_id:{group_id} AND ({status_clause})"
    else:
        query = f"group_id:{group_id}"

    # Add country filter if set
    if country:
        query += f" AND cf_country:'{country}'"

    logger.info(f"Freshdesk search query: {query}")
    all_tickets = []
    page = 1

    while True:
        result = freshdesk_request(
            "GET", "/search/tickets", api_key, domain,
            params={"query": f'"{query}"', "page": page}
        )
        results = result.get("results", [])
        if not results:
            break
        all_tickets.extend(results)
        if len(results) < 30:
            break
        page += 1

    return all_tickets


def get_ticket_details(api_key, domain, ticket_id):
    """Get full ticket details + conversations."""
    ticket = freshdesk_request(
        "GET", f"/tickets/{ticket_id}", api_key, domain,
        params={"include": "requester,stats"}
    )
    conversations = freshdesk_request(
        "GET", f"/tickets/{ticket_id}/conversations", api_key, domain
    )
    if not isinstance(conversations, list):
        conversations = conversations.get("conversations", [])
    return ticket, conversations


def compile_ticket_thread(ticket_data, conversations, domain):
    """Compile ticket + conversations into plain text."""
    tid = ticket_data.get("id", "?")
    subject = ticket_data.get("subject", "")
    status = STATUS_MAP.get(ticket_data.get("status", 0), "Unknown")
    priority = PRIORITY_MAP.get(ticket_data.get("priority", 0), "Unknown")
    group_id = ticket_data.get("group_id", "")

    requester = ticket_data.get("requester", {})
    if isinstance(requester, dict):
        req_name = requester.get("name", "Unknown")
        req_email = requester.get("email", "Unknown")
    else:
        req_name = "Unknown"
        req_email = "Unknown"

    cf = ticket_data.get("custom_fields", {}) or {}
    country = cf.get("cf_country", "")

    desc = ticket_data.get("description_text") or strip_html(ticket_data.get("description", ""))

    lines = [
        "=" * 70,
        f"FRESHDESK TICKET #{tid}",
        "=" * 70, "",
        "METADATA:",
        f"  Subject: {subject}",
        f"  Status: {status}",
        f"  Priority: {priority}",
        f"  Group ID: {group_id}",
        f"  Requester: {req_name} ({req_email})",
        f"  Country: {country}",
        f"  Created: {ticket_data.get('created_at', '')}",
        f"  Updated: {ticket_data.get('updated_at', '')}", "",
        "DESCRIPTION:", "-" * 70, desc, "",
    ]

    if conversations:
        lines += ["CONVERSATION HISTORY:", "-" * 70]
        sorted_convs = sorted(conversations, key=lambda c: c.get("created_at", ""))
        for conv in sorted_convs:
            from_name = "Agent" if conv.get("incoming", False) is False else "Customer"
            user = conv.get("user", {})
            if isinstance(user, dict):
                from_name = user.get("name", from_name)
            body = strip_html(conv.get("body", ""))
            lines += ["", f"From: {from_name}", f"At: {conv.get('created_at', '')}", "-" * 40, body]

    lines += ["", "=" * 70]
    return "\n".join(lines)


def extract_and_download_screenshots(ticket_data, conversations, api_key, domain, ticket_id):
    """Extract screenshot/image attachments from a Freshdesk ticket and its conversations.
    Downloads images to a local folder and returns a list of screenshot metadata dicts.
    Each dict: {"filename": str, "path": str, "source": str, "created_at": str, "content_type": str}
    """
    screenshots_dir = os.path.join(_DATA_DIR, "screenshots", str(ticket_id))
    os.makedirs(screenshots_dir, exist_ok=True)

    image_extensions = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff"}
    image_content_types = {"image/png", "image/jpeg", "image/jpg", "image/gif", "image/bmp", "image/webp", "image/tiff"}
    screenshots = []

    # Patterns that indicate junk images (logos, signatures, icons, tracking pixels)
    junk_name_patterns = re.compile(
        r'(logo|signature|banner|footer|header|icon|badge|avatar|'
        r'spacer|pixel|tracking|linkedin|facebook|twitter|instagram|'
        r'social[-_]?media|email[-_]?sig|vcard|phone[-_]?icon|'
        r'1x1|transparent|blank|divider|separator)',
        re.IGNORECASE
    )
    # Minimum size in bytes — images under this are likely icons/logos (5KB)
    MIN_IMAGE_SIZE = 5000

    def download_attachment(att, source_label, created_at="", is_inline=False):
        """Download a single attachment if it's an image and not junk."""
        name = att.get("name", "")
        url = att.get("attachment_url", "")
        content_type = att.get("content_type", "").lower()
        ext = os.path.splitext(name)[1].lower() if name else ""

        if not url:
            return
        if ext not in image_extensions and content_type not in image_content_types:
            return

        # Skip images whose filename matches known junk patterns
        if name and junk_name_patterns.search(name):
            logger.info(f"Skipping junk image: {name} (filename match) for ticket {ticket_id}")
            return

        # Skip images whose URL matches junk patterns (common for inline email images)
        if junk_name_patterns.search(url):
            logger.info(f"Skipping junk image URL: {url[:100]} for ticket {ticket_id}")
            return

        safe_name = secure_filename(name) or f"screenshot_{len(screenshots)}.png"
        # Avoid duplicates
        local_path = os.path.join(screenshots_dir, safe_name)
        counter = 1
        while os.path.exists(local_path):
            base, ext_part = os.path.splitext(safe_name)
            local_path = os.path.join(screenshots_dir, f"{base}_{counter}{ext_part}")
            counter += 1

        try:
            resp = requests.get(url, auth=HTTPBasicAuth(api_key, "X"), timeout=30)
            file_size = len(resp.content)
            if resp.status_code == 200 and file_size > 100:
                # Skip tiny images — likely logos, icons, or tracking pixels
                # For inline images, use a higher threshold (they're often email decorations)
                min_size = MIN_IMAGE_SIZE if is_inline else 2000
                if file_size < min_size:
                    logger.info(f"Skipping tiny image: {safe_name} ({file_size} bytes) for ticket {ticket_id}")
                    return

                # Check image dimensions if possible — skip very small images
                try:
                    from PIL import Image
                    import io
                    img = Image.open(io.BytesIO(resp.content))
                    w, h = img.size
                    # Skip images that are clearly icons/logos (under 80px in both dimensions)
                    # or extremely wide and thin (likely banners/separators)
                    if w < 80 and h < 80:
                        logger.info(f"Skipping small image: {safe_name} ({w}x{h}px) for ticket {ticket_id}")
                        return
                    if (w > 10 * h) or (h > 10 * w):
                        logger.info(f"Skipping banner/divider image: {safe_name} ({w}x{h}px) for ticket {ticket_id}")
                        return
                except ImportError:
                    pass  # PIL not installed — skip dimension check
                except Exception:
                    pass  # Corrupted image or unsupported format — still save it

                with open(local_path, "wb") as f:
                    f.write(resp.content)
                screenshots.append({
                    "filename": os.path.basename(local_path),
                    "path": local_path,
                    "source": source_label,
                    "created_at": created_at,
                    "content_type": content_type or "image/png",
                    "size": file_size,
                })
                logger.info(f"Downloaded screenshot: {safe_name} ({file_size} bytes) for ticket {ticket_id}")
        except Exception as e:
            logger.warning(f"Failed to download attachment {name} for ticket {ticket_id}: {e}")

    # Also extract inline images from HTML description
    def extract_inline_images(html_content, source_label, created_at=""):
        """Extract image URLs from HTML content (inline images in Freshdesk).
        Skips images that appear in email signatures (after common signature markers)."""
        if not html_content:
            return

        # Try to strip email signature content before extracting images.
        # Common signature markers in Freshdesk emails:
        sig_markers = [
            '<div class="signature"', '<div id="signature"',
            '-- <br', '--<br',  # RFC signature delimiter
            '<table class="signature', '<div class="email-signature',
            'Cordialement,', 'Cordialement<br', 'Kind regards,', 'Kind regards<br',
            'Best regards,', 'Best regards<br', 'Regards,', 'Regards<br',
            'Met vriendelijke groet', 'Mit freundlichen',
            'Bien à vous', 'Bien cordialement',
        ]
        # Find the earliest signature marker position
        content_for_images = html_content
        sig_pos = len(html_content)
        for marker in sig_markers:
            idx = html_content.lower().find(marker.lower())
            if idx != -1 and idx < sig_pos:
                sig_pos = idx
        if sig_pos < len(html_content):
            content_for_images = html_content[:sig_pos]
            logger.info(f"Stripped email signature at position {sig_pos}/{len(html_content)} for image extraction")

        img_pattern = re.compile(r'<img[^>]+src=["\']([^"\']+)["\']', re.IGNORECASE)
        for match in img_pattern.finditer(content_for_images):
            img_url = match.group(1)
            if not img_url.startswith("http"):
                continue
            att = {
                "name": os.path.basename(img_url.split("?")[0]) or f"inline_{len(screenshots)}.png",
                "attachment_url": img_url,
                "content_type": "image/png",
            }
            download_attachment(att, source_label, created_at, is_inline=True)

    # 1. Ticket description attachments
    for att in (ticket_data.get("attachments") or []):
        download_attachment(att, "Ticket description", ticket_data.get("created_at", ""))

    # 1b. Inline images in ticket HTML description
    extract_inline_images(
        ticket_data.get("description", ""),
        "Ticket description (inline)",
        ticket_data.get("created_at", "")
    )

    # 2. Conversation attachments
    for conv in (conversations or []):
        source = "Customer" if conv.get("incoming", True) else "Agent"
        user = conv.get("user", {})
        if isinstance(user, dict):
            source = user.get("name", source)
        created = conv.get("created_at", "")
        for att in (conv.get("attachments") or []):
            download_attachment(att, f"Conversation – {source}", created)

        # 2b. Inline images in conversation HTML
        extract_inline_images(conv.get("body", ""), f"Conversation – {source} (inline)", created)

    return screenshots


# ── Knowledge Base Helper ────────────────────────────────────────────────────

KNOWLEDGE_CATEGORIES = [
    "Annual Account Presentation Requirements",
    "Accounting Law",
    "Tax Law",
    "Commercial Law",
    "Silverfin UX",
    "Previous Freshdesk Responses",
    "Others",
]

WRITING_STYLES = {
    "customer_support": {
        "label": "Customer Support (Soft & Professional)",
        "instructions": """Writing style: Customer-facing support agent.
- Start with "Dear [Name]" or "Dear Customer"
- Be warm, empathetic, and acknowledge the customer's frustration or question
- Use simple, clear language — avoid technical jargon unless necessary
- Never promise specific timelines or dates
- Always include concrete next steps
- End with a reassuring closing (e.g. "Please don't hesitate to reach out if you have further questions")
- Keep the tone soft, patient, and professional
- Sign off with "Kind regards" or "Best regards\"""",
    },
    "product_manager": {
        "label": "Product Owner / Manager (Technical & Direct)",
        "instructions": """Writing style: Internal product owner / manager.
- Be direct, concise, and technically precise
- Use proper technical terminology (Silverfin features, accounting terms, etc.)
- Focus on root cause analysis and actionable items
- Reference specific product areas, templates, or features when relevant
- Include clear next steps and ownership (who does what)
- Tone should be professional but efficient — no fluff
- Structure the response with clear sections if needed
- Sign off professionally""",
    },
}


def get_knowledge_base_context(db, max_per_entry=6000, max_total=30000):
    """Fetch all knowledge base entries and format them for the AI prompt.
    Caps individual entries and total size to keep prompt tokens reasonable."""
    rows = db.execute(
        "SELECT category, title, content, entry_type, file_path, url FROM knowledge_base ORDER BY category, title"
    ).fetchall()
    if not rows:
        return ""

    sections = {}
    for row in rows:
        cat = row["category"]
        if cat not in sections:
            sections[cat] = []

        entry_type = row["entry_type"] or "text"

        if entry_type == "file" and row["file_path"]:
            content = extract_text_from_file(row["file_path"])
            if len(content) > max_per_entry:
                content = content[:max_per_entry] + f"\n... [truncated from {len(content)} chars — full document available in knowledge base]"
            user_note = row["content"] or ""
            if user_note:
                content = f"NOTE: {user_note}\n\n{content}"
            sections[cat].append(f"### {row['title']} (from uploaded file)\n{content}")
        elif entry_type == "url" and row["url"]:
            note = row["content"] if row["content"] else ""
            sections[cat].append(f"### {row['title']} (reference: {row['url']})\n{note}")
        else:
            content = row["content"] or ""
            if len(content) > max_per_entry:
                content = content[:max_per_entry] + "\n... [truncated]"
            sections[cat].append(f"### {row['title']}\n{content}")

    lines = ["\n\nKNOWLEDGE BASE (use this as reference — check relevant entries before responding):"]
    lines.append("=" * 60)
    total_chars = 0
    for cat, entries in sections.items():
        lines.append(f"\n## {cat}")
        lines.append("-" * 40)
        for entry in entries:
            if total_chars + len(entry) > max_total:
                remaining = max_total - total_chars
                if remaining > 200:
                    lines.append(entry[:remaining] + "\n... [knowledge base truncated for prompt size]")
                    total_chars = max_total
                break
            lines.append(entry)
            total_chars += len(entry)
        if total_chars >= max_total:
            lines.append("\n... [additional entries omitted — knowledge base exceeds prompt limit]")
            break
    lines.append("=" * 60)

    # Append Google Drive KB content if configured
    try:
        if not GOOGLE_AVAILABLE:
            raise ImportError("Google libraries not installed")
        sa_json = get_setting("google_sa_json", db=db)
        kb_folder = get_setting("google_kb_folder", db=db)
        if sa_json and kb_folder:
            remaining = max_total - total_chars
            if remaining > 500:
                drive_context = get_kb_context_from_drive(sa_json, kb_folder, max_chars=min(remaining, 10000))
                if drive_context:
                    lines.append("\n\nGOOGLE DRIVE KNOWLEDGE BASE:")
                    lines.append("=" * 60)
                    lines.append(drive_context[:remaining])
                    lines.append("=" * 60)
    except Exception as e:
        log.warning(f"Failed to load Google Drive KB context: {e}")

    return "\n".join(lines)


def get_terminology_context(db):
    """Fetch all terminology entries and format them as a structured glossary for AI prompts."""
    rows = db.execute(
        "SELECT term_fr, term_en, definition, category, ecdf_reference, usage_context FROM terminology ORDER BY category, term_fr"
    ).fetchall()

    if not rows:
        return ""

    sections = {}
    for row in rows:
        cat = row["category"] or "general"
        if cat not in sections:
            sections[cat] = []

        term_entry = f"- FR: {row['term_fr']} | EN: {row['term_en']}"
        if row["definition"]:
            term_entry += f" ({row['definition']})"
        if row["ecdf_reference"]:
            term_entry += f" [eCDF: {row['ecdf_reference']}]"
        sections[cat].append(term_entry)

    # Map internal category names to display names
    category_display = {
        "legal_entity": "Legal Entity Terms",
        "accounting": "Accounting Terms",
        "legal_document": "Legal Documents",
        "tax": "Tax Terms",
        "ecdf": "eCDF Terms",
        "general": "General Terms",
    }

    lines = ["OFFICIAL LUXEMBOURG TERMINOLOGY GLOSSARY (always use these exact terms):"]
    for cat in ["legal_entity", "accounting", "legal_document", "tax", "ecdf", "general"]:
        if cat in sections and sections[cat]:
            display_cat = category_display.get(cat, cat)
            lines.append(f"\n[{display_cat}]")
            lines.extend(sections[cat])

    return "\n".join(lines)


def resolve_template_path(code_path):
    """Resolve the template code path, handling macOS local paths in sandboxed environments."""
    if not code_path:
        return ""
    # Direct path works — use it
    if os.path.isdir(code_path):
        return code_path
    # Try to map macOS user paths to the sandboxed mount
    # e.g. /Users/someone/Desktop/MyFolder -> /sessions/.../mnt/Desktop/MyFolder
    mount_base = os.environ.get("COWORK_MOUNT", "/sessions")
    # Find any session mount point
    import glob as globmod
    mount_candidates = globmod.glob("/sessions/*/mnt/Desktop")
    if not mount_candidates:
        mount_candidates = globmod.glob("/sessions/*/mnt")
    for mount_point in mount_candidates:
        # Try matching common user path patterns
        # /Users/<user>/Desktop/<folder> or C:\Users\<user>\Desktop\<folder>
        for pattern_prefix in ["/Users/", "C:\\Users\\", "C:/Users/"]:
            if code_path.startswith(pattern_prefix):
                # Extract the part after Desktop/
                desktop_idx = code_path.lower().find("desktop")
                if desktop_idx >= 0:
                    # Get everything after "Desktop/" or "Desktop\"
                    after_desktop = code_path[desktop_idx + len("desktop"):]
                    after_desktop = after_desktop.lstrip("/\\")
                    candidate = os.path.join(mount_point.rsplit("/mnt", 1)[0], "mnt", "Desktop", after_desktop)
                    if os.path.isdir(candidate):
                        return candidate
        # Also try just the folder name at the end of the path
        folder_name = os.path.basename(code_path.rstrip("/\\"))
        for base in mount_candidates:
            candidate = os.path.join(base, folder_name)
            if os.path.isdir(candidate):
                return candidate
    return ""


def find_template_code(ticket_subject, ticket_analysis="", db=None):
    """Search the configured code path for template files matching the ticket's template name.
    Uses a comprehensive alias map and fuzzy matching to find the right template folder,
    then returns the main.liquid and key text_parts files."""
    raw_path = get_setting("template_code_path", db=db)
    code_path = resolve_template_path(raw_path)
    if not code_path:
        return ""

    search_text = ((ticket_subject or "") + " " + (ticket_analysis or "")).lower()

    # Comprehensive alias map: search terms -> folder name patterns
    # Each entry: (list of search phrases, list of folder name patterns to match)
    TEMPLATE_ALIASES = [
        # Annual Accounts notes
        (["staff cost", "frais de personnel", "personnel", "employee", "employé", "salaries", "wages", "salaire"],
         ["staff_cost", "wages", "remunerations"]),
        (["equity", "capitaux propres", "capital", "share capital", "fonds propres"],
         ["equity", "equity_movement", "shareholder_register"]),
        (["financial fixed asset", "immobilisation financ", "investments held", "financial asset", "participat"],
         ["fin_fa", "financial_fixed_assets", "investments", "overview_fin_assets"]),
        (["tangible", "immobilisation corpor", "corporel"],
         ["tangible_fa", "fixed_assets"]),
        (["intangible", "immobilisation incorpor", "incorporel"],
         ["intangible_fa", "formation_expenses"]),
        (["receivable", "créance", "creance", "debtor", "débiteur"],
         ["curr_assets", "prepayments", "advances_and_loans"]),
        (["payable", "dette", "debt", "creditor", "créancier", "split between", "becoming due"],
         ["debts", "debenture_loans", "credit_institutions", "other_debts"]),
        (["provision"],
         ["provisions"]),
        (["inventory", "inventories", "stock", "raw material", "matière première"],
         ["inventories", "inventory", "raw_materials", "variation_stocks"]),
        (["turnover", "chiffre d'affaires", "revenue"],
         ["turnover", "gross_profit"]),
        (["cash", "trésorerie", "bank"],
         ["cash", "overview_cash"]),
        (["tax note", "impôt", "note fiscale", "deferred tax"],
         ["tax", "overview_taxes"]),
        (["off balance", "hors bilan", "engagement"],
         ["off_balance_comm"]),
        (["accounting polic", "politique comptable", "going concern", "non going concern"],
         ["acc_policies"]),
        (["audit", "commissaire", "réviseur"],
         ["audit"]),
        (["subsequent event", "événement postérieur"],
         ["subsequent_events"]),
        (["related part", "partie liée"],
         ["related_parties"]),
        (["general info", "informations générales", "company information"],
         ["general_info", "company_information"]),
        (["governance", "gérant", "administrateur", "board", "management body"],
         ["governance_info", "lux_aa_governance"]),
        (["interest payable", "charges d'intérêt", "interest expense"],
         ["interest_payable"]),
        (["other operating exp", "autres charges d'exploitation"],
         ["other_operating_expenses"]),
        (["other operating inc", "autres produits d'exploitation"],
         ["other_operating_income"]),
        (["value adjust", "correction de valeur", "dépréciation"],
         ["value_adj", "value_adjustments"]),
        (["deferred income", "produits constatés d'avance", "comptes de régularisation"],
         ["deferred_income"]),
        (["work performed", "travaux effectués", "production immobilisée"],
         ["work_performed"]),
        (["ebitda", "produit d'exploitation"],
         ["gross_profit", "turnover", "other_operating"]),
        # Annual Accounts structure
        (["balance sheet", "bilan"],
         ["ecdf_balance_sheet"]),
        (["profit loss", "compte de résultat", "profit and loss", "p&l"],
         ["ecdf_profit_loss"]),
        (["cover page", "page de garde"],
         ["cover"]),
        (["management report", "rapport de gestion"],
         ["management_report"]),
        (["validation", "contrôle"],
         ["validation_rules", "overview_validations"]),
        (["settings", "paramètr", "réglages"],
         ["settings"]),
        # Corporate Tax
        (["corporate tax", "impôt sur le revenu", "irc", "icc", "business profit"],
         ["lux_ct_cit_business_profit", "lux_ct_cit", "lux_ct_form_500"]),
        (["addition", "déduction", "ajustement"],
         ["additions", "deductions", "business_profit"]),
        (["net wealth", "impôt sur la fortune", "fortune"],
         ["nwt"]),
        (["municipal business tax", "icc", "taxe commerciale"],
         ["mbt"]),
        (["investment tax credit", "crédit d'impôt investissement", "bonification"],
         ["investment_tax_credit"]),
        (["tax credit unemploy", "embauche chômeur"],
         ["tax_credit_recruiting"]),
        (["specific fiscal", "disposition fiscale"],
         ["specific_fiscal"]),
        (["controlled foreign", "société étrangère contrôlée"],
         ["controlled_foreign"]),
        (["tax consolidat", "intégration fiscal"],
         ["tax_consolidation"]),
        (["form 500", "formulaire 500"],
         ["form_500"]),
        (["form 506", "formulaire 506"],
         ["form_506"]),
        # Legal Documents
        (["board minute", "pv du conseil", "minutes of the board", "procès-verbal"],
         ["legal_minutes_board"]),
        (["general meeting", "assemblée générale", "ag ", "ago"],
         ["legal_general_meeting"]),
        (["written resolution", "résolution circulaire", "voie circulaire"],
         ["legal_written_resolutions"]),
        (["power of attorney", "procuration", "proxy"],
         ["legal_power_of_attorney", "proxy"]),
        (["convening", "convocation"],
         ["legal_convening"]),
        (["result allocation", "affectation du résultat", "profit allocation"],
         ["legal_result_allocation", "legal_final_result"]),
        (["dividend", "dividende", "interim div"],
         ["legal_interim_div", "legal_result_allocation"]),
        (["special report", "rapport spécial"],
         ["legal_special_report"]),
        (["legal setting", "legal doc setting"],
         ["legal_settings"]),
    ]

    # Find matching template folders
    matched_folders = set()
    for search_phrases, folder_patterns in TEMPLATE_ALIASES:
        if any(phrase in search_text for phrase in search_phrases):
            for pattern in folder_patterns:
                matched_folders.add(pattern)

    # Also extract meaningful words from subject for direct folder name matching
    stop_words = {"silverfin", "urgent", "please", "note", "template", "issue", "problem",
                  "comptes", "annuels", "annual", "accounts", "the", "for", "and", "des",
                  "les", "une", "dans", "avec", "this", "that", "from", "fwd"}
    subject_words = []
    for word in (ticket_subject or "").split():
        clean = word.strip(":-–()[].,;\"'#/").lower()
        if len(clean) > 3 and clean not in stop_words:
            subject_words.append(clean)

    # Walk the code directory and score template folders
    matched_files = []
    try:
        for root, dirs, files in os.walk(code_path):
            dirs[:] = [d for d in dirs if not d.startswith('.') and d not in ('node_modules', '__pycache__', '.git')]
            for fname in files:
                if not fname.endswith(('.liquid', '.yml', '.yaml')):
                    continue
                fpath = os.path.join(root, fname)
                rel_path = os.path.relpath(fpath, code_path).lower()

                score = 0
                # Score from alias matching
                for pattern in matched_folders:
                    if pattern in rel_path:
                        score += 10
                        # Bonus for main.liquid files (most important)
                        if fname == "main.liquid":
                            score += 5
                        break

                # Score from direct word matching against path
                for word in subject_words:
                    if word in rel_path:
                        score += 2

                if score > 0:
                    matched_files.append((score, os.path.relpath(fpath, code_path), fpath))

    except Exception as e:
        logger.warning(f"Error scanning template code path: {e}")
        return ""

    if not matched_files:
        return ""

    # Sort by score, deduplicate by template folder, take top 3 folders
    matched_files.sort(key=lambda x: -x[0])
    seen_folders = set()
    top_files = []
    for score, rel_path, fpath in matched_files:
        # Get the template folder (2nd level, e.g. reconciliation_texts/lux_aa_an_staff_cost)
        parts = rel_path.split(os.sep)
        folder_key = os.sep.join(parts[:3]) if len(parts) > 2 else parts[0]
        if folder_key in seen_folders:
            continue
        seen_folders.add(folder_key)
        # For this folder, find main.liquid first, then key text_parts
        folder_path = os.path.dirname(fpath)
        # Go up to the template root folder
        template_root = fpath
        for p in parts[:-1]:
            if p in ("text_parts", "tests"):
                template_root = os.path.dirname(template_root)
        # Now gather important files from this template folder
        template_dir = os.path.dirname(fpath)
        while os.path.basename(template_dir) in ("text_parts", "tests"):
            template_dir = os.path.dirname(template_dir)

        main_liquid = os.path.join(template_dir, "main.liquid")
        if os.path.isfile(main_liquid):
            top_files.append((score + 5, os.path.relpath(main_liquid, code_path), main_liquid))
        # Also add text_parts .liquid files (they contain the actual display logic)
        text_parts_dir = os.path.join(template_dir, "text_parts")
        if os.path.isdir(text_parts_dir):
            for tp_file in sorted(os.listdir(text_parts_dir)):
                if tp_file.endswith(".liquid"):
                    tp_path = os.path.join(text_parts_dir, tp_file)
                    top_files.append((score, os.path.relpath(tp_path, code_path), tp_path))

        if len(seen_folders) >= 2:  # Top 2 template folders
            break

    # Sort by score (main.liquid files first), deduplicate
    top_files.sort(key=lambda x: -x[0])
    seen_paths = set()
    unique_files = []
    for score, rel_path, fpath in top_files:
        if fpath not in seen_paths:
            seen_paths.add(fpath)
            unique_files.append((score, rel_path, fpath))
    top_files = unique_files

    code_parts = []
    total_chars = 0
    max_chars = 12000  # Cap total code context

    for score, rel_path, fpath in top_files:
        try:
            with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            # Truncate individual files
            if len(content) > 4000:
                content = content[:4000] + "\n... [truncated]"
            if total_chars + len(content) > max_chars:
                remaining = max_chars - total_chars
                if remaining > 500:
                    content = content[:remaining] + "\n... [truncated]"
                else:
                    break
            code_parts.append(f"── File: {rel_path} ──\n{content}")
            total_chars += len(content)
        except Exception as e:
            logger.warning(f"Could not read template file {fpath}: {e}")

    if not code_parts:
        return ""

    return "ACTUAL TEMPLATE CODE (from repository):\n" + "\n\n".join(code_parts)


# ── AI Analyzer ──────────────────────────────────────────────────────────────

def call_anthropic_with_retry(client, max_retries=3, **kwargs):
    """Call Anthropic API with exponential backoff on rate limit errors (429)."""
    for attempt in range(max_retries + 1):
        try:
            resp = client.messages.create(**kwargs)
            return resp
        except Exception as e:
            error_str = str(e)
            is_rate_limit = "429" in error_str or "rate_limit" in error_str.lower()

            if is_rate_limit and attempt < max_retries:
                # Exponential backoff: 15s, 30s, 60s
                wait = 15 * (2 ** attempt)
                logger.warning(f"Rate limited (attempt {attempt+1}/{max_retries+1}), waiting {wait}s...")
                time.sleep(wait)
            else:
                raise  # Not a rate limit error, or out of retries


def truncate_text(text, max_chars=2000):
    """Truncate text to fit within token budget. ~4 chars per token."""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n\n[... truncated for brevity ...]"


def strip_code_from_output(text):
    """Post-processing: aggressively remove code patterns that leak into AI output.
    This is a safety net — the prompts should prevent this, but models frequently slip.
    Applied to ALL AI outputs before saving to DB or returning to user."""
    import re

    if not text:
        return text

    # ── 1. Remove code blocks and inline code ────────────────────────────────
    text = re.sub(r'```[\s\S]*?```', '', text)         # fenced code blocks
    text = re.sub(r'`[^`]+`', '', text)                 # inline backticks

    # ── 2. Remove Liquid template syntax ─────────────────────────────────────
    text = re.sub(r'\{%-?\s*.*?\s*-?%\}', '', text)    # {% ... %} and {%- ... -%}
    text = re.sub(r'\{\{-?\s*.*?\s*-?\}\}', '', text)  # {{ ... }} and {{- ... -}}
    text = re.sub(r'\{%.*?%\}', '', text)               # fallback for multiline
    text = re.sub(r'\{\{.*?\}\}', '', text)             # fallback for multiline

    # ── 3. Remove file paths and template structure references ───────────────
    # Lines containing Silverfin folder structures
    text = re.sub(r'^.*(?:reconciliation_texts|text_parts|account_templates|export_files|shared_parts|working_papers)/[^\s]+.*$',
                  '', text, flags=re.MULTILINE)
    # Any path with forward slashes that looks like code structure
    text = re.sub(r'^.*(?:lu_market-main|lux_[a-z])/[^\s]+.*$', '', text, flags=re.MULTILINE)
    # "Fichiers concernés" / "Files affected" lines
    text = re.sub(r'^(?:Fichiers?\s+concern[ée]s?|Files?\s+affected|Affected\s+files?|Code\s+files?|Template\s+files?)\s*[:：].*$',
                  '', text, flags=re.MULTILINE | re.IGNORECASE)

    # ── 4. Remove template technical IDs ─────────────────────────────────────
    text = re.sub(r'\blux_[a-z]{2}_[a-z_]+(?:/[^\s]*)?\b', '', text)  # lux_aa_an_staff_cost and paths
    text = re.sub(r'\blux_[a-z]{2}_[a-z]+(?:/[^\s]*)?\b', '', text) # lux_ci_general and paths
    text = re.sub(r'\blu_market-main(?:/[^\s]*)?\b', '', text)       # lu_market-main and paths
    text = re.sub(r'\blu_[a-z][a-z_-]+(?:/[^\s]+)?\b', '', text)    # lu_xxx paths

    # ── 5. Remove .liquid and other code file references ─────────────────────
    text = re.sub(r'\b\w+\.liquid\b', '', text)
    text = re.sub(r'\b\w+\.(?:rb|erb|yml|yaml|haml)\b', '', text)

    # ── 6. Remove Silverfin Liquid variable patterns ─────────────────────────
    # Dotted accessor chains: company.custom.xxx, period.year_end_date, account.xxx
    text = re.sub(r'\b(?:company|period|custom|account|result|adjustments?|people|rollforward|filler)\.[a-z_.]+\b', '', text)
    # Silverfin-specific namespace patterns
    text = re.sub(r'\b(?:local_var|local_variable|include_variable)\b', '', text)
    # Drop.xxx patterns (Silverfin Liquid drops)
    text = re.sub(r'\b(?:accounts|periods?|companies|results?|customs?)\.[a-z_]+(?:\.[a-z_]+)*\b', '', text)

    # ── 7. Remove snake_case code variable patterns ──────────────────────────
    # Known code prefixes (won't match normal French words)
    text = re.sub(r'\b(?:employees?_[a-z_]{2,}|hide_[a-z_]{2,}|show_[a-z_]{2,}|total_[a-z_]{2,}|amount_[a-z_]{2,})\b', '', text)
    text = re.sub(r'\b(?:is_[a-z_]{2,}|has_[a-z_]{2,}|get_[a-z_]{2,}|set_[a-z_]{2,}|num_[a-z_]{2,}|var_[a-z_]{2,})\b', '', text)
    text = re.sub(r'\b(?:enable_[a-z_]{2,}|disable_[a-z_]{2,}|display_[a-z_]{2,}|include_[a-z_]{2,})\b', '', text)
    text = re.sub(r'\b(?:current_[a-z_]{2,}|previous_[a-z_]{2,}|default_[a-z_]{2,}|custom_[a-z_]{2,})\b', '', text)
    # Translation key patterns (t_partnership_interests, t_share_capital, t_xxx_yyy)
    text = re.sub(r'\bt_[a-z_]{2,}\b', '', text)
    # Config/code variables with known suffixes (_code, _type, _name, _id, _value, _date, _key)
    text = re.sub(r'\b[a-z]+_(?:code|type|name|id|value|date|key|flag|count|index|mode|status|path)\b', '', text)
    # Silverfin-specific variable patterns (subscribed_capital_xxx, legal_form_xxx)
    text = re.sub(r'\b(?:subscribed|legal|nominal|registered|authorized|issued)_[a-z_]+\b', '', text)
    # Variables with 2+ underscores that look like code (e.g. staff_cost_breakdown, share_capital_value)
    text = re.sub(r'\b[a-z]+(?:_[a-z]+){2,}\b', '', text)

    # ── 8. Remove code logic expressions that leak ───────────────────────────
    # "if variable_name == value" / "unless condition" / "for item in collection"
    text = re.sub(r'\b(?:if|unless|elsif|endfor|endif|endunless|endcase|when|capture|endcapture)\s+[a-z_]+', '', text)
    # == != > < comparisons with variable names
    text = re.sub(r'\b[a-z_]+\s*(?:==|!=|>=|<=|<>)\s*(?:\d+|"[^"]*"|true|false|nil|blank|empty)\b', '', text)

    # ── 9. Remove "INTRO SECTION", "BREAKDOWN TABLE" code section markers ───
    text = re.sub(r'\b(?:INTRO\s+SECTION|BREAKDOWN\s+TABLE|SECTION\s+\d+[A-Z]?)\b', '', text)

    # ── 10. Remove markdown formatting ───────────────────────────────────────
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)  # **bold** → bold
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)  # # Header → Header

    # ── 11. Clean up artifacts ───────────────────────────────────────────────
    # Multiple spaces left by removals
    text = re.sub(r'  +', ' ', text)
    # Lines that became empty or whitespace-only after stripping
    text = re.sub(r'^\s+$', '', text, flags=re.MULTILINE)
    # Multiple blank lines
    text = re.sub(r'\n{3,}', '\n\n', text)

    return text.strip()


def load_screenshots_for_ai(ticket_id_or_screenshots, max_images=5, max_size_bytes=4_000_000):
    """Load ticket screenshots as base64 content blocks for the Anthropic vision API.
    Accepts either a ticket_id (int) to load from DB, or a list of screenshot dicts.
    Returns a list of content blocks: [{"type": "image", "source": {...}}, {"type": "text", "text": "..."}]
    """
    import base64

    screenshots = []
    if isinstance(ticket_id_or_screenshots, list):
        screenshots = ticket_id_or_screenshots
    elif isinstance(ticket_id_or_screenshots, (int, str)):
        try:
            try:
                db = get_db()
            except RuntimeError:
                # Outside Flask request context (e.g. background job) — use standalone connection
                db = get_db_standalone()
            row = db.execute("SELECT screenshots_json FROM tickets WHERE ticket_id = ?", (int(ticket_id_or_screenshots),)).fetchone()
            if row and row["screenshots_json"]:
                screenshots = json.loads(row["screenshots_json"] or "[]")
        except Exception:
            screenshots = []

    if not screenshots:
        return []

    # Filter out likely junk images before sending to AI
    _junk_re = re.compile(
        r'(logo|signature|banner|footer|header|icon|badge|avatar|'
        r'spacer|pixel|tracking|linkedin|facebook|twitter|instagram|'
        r'social[-_]?media|email[-_]?sig|vcard|phone[-_]?icon|'
        r'1x1|transparent|blank|divider|separator)',
        re.IGNORECASE
    )

    content_blocks = []
    loaded = 0
    for ss in screenshots:
        if loaded >= max_images:
            break
        fpath = ss.get("path", "")
        fname = ss.get("filename", "")
        if not fpath or not os.path.exists(fpath):
            continue
        # Skip images with junk filenames
        if fname and _junk_re.search(fname):
            continue
        try:
            file_size = os.path.getsize(fpath)
            if file_size > max_size_bytes or file_size < 5000:
                continue

            with open(fpath, "rb") as f:
                img_data = f.read()

            b64 = base64.standard_b64encode(img_data).decode("utf-8")

            # Determine media type from actual file bytes (not extension — Freshdesk
            # can serve JPEGs with .png extensions, causing Anthropic API errors)
            def _detect_media_type(data, path, fallback):
                """Detect image format from magic bytes."""
                if data[:8] == b'\x89PNG\r\n\x1a\n':
                    return "image/png"
                if data[:3] == b'\xff\xd8\xff':
                    return "image/jpeg"
                if data[:4] == b'GIF8':
                    return "image/gif"
                if data[:4] == b'RIFF' and data[8:12] == b'WEBP':
                    return "image/webp"
                # Fall back to extension if magic bytes don't match
                ext = os.path.splitext(path)[1].lower()
                ext_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                           ".gif": "image/gif", ".webp": "image/webp"}
                return ext_map.get(ext, fallback)

            media_type = _detect_media_type(img_data, fpath, ss.get("content_type", "image/png"))

            # Add label text block before the image
            source_label = ss.get("source", "Screenshot")
            filename = ss.get("filename", os.path.basename(fpath))
            content_blocks.append({
                "type": "text",
                "text": f"[Screenshot {loaded+1}: {source_label} — {filename}]"
            })
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": b64,
                }
            })
            loaded += 1
        except Exception as e:
            logger.warning(f"Failed to load screenshot {fpath} for AI: {e}")

    if content_blocks:
        content_blocks.append({
            "type": "text",
            "text": f"[{loaded} screenshot(s) from the Freshdesk ticket are shown above. Reference them in your analysis where relevant, e.g. 'as shown in the client\'s screenshot' or 'the highlighted section in the screenshot shows...']"
        })

    return content_blocks


def search_jira_for_ticket(ticket_subject, template_name="", db=None):
    """Search Jira for issues related to a Freshdesk ticket.
    Uses the Jira search settings from the database.
    Returns a text summary of related Jira issues, or empty string if Jira is not configured.
    """
    if not db:
        return ""

    jira_domain = get_setting("jira_domain", "", db=db)
    jira_project = get_setting("jira_project", "", db=db)

    if not jira_domain:
        return ""

    try:
        import requests as req  # noqa: may not be installed

        jira_email = get_setting("jira_email", "", db=db)
        jira_token = get_setting("jira_api_token", "", db=db)

        if not jira_email or not jira_token:
            return ""

        # Build JQL search: look for related issues by template name or keywords
        # Escape quotes to prevent JQL injection
        def _jql_escape(s):
            return s.replace('\\', '\\\\').replace('"', '\\"')

        search_terms = []
        if template_name:
            search_terms.append(f'text ~ "{_jql_escape(template_name)}"')

        # Extract meaningful keywords from subject
        stop_words = {"the", "a", "an", "in", "on", "at", "for", "to", "of", "and", "or", "is",
                      "le", "la", "les", "de", "du", "des", "en", "un", "une", "et", "ou",
                      "silverfin", "template", "issue", "bug", "error", "problem", "request", "-"}
        keywords = [w for w in ticket_subject.split() if len(w) > 3 and w.lower() not in stop_words]
        if keywords:
            kw_query = " OR ".join(f'text ~ "{_jql_escape(kw)}"' for kw in keywords[:3])
            search_terms.append(f"({kw_query})")

        if not search_terms:
            return ""

        jql = " OR ".join(search_terms)
        if jira_project:
            jql = f"project = {jira_project} AND ({jql})"
        jql += " ORDER BY updated DESC"

        url = f"https://{jira_domain}/rest/api/3/search"
        resp = req.get(
            url,
            params={"jql": jql, "maxResults": 5, "fields": "summary,status,priority,assignee,updated"},
            auth=(jira_email, jira_token),
            timeout=10,
        )

        if resp.status_code != 200:
            logger.warning(f"Jira search failed: {resp.status_code}")
            return ""

        data = resp.json()
        issues = data.get("issues", [])
        if not issues:
            return ""

        context_parts = []
        for issue in issues:
            fields = issue.get("fields", {})
            key = issue.get("key", "?")
            summary = fields.get("summary", "")
            status = (fields.get("status") or {}).get("name", "?")
            priority = (fields.get("priority") or {}).get("name", "?")
            assignee = (fields.get("assignee") or {}).get("displayName", "Unassigned")
            updated = (fields.get("updated") or "")[:10]
            context_parts.append(
                f"- {key}: {summary} [Status: {status}, Priority: {priority}, Assignee: {assignee}, Updated: {updated}]"
            )

        return "\n".join(context_parts)

    except Exception as e:
        logger.warning(f"Jira search error: {e}")
        return ""


# ── Jira Ticket Creation ──────────────────────────────────────────────────────

# Classification → Jira issue type mapping
JIRA_ISSUE_TYPE_MAP = {
    "bug": "Bug",
    "feature_request": "Feature request",
    "enhancement": "Feature request",
    "how_to": "Customer request",
    "sync": "Bug",
    "data": "Bug",
    "other": "Customer request",
}

# Priority mapping (internal → Jira priority name)
JIRA_PRIORITY_MAP = {
    "low": "Low Priority",
    "normal": "Normal Prio",
    "medium": "Normal Prio",
    "high": "High Priority",
    "critical": "High Priority",
    "urgent": "High Priority",
    "1": "Low Priority",
    "2": "Normal Prio",
    "3": "High Priority",
    "4": "High Priority",
}


def _jira_request(method, path, jira_domain, jira_email, jira_token, json_data=None, params=None):
    """Make an authenticated request to the Jira REST API."""
    url = f"https://{jira_domain}/rest/api/3{path}"
    resp = requests.request(
        method, url,
        auth=(jira_email, jira_token),
        json=json_data,
        params=params,
        headers={"Accept": "application/json", "Content-Type": "application/json"},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json() if resp.text else {}


def _jira_search_jql(jql, fields, jira_domain, jira_email, jira_token, max_results=50):
    """Search Jira issues using JQL — tries new /search/jql endpoint first, falls back to /search.

    Atlassian deprecated /rest/api/3/search (returns 410 Gone) on newer Jira Cloud instances.
    The replacement is POST /rest/api/3/search/jql with a JSON body.
    """
    # Try new endpoint first (POST /search/jql with JSON body)
    try:
        body = {"jql": jql, "fields": fields if isinstance(fields, list) else fields.split(","), "maxResults": max_results}
        result = _jira_request("POST", "/search/jql", jira_domain, jira_email, jira_token, json_data=body)
        return result
    except Exception as e1:
        logger.debug(f"New /search/jql failed ({e1}), trying legacy /search...")

    # Fallback to legacy GET /search
    params = {"jql": jql, "fields": fields if isinstance(fields, str) else ",".join(fields), "maxResults": max_results}
    return _jira_request("GET", "/search", jira_domain, jira_email, jira_token, params=params)


def _build_jira_description(ticket, freshdesk_url=""):
    """Build a structured Jira issue description from the ticket data."""
    parts = []

    # Freshdesk link
    if freshdesk_url or ticket.get("ticket_url"):
        fd_url = freshdesk_url or ticket.get("ticket_url", "")
        parts.append(f"Freshdesk Ticket: {fd_url}\n")

    # Client info
    req_name = ticket.get("requester_name", "")
    req_email = ticket.get("requester_email", "")
    if req_name or req_email:
        parts.append(f"Client: {req_name} ({req_email})")

    # Summary from AI analysis
    summary = ticket.get("summary", "")
    if summary:
        parts.append(f"\n## Summary\n{summary}")

    # Analysis
    analysis = ticket.get("analysis", "")
    if analysis:
        parts.append(f"\n## Analysis\n{analysis}")

    # Extract backlog ticket section from draft if available
    draft = ticket.get("draft_response", "") or ""
    backlog_section = ""
    if "--- BACKLOG TICKET ---" in draft:
        backlog_section = draft.split("--- BACKLOG TICKET ---", 1)[1].strip()
        # Remove any trailing section headers
        for marker in ["--- CLIENT RESPONSE ---", "--- INTERNAL NOTE"]:
            if marker in backlog_section:
                backlog_section = backlog_section.split(marker, 1)[0].strip()

    if backlog_section:
        parts.append(f"\n## Backlog Details\n{backlog_section}")

    # RICE score
    rice_score = ticket.get("rice_score", 0)
    if rice_score:
        parts.append(f"\n## RICE Score: {rice_score}")
        parts.append(f"Reach: {ticket.get('rice_reach', 0)} | Impact: {ticket.get('rice_impact', 0)} | "
                     f"Confidence: {ticket.get('rice_confidence', 0)} | Effort: {ticket.get('rice_effort', 0)}")

    # Risk level
    risk = ticket.get("risk_level", "")
    if risk:
        parts.append(f"\nRisk Level: {risk}")

    return "\n".join(parts)


def create_jira_ticket_from_freshdesk(ticket, db=None, overrides=None):
    """Create a Jira ticket from a Freshdesk ticket analysis.

    overrides (dict, optional): User-selected Jira field overrides:
        - issue_type (str): Jira issue type name (e.g. "Bug", "Story", "Task")
        - priority (str): Jira priority name (e.g. "High Priority")
        - epic_key (str): Epic key to link to (e.g. "LUX-42")
        - components (list[str]): Component names
        - labels (list[str]): Additional labels
        - summary (str): Custom summary override

    Returns (jira_key, jira_url) or raises an exception."""
    if not db:
        db = get_db()
    if not overrides:
        overrides = {}

    jira_domain = get_setting("jira_domain", "", db=db)
    jira_email = get_setting("jira_email", "", db=db)
    jira_token = get_setting("jira_api_token", "", db=db)
    jira_project = get_setting("jira_project", "LUX", db=db)

    if not jira_domain or not jira_email or not jira_token:
        raise ValueError("Jira is not configured. Go to Settings → Jira Integration.")

    # Issue type: user override → auto-map from classification
    if overrides.get("issue_type"):
        issue_type = overrides["issue_type"]
    else:
        classification = (ticket.get("classification") or "other").lower().strip()
        issue_type = JIRA_ISSUE_TYPE_MAP.get(classification, "Customer request")

    # Priority: user override → auto-map from risk
    if overrides.get("priority"):
        jira_priority = overrides["priority"]
    else:
        risk = (ticket.get("risk_level") or "medium").lower().strip()
        priority_str = (str(ticket.get("priority") or "")).lower().strip()
        jira_priority = JIRA_PRIORITY_MAP.get(risk, JIRA_PRIORITY_MAP.get(priority_str, "Normal Prio"))

    # Summary: user override or auto-build
    fd_id = ticket.get("ticket_id", "?")
    subject = ticket.get("subject", "No subject")
    if overrides.get("summary"):
        jira_summary = overrides["summary"]
    else:
        jira_summary = f"[FD#{fd_id}] {subject}"
    if len(jira_summary) > 255:
        jira_summary = jira_summary[:252] + "..."

    # Build description
    description = _build_jira_description(ticket)

    # Labels: auto-generated + user additions
    labels = []
    template_name = (ticket.get("template_name") or "").strip()
    workflow_name = (ticket.get("workflow_name") or "").strip()
    if template_name:
        label = template_name.replace(" ", "_").replace("/", "_")[:50]
        labels.append(label)
    if workflow_name:
        label = workflow_name.replace(" ", "_").replace("/", "_")[:50]
        labels.append(label)
    labels.append(f"freshdesk_{fd_id}")
    # Add user-specified labels
    for extra_label in (overrides.get("labels") or []):
        if extra_label and extra_label not in labels:
            labels.append(extra_label.replace(" ", "_")[:50])

    # Build the fields dict
    fields = {
        "project": {"key": jira_project},
        "issuetype": {"name": issue_type},
        "summary": jira_summary,
        "description": {
            "type": "doc",
            "version": 1,
            "content": [
                {
                    "type": "paragraph",
                    "content": [{"type": "text", "text": line}]
                }
                for line in description.split("\n") if line.strip()
            ]
        },
        "priority": {"name": jira_priority},
        "labels": labels,
    }

    # Components (user-selected)
    if overrides.get("components"):
        fields["components"] = [{"name": c} for c in overrides["components"] if c]

    # Parent ticket — set directly on the issue if provided (for sub-tasks)
    parent_key = (overrides.get("parent_key") or "").strip()
    epic_key = (overrides.get("epic_key") or "").strip()

    # Decide which parent to set at creation time:
    # 1. If explicit parent_key is given (sub-task → Story/Task), use it as parent
    # 2. If epic_key is given (Story/Task → Epic), try setting epic as parent
    # 3. Don't set both — parent_key takes precedence
    if parent_key:
        fields["parent"] = {"key": parent_key}
    elif epic_key:
        # Try to set epic as parent at creation (works for team-managed projects)
        fields["parent"] = {"key": epic_key}

    # Create the issue
    issue_data = _jira_request(
        "POST", "/issue",
        jira_domain, jira_email, jira_token,
        json_data={"fields": fields}
    )

    # If epic was requested but the parent field approach might have failed
    # (company-managed projects use a different mechanism), try issue link as fallback
    if epic_key and not parent_key:
        # Check if the parent was actually set by trying to read the created issue
        try:
            created = _jira_request("GET", f"/issue/{issue_data.get('key', '')}?fields=parent",
                                     jira_domain, jira_email, jira_token)
            has_parent = bool(created.get("fields", {}).get("parent"))
        except Exception:
            has_parent = False

        if not has_parent:
            # Fall back to creating an issue link for company-managed projects
            try:
                _jira_request(
                    "POST", "/issueLink",
                    jira_domain, jira_email, jira_token,
                    json_data={
                        "type": {"name": "Epic-Story Link"},
                        "inwardIssue": {"key": epic_key},
                        "outwardIssue": {"key": issue_data.get("key", "")},
                    }
                )
            except Exception as link_err:
                logger.warning(f"Could not link {issue_data.get('key', '')} to epic {epic_key}: {link_err}")

    jira_key = issue_data.get("key", "")
    jira_url = f"https://{jira_domain}/browse/{jira_key}"

    # Add Freshdesk ticket as a remote link on the Jira issue
    fd_url = ticket.get("ticket_url") or f"https://silverfin.freshdesk.com/a/tickets/{fd_id}"
    try:
        _jira_request(
            "POST", f"/issue/{jira_key}/remotelink",
            jira_domain, jira_email, jira_token,
            json_data={
                "globalId": f"freshdesk={fd_id}",
                "application": {
                    "type": "com.freshdesk",
                    "name": "Freshdesk",
                },
                "object": {
                    "url": fd_url,
                    "title": f"Freshdesk #{fd_id}: {subject[:100]}",
                    "icon": {
                        "url16x16": "https://freshdesk.com/favicon.ico",
                        "title": "Freshdesk"
                    },
                }
            }
        )
    except Exception as link_err:
        logger.warning(f"Could not add Freshdesk remote link to {jira_key}: {link_err}")

    # Store Jira key + URL in the DB
    db.execute(
        "UPDATE tickets SET jira_ticket_key = ?, jira_ticket_url = ? WHERE ticket_id = ?",
        (jira_key, jira_url, fd_id)
    )
    db.commit()

    logger.info(f"Created Jira {jira_key} for Freshdesk #{fd_id} (type={issue_type}, priority={jira_priority})")
    return jira_key, jira_url


def analyze_and_draft_ai(compiled_thread, anthropic_key, writing_style="customer_support", kb_context="", project_instructions="", terminology_context="", code_context="", client_context=""):
    """Analyze a ticket: classify, score RICE, and produce a detailed analysis summary.
    Draft responses are generated SEPARATELY after PO approval — not in this call."""
    client = Anthropic(api_key=anthropic_key)

    project_ctx = ""
    if project_instructions:
        project_ctx = f"\nProject context & instructions:\n{truncate_text(project_instructions, 2000)}\n"

    term_ctx = ""
    if terminology_context:
        term_ctx = f"\n{terminology_context}\n"

    code_ctx = ""
    if code_context:
        is_agent_brief = code_context.startswith("CODE AGENT BRIEF") or "TEMPLATE OVERVIEW" in code_context[:200]
        if is_agent_brief:
            code_ctx = f"""

TEMPLATE ANALYSIS (functional description from Code Agent — use for RICE Effort scoring):
{truncate_text(code_context, 4000)}
This is already in plain language. Do NOT add code references to your analysis output.
"""
        else:
            code_ctx = f"""

TEMPLATE CODE REFERENCE (use INTERNALLY for RICE Effort scoring — NEVER reference in output):
Use this to assess fix complexity, verify client claims, and check bug vs working-as-designed.
DO NOT write variable names, file paths, template IDs, or code logic in your analysis.
{truncate_text(code_context, 4000)}
"""

    client_ctx_block = ""
    if client_context:
        client_ctx_block = f"\n   CLIENT DATA (use ONLY for RICE scoring — do NOT include these numbers in your analysis or summary):\n   {client_context}\n"

    system = f"""You are a senior product analyst for Silverfin's Luxembourg templates team (BSO LUX).
You think like a PO who has been managing these templates for years. You know the code, the accounting, the law, and the clients. You've seen every type of ticket before. Your analysis is sharp, practical, and never robotic.

You receive tickets from accounting firms (BDO, EY, KPMG, Deloitte, JTC, etc.) using Silverfin for Luxembourg financial reporting.
{project_ctx}{term_ctx}{code_ctx}

SILVERFIN DOMAIN:
- Cloud accounting platform. Luxembourg: annual accounts, tax returns, legal documents.
- Annual accounts: balance sheet, P&L, notes (annexes), eCDF filing.
- Legal docs: board minutes (PV du Conseil d'Administration / Conseil de Gérance), AG/AGO, profit allocation, written resolutions, proxy, convening notices.
- Tax: IRC/ICC, ICCo, IF, forms 500/506A/506B.
- Workflows: "Luxembourg Annual Accounts", "Luxembourg Corporate Tax", "Legal Documents (Legal Docs)".
- Template types: reconciliation text (RT), working paper (WP), account template, export template.
- Common notes: Financial assets, Receivables, Payables, Staff costs, Provisions, Equity, Stocks, Tangible/Intangible fixed assets, Debtors, Creditors.
- Chart of accounts (PCN): class 1 = Equity/Liabilities, class 2 = Fixed assets, class 4 = Receivables/Payables, class 6 = Expenses, class 7 = Income.
- Templates support FR/EN/DE, multiple periods (N, N-1, N-2), visibility conditions, linked templates.
- Law refs: Loi modifiée du 19 décembre 2002, RGD 18/12/2015, PCN, eCDF (RCSL/CNC), Code de Commerce.

IMPORTANT — Templates have FUNCTIONAL INTENT, not just code. Many tables, conditions, and logic are implemented ON PURPOSE:
- Dividend distribution rows, equity movement tables, depreciation schedules — these follow specific accounting logic and legal requirements.
- Dropdown restrictions, mandatory fields, conditional visibility — these PREVENT errors. They exist for a reason.
- When a client reports something as a "bug", it may actually be the template working as designed. The client may be misusing it.
- Example: a dividend distribution row that behaves differently from what the client expects — that's because the template enforces the legal flow (profit allocation → reserves → dividends). The client may not understand this flow.

YOUR THINKING PROCESS — Follow this order for EVERY ticket:

STEP 1: WHAT IS THE CLIENT ACTUALLY SAYING?
Read past the urgency ("urgent", "critical", "ASAP" — ignore these) and past imprecise language. Identify the REAL issue: what template, what field/section, what did they expect vs what happened? Sometimes the client describes a symptom, not the root cause.

STEP 2: CHECK WHAT ALREADY EXISTS IN THE TEMPLATE.
THIS IS THE MOST IMPORTANT STEP. Many clients simply don't know how to use Silverfin.
- If template code or functional analysis is provided above, READ IT FIRST.
- Look for: visibility conditions, toggles, dropdowns, settings that control what the client is asking about.
- Look for: existing calculations, sections, or fields that already do what they want.
- Common pattern: client says "X is missing" → X is behind a setting or dropdown they haven't activated. This is a HOW-TO.
- Common pattern: client says "Y is wrong" → Y depends on account entries they haven't made. This is a HOW-TO.
- Common pattern: client reports a "bug" → the template is working as designed (e.g. dividend distribution, depreciation logic, equity movement). The client misunderstands the accounting flow. This is a HOW-TO.
If the solution already exists, STOP HERE. Classify as how_to and explain where to find it.

STEP 3: IS IT A QUICK FIX?
Typos, wrong translations, cropped text, font issues, table alignment, missing labels — these are straightforward. Don't over-analyse them. Classify as bug, give low effort (1-2), write a short analysis. No need for legal checks or elaborate reasoning.

STEP 4: IF IT'S A REAL ISSUE — CLASSIFY CORRECTLY.
- BUG: Something broke or produces wrong results. Evidence: "it was working before", wrong calculation, missing data that should be there based on account entries.
- FEATURE REQUEST: Client wants something NEW that the template was never designed to do. Additional columns, new sections, new exports.
- HOW-TO: Client doesn't know how to use an existing feature (identified in Step 2).
- SYNC/DATA: Data synchronisation or import/export problems.
- OTHER: Doesn't fit above.

STEP 5: CHECK THE LEGAL ANGLE.
- Is this required by Luxembourg law, eCDF, PCN? → Higher impact, likely a real bug.
- Is this a client preference? → Lower impact, feature request.
- APPLICABILITY: Does this even apply to the client's entity type? SA vs SARL vs SCSp have different requirements. Micro/small/full regime differences.
- If the client cites a law, VERIFY IT. Clients sometimes misquote regulations.

STEP 6: THINK ABOUT THE SOLUTION.
- For bugs: what's the fix? Is there a workaround the client can use NOW while we work on a permanent fix?
- For complex fixes: ALWAYS propose a workaround first, then describe the permanent fix separately.
- For feature requests: is it feasible in Silverfin Liquid? Does it make accounting sense?
- For how-to: just explain how to use the feature. No development needed.
- PATTERN CHECK: look at how other sections in the same template handle similar logic. Propose solutions consistent with existing patterns.

STEP 7: ASSESS EFFORT HONESTLY.
Don't inflate effort. A label fix = 1 hour. A new section with conditions = 1 day. Be realistic based on what you see in the code.

YOUR OUTPUT:

1. CLASSIFICATION: From Step 4 above. What this ACTUALLY is, not what the client calls it.

2. SUMMARY: One clear, natural sentence in French. Mention the template/note by name. Write it like you'd say it to a colleague.
   Good: "Le texte de la note frais de personnel est incohérent quand l'effectif passe de >0 à 0 entre deux exercices."
   Good: "Le client demande d'ajouter des colonnes 'Siège social' et 'Participation' au tableau des immobilisations financières — c'est une nouvelle fonctionnalité."
   Good: "Le client ne sait pas comment activer l'affichage multi-périodes dans la note du bilan — la fonctionnalité existe déjà dans les paramètres."
   Bad: "Le client signale un problème avec un template." (too vague)

3. ANALYSIS: 3-6 sentences in French, written naturally like a senior analyst would explain to the PO. Cover:
   - What template/note is affected and what it does
   - What the real issue is (your assessment, not just repeating the client)
   - For how-to: WHERE the feature is and HOW to use it
   - For bugs: what's broken and what the fix looks like (+ workaround if applicable)
   - For feature requests: feasibility and whether it makes sense
   - Legal basis: "Exigence légale: Oui — [reference]" or "Non — préférence client"
   - Code check result: does the feature exist already? Is the template working as designed?
   - DON'T pad with generic statements. Every sentence should add information.
   - DON'T write section headers like "LEGAL CHECK:" or "CODE CHECK:" — weave it naturally into the analysis.

4. RISK: critical/high/medium/low based on actual severity, not client words.

5. RICE SCORING:
{client_ctx_block}
   REACH (1-10): 1=one entity, 5=anyone using this template, 10=every client+every file
   IMPACT (1-5): 1=cosmetic, 3=moderate with workaround, 5=compliance/calculation error
   CONFIDENCE (1-5): 1=vague ticket, 3=reasonable understanding, 5=fully clear with evidence
   EFFORT (1-10): 1=~1h label fix, 3=half day logic change, 5=2 days complex logic, 7=1 week, 10=1 month+
   FORMULA: (Reach × Impact × Confidence) / Effort

LANGUAGE: Summary and analysis in FRENCH. Use proper terminology. Keep template names as-is in English. Other JSON fields stay in English.

TONE: Write like a colleague who knows their stuff, not like a robot filling out a form. Be direct, practical, and to the point. Short for simple issues. Detailed only when complexity demands it.

Reply ONLY with valid JSON (no markdown, no code blocks):
{{"classification":"bug|feature_request|how_to|sync|data|other","confidence":0-100,"needs_review":true/false,"summary":"une phrase claire et naturelle en français — mentionner le template/note","analysis":"analyse concise et pratique en français, 3-6 phrases, pas robotique","risk_level":"low|medium|high|critical","rice_reach":number_1to10,"rice_impact":number_1to5,"rice_confidence":number_1to5,"rice_effort":number_1to10,"template_name":"name of affected template (or empty string)","workflow_name":"Annual Accounts|Corporate Tax|Legal Docs (or empty string)"}}"""

    short_thread = truncate_text(compiled_thread, 6000)
    short_kb = truncate_text(kb_context, 5000) if kb_context else ""

    user_msg = short_thread
    if short_kb:
        user_msg += f"\n\nKNOWLEDGE BASE CONTEXT — YOU MUST READ AND VERIFY AGAINST THIS BEFORE CLASSIFYING OR SCORING. Contains commercial law provisions, chart of accounts, template docs. Never contradict this KB:\n{short_kb}"

    # Retry loop: if JSON parse fails on first attempt, retry once with a nudge
    max_attempts = 2
    last_raw_output = ""
    for attempt in range(max_attempts):
        resp = call_anthropic_with_retry(
            client,
            model="claude-sonnet-4-5",
            max_tokens=1500,
            system=system,
            messages=[{"role": "user", "content": user_msg}] if attempt == 0 else [
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": last_raw_output},
                {"role": "user", "content": "Your previous response was not valid JSON. Please reply with ONLY valid JSON (no markdown, no code blocks, no explanations). Start with { and end with }."}
            ],
        )
        text = resp.content[0].text.strip()
        last_raw_output = text

        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip()
        try:
            result = json.loads(text)
            result.setdefault("classification", "other")
            result.setdefault("confidence", 0)
            result.setdefault("needs_review", True)
            result.setdefault("summary", "")
            result.setdefault("analysis", "")
            result.setdefault("risk_level", "medium")
            result.setdefault("draft_response", "")
            result.setdefault("template_name", "")
            result.setdefault("workflow_name", "")
            # Strip any code that leaked into text fields
            if result.get("analysis"):
                result["analysis"] = strip_code_from_output(result["analysis"])
            if result.get("summary"):
                result["summary"] = strip_code_from_output(result["summary"])
            return result
        except json.JSONDecodeError:
            if attempt < max_attempts - 1:
                logger.warning(f"Analysis JSON parse failed (attempt {attempt+1}), retrying...")
                continue
            # Final attempt failed — log the raw output for debugging and return a useful error
            logger.error(f"Analysis JSON parse failed after {max_attempts} attempts. Raw output: {text[:500]}")
            return {"classification": "other", "confidence": 0, "needs_review": True,
                    "summary": "Analysis failed — AI returned invalid format (see raw output below)",
                    "analysis": f"[PARSE ERROR] The AI did not return valid JSON after {max_attempts} attempts. Raw output preserved for debugging:\n\n{text[:1000]}",
                    "risk_level": "medium",
                    "draft_response": "", "template_name": "", "workflow_name": "",
                    "_raw_output": text}  # Store full raw output for DB logging


def generate_draft_response(compiled_thread, anthropic_key, lang="fr", writing_style="customer_support",
                            kb_context="", project_instructions="", analysis="", po_reason="", code_context="", terminology_context="", ticket_id=None, force_simple=False,
                            classification="", priority=""):
    """Generate a draft response (in a specific language) for an APPROVED ticket.
    This is called AFTER PO approval, separately from the analysis step.
    Produces 3 sections: CLIENT RESPONSE, INTERNAL NOTE (BSO LUX), BACKLOG TICKET."""
    client = Anthropic(api_key=anthropic_key)
    style = WRITING_STYLES.get(writing_style, WRITING_STYLES["customer_support"])

    project_ctx = ""
    if project_instructions:
        project_ctx = f"\nProject context & instructions:\n{truncate_text(project_instructions, 2000)}\n"

    term_ctx = ""
    if terminology_context:
        term_ctx = f"\n{terminology_context}\n"

    # Map classification and priority to human-readable labels for BSO note
    classification_map = {
        "bug": "Bug",
        "feature_request": "Feature Request",
        "enhancement": "Enhancement",
        "how_to": "How-to",
        "pending": "Pending",
    }
    priority_map = {
        "1": "Low",
        "2": "Normal",
        "3": "High",
        "4": "Urgent",
        "low": "Low",
        "normal": "Normal",
        "high": "High",
        "urgent": "Urgent",
    }
    classification_label = classification_map.get(str(classification).lower().strip(), classification or "Unknown")
    priority_label = priority_map.get(str(priority).lower().strip(), priority or "Normal")

    code_ctx = ""
    if code_context:
        # Check if this is a Code Agent functional brief (no raw code) vs raw template code
        is_agent_brief = code_context.startswith("CODE AGENT BRIEF") or "TEMPLATE OVERVIEW" in code_context[:200]
        if is_agent_brief:
            code_ctx = f"""
TEMPLATE ANALYSIS (functional description — produced by the Code Agent, already in plain language):
{code_context}

USE THIS ANALYSIS TO:
1. Understand what the template CURRENTLY does — verify the client's claim.
2. Identify REFERENCE PATTERNS: other sections in the same template that already handle similar logic.
3. Propose solutions CONSISTENT with existing patterns — never simplify or remove safeguards.
4. Understand visibility rules, dropdown logic, and conditional display behaviour.
REMINDER: This analysis is already in plain language. Do NOT add code references, variable names, or file paths to your output.
"""
        else:
            code_ctx = f"""
TEMPLATE CODE REFERENCE (for your understanding ONLY — ABSOLUTELY NEVER reference any code, variable names, file paths, or template IDs in your output):
{code_context}

HOW TO USE THIS (internally only — NEVER expose any of this in your output):
1. Read it to understand what the template CURRENTLY does — verify the client's claim.
2. FIND REFERENCE PATTERNS in OTHER SECTIONS of the same template.
3. PROPOSE SOLUTIONS CONSISTENT WITH EXISTING PATTERNS.
4. Understand WHY things exist: mandatory dropdowns = data consistency, conditional visibility = prevent invalid combos.
WARNING: If you write ANY variable name, file path, template ID, or code logic in your output, it is a CRITICAL ERROR.
"""

    if lang == "fr":
        lang_instruction = """LANGUAGE RULE — THIS IS MANDATORY AND NON-NEGOTIABLE:
Write ALL responses in FRENCH. Every single word, sentence, paragraph must be in French.
Do NOT copy English phrases from the context data below — translate everything.
The ONLY exception: "Proposed new wording" section must have BOTH FR and EN text for developer reference.
If you see English text in the KB AGENT BRIEF or RESEARCH AGENT BRIEF below, that is internal context data —
you must TRANSLATE the relevant information into French, never copy it in English."""
    else:
        lang_instruction = """LANGUAGE RULE — THIS IS MANDATORY AND NON-NEGOTIABLE:
Write ALL responses in ENGLISH. Every single word, sentence, paragraph must be in English.
Do NOT copy French phrases from the context data below — translate everything.
The ONLY exception: "Proposed new wording" section must have BOTH FR and EN text for developer reference.
If you see French text in ticket data below, that is source data — translate to English in your output."""

    system = f"""You are the Product Owner for Silverfin's Luxembourg templates team (BSO LUX).
You have APPROVED this ticket. Now draft the 3 responses needed to action it.

{lang_instruction}

BEFORE WRITING, THINK THROUGH THIS PROCESS (do not output this — use it internally):

1. WHAT'S REALLY GOING ON?
   Re-read the ticket and analysis. What is the client's ACTUAL problem? Not what they said — what they NEED.
   - Did they report a bug that's actually the template working as designed? (e.g. dividend distribution logic, equity movement rules)
   - Did they ask for a feature that already exists behind a setting they don't know about?
   - Is this just a typo/label/translation fix? → Keep everything short.

2. IS THE TEMPLATE WORKING AS DESIGNED?
   If template code/analysis was provided, check what the template ACTUALLY does.
   Many things clients report as "bugs" are intentional:
   - Mandatory dropdowns that restrict choices → they prevent invalid data combinations
   - Conditional visibility that hides sections → it prevents showing irrelevant information
   - Specific row behaviour in tables (dividend distribution, depreciation, etc.) → follows legal/accounting logic
   If the template is working as designed, the client needs guidance, not a code change.

3. DOES THE SOLUTION ALREADY EXIST?
   Check for: settings, toggles, dropdowns, visibility conditions that already do what the client wants.
   If yes → explain HOW to use it. No development work needed.

4. LEGAL CHECK:
   Is this required by law (Loi 2002, Code de Commerce, eCDF, CSSF)? Or just a preference?
   Does this apply to the client's entity type (SA/SARL/SCSp) and regime (micro/small/full)?

5. SOLUTION APPROACH:
   - Quick fixes (typo, label, font, cropped text, table alignment): just describe what to change. Done.
   - Bugs with workaround: propose the WORKAROUND first (what the client can do NOW), then the permanent fix.
   - Complex changes: describe the fix in detail, following existing patterns in the template.
   - Feature requests: assess feasibility, propose how it would work, consistent with existing patterns.
   - How-to: explain how to use the existing feature. No fix needed.

6. PATTERN CHECK: Look at other sections in the template. How do they handle similar logic? Follow the same approach.

7. SCOPE: Does this affect other templates, periods (N-1, N-2), or company types (SA vs SARL)?

{style['instructions']}{project_ctx}{term_ctx}

SILVERFIN CONTEXT:
- Cloud accounting platform for Luxembourg firms (BDO, EY, KPMG, Deloitte, JTC, etc.)
- Workflows: "Luxembourg Annual Accounts", "Luxembourg Corporate Tax", "Legal Documents"
- Template types: reconciliation texts, working papers, account templates
- Notes: Financial assets, Receivables, Payables, Staff costs, Provisions, Equity, Stocks, Fixed assets
- Legal docs: board minutes (PV du Conseil de Gérance / PV du Conseil d'Administration), general meeting (AG/AGO), written resolutions, proxy, convening notices
- Company types: SA (société anonyme) and SARL (société à responsabilité limitée)

LUXEMBOURG LEGAL TERMINOLOGY — EXACT terms (critical):
- SARL 1 manager: "Gérant Unique" / "Sole Manager"
- SARL 2+ managers: "Conseil de Gérance" / "Board of Managers" (NEVER "Gérants")
- SA 1 director: "Administrateur Unique" / "Sole Director"
- SA 2+ directors: "Conseil d'Administration" / "Board of Directors"
- Commissaire aux comptes / Réviseur d'entreprises agréé for auditors

{code_ctx}

The PO's analysis of this ticket:
{truncate_text(analysis, 1500) if analysis else "No detailed analysis available."}

{("PO decision notes: " + po_reason) if po_reason else ""}

COMPLEXITY MATCHING — CRITICAL (this determines the length of your ENTIRE response):
First, CLASSIFY this issue before writing anything:
- SIMPLE (translation fix, wording change, label update, typo, missing/wrong word):
  Client response = 2-3 sentences. "Bien reçu, on corrige. Merci." style.
  Internal note = state the problem, give EXACT new wording (FR + EN), done.
  Skip BACKLOG TICKET entirely. Total response UNDER 200 words.
  Examples: "change 'shares' to 'partnership interests'", "add missing comma", "translate label X", "fix typo in note Y"
- MEDIUM (visibility condition, display logic, single template change, HOW-TO explanation):
  Client response = 3-5 sentences. Direct, clear, shows understanding.
  Internal note = issue + conditions + next step + proposed fix. Under 500 words total.
- COMPLEX (multi-template change, new feature, structural redesign, legal/compliance issue):
  Full detail with reference patterns, edge cases, backlog ticket. Up to 800 words total.

GOLDEN RULE: If the fix can be described in one sentence ("change X to Y"), the entire response should be SHORT.
A senior PO does not write 500 words about a typo. That signals insecurity, not thoroughness.
LESS IS MORE for simple issues. Brevity shows mastery — you understood immediately and acted.

{'''FORCED SIMPLICITY MODE — ABSOLUTE BREVITY REQUIRED. THIS OVERRIDES EVERYTHING BELOW.

This ticket is a SIMPLE wording/translation/label fix. Your ENTIRE output must fit the template below. Violating the length limit is a CRITICAL ERROR.

HARD LIMITS (non-negotiable):
- CLIENT RESPONSE: 2 sentences MAX. Just: acknowledge + confirm the fix will be made. Sign off.
- INTERNAL NOTE: 5-7 lines MAX. Structure:
  "Hi team,
   Agreed — [one-line problem statement].
   Next step: update the wording in the [note name] (FR + EN).
   FR : '<new French text>'
   EN : '<new English text>'
   Thanks"
- BACKLOG TICKET: OMIT ENTIRELY. Do NOT output the "--- BACKLOG TICKET ---" header at all.

BANNED in simple mode:
- No "I agree with the client" paragraph explaining reasoning
- No bullet-point condition lists (no "* when X = 0...")
- No edge case sections (no "Quand les deux exercices = 0...")
- No legal/accounting justifications
- No reference-pattern discussion
- No "Proposed new wording (when...)" conditional framings
- No template logic explanations
- No "Current text:" block unless the fix is replacing a specific quoted sentence (then max 1 line)

TOTAL OUTPUT: under 120 words combined (client + internal). If you write more, you have failed.''' if force_simple else ""}

ABSOLUTE RULES — VIOLATION OF ANY OF THESE IS A CRITICAL ERROR:

1. ABSOLUTELY ZERO CODE IN OUTPUT. This is the most important rule. NEVER write:
   - Variable names (employees_cy, hide_breakdown_due_to_no_fte, company.custom.people, etc.)
   - File names or paths (main.liquid, lux_aa_an_staff_cost, text_parts/translations.liquid, etc.)
   - Template technical IDs (lux_ci_general_information, lux_aa_an_equity, etc.)
   - Programming logic descriptions (if employees_cy == 0, when variable X > Y, etc.)
   - Section names from code (INTRO SECTION, etc.)
   - ANY reference to how the code works internally

   You USE the code to UNDERSTAND the issue. You then describe it in PLAIN HUMAN LANGUAGE.
   BAD: "La variable employees_cy = 0 et employees_py = 0.5"
   GOOD: "L'effectif de l'exercice en cours est de 0 et celui de l'exercice précédent est de 0,5"
   BAD: "Modifier la condition dans la section INTRO SECTION du template lux_aa_an_staff_cost/main.liquid"
   GOOD: "Modifier la phrase d'introduction de la note frais de personnel"
   BAD: "Fichiers concernés : lu_market-main/reconciliation_texts/..."
   GOOD: (never list file names — the BSO team knows which files to change from the template name)

2. NEVER mix languages. French version = 100% French. English version = 100% English.
   The ONLY exception: when proposing a wording change in the template, provide BOTH the FR and EN text
   because BSO needs to update both language versions. This is the ONLY place where both languages appear.

3. NO MARKDOWN. No **, no #, no ```. Plain text only.

4. "Next step:" is written in ONE language only — the language of this draft version.
   In the FR version: "Next step:" followed by French text ONLY.
   In the EN version: "Next step:" followed by English text ONLY.
   NEVER duplicate the next step in both languages.

5. Be SPECIFIC. Don't repurpose the client's words. Analyse the code, form your own opinion, propose the fix.

6. Match length to complexity. Wording fix = concise. Multi-condition bug = detailed.

EXAMPLE OF A GOOD INTERNAL NOTE (French version — follow this exact style and ORDER):
"Hi team,

Je suis d'accord avec le client. La formulation de la note frais de personnel est incohérente quand l'effectif de l'exercice en cours est de 0 mais que l'exercice précédent avait des employés.

Texte actuel : 'La Société n'a pas employé de personnel durant l'exercice clos au 31 décembre 2025 (exercice précédent: 5,00) répartie comme suit:'
Cette phrase n'a pas de sens avec un effectif à zéro — on ne peut pas montrer une ventilation de 0 personnes.

Le problème survient quand :
* L'effectif de l'exercice en cours = 0
* L'effectif de l'exercice précédent > 0
* Le tableau de ventilation apparaît car les données de l'exercice précédent existent
* Mais la structure de la phrase est incorrecte pour un effectif de zéro

Next step :
Modifier la section d'introduction de la note frais de personnel pour gérer le cas effectif exercice en cours = 0 avec exercice précédent > 0. Mettre à jour les versions française et anglaise du texte.

Proposed new wording (quand effectif exercice en cours = 0 et exercice précédent > 0) :
FR : 'La Société n'a employé aucun personnel au cours de l'exercice clôturé le [DATE]. Au cours de l'exercice précédent clôturé le [DATE_PY], l'effectif moyen était de [X] personnes, La répartition était la suivante :'
EN : 'The Company had no employees during the financial year ended [DATE]. In the prior year ended [DATE_PY], the average number of employees was [X] persons, the breakdown was as follows:'

Quand l'effectif de l'exercice en cours > 0 : aucun changement.
Quand les deux exercices = 0 : garder le comportement existant (pas de tableau de ventilation, phrase différente).

Thanks"

NOTE: The example above shows BOTH FR and EN only for the proposed wording change (because BSO needs to implement both). Everything else is in French only because this is the French version.
IMPORTANT: The "Next step:" ALWAYS comes BEFORE the "Proposed new wording". This order is mandatory.

YOUR RESPONSE MUST CONTAIN EXACTLY 3 SECTIONS with these EXACT headers:

--- CLIENT RESPONSE ---
Write as a PO who knows these templates inside out. Your tone is natural, confident, and human — like a knowledgeable colleague replying to another professional.

KEY PRINCIPLES:
- Get straight to the point. No generic openings ("Thank you for bringing this to our attention", "We take all feedback seriously" — NEVER).
- Show you understood the specific issue by naming the template/note and the exact scenario.
- Be direct about what will happen: fix it, explain how to use it, or explain why not.
- QUICK FIXES (typo, label, font): 2 sentences max. "Bien vu, on corrige. Ce sera dans la prochaine mise à jour."
- BUGS WITH WORKAROUND: acknowledge the issue, give the workaround they can use NOW, then mention the fix is coming.
- HOW-TO / WORKING AS DESIGNED: explain HOW it works and WHY (the functional reason, not the code reason). If the client misunderstands the logic (e.g. dividend distribution flow), explain the accounting logic naturally. Don't be condescending — just clear.
- FEATURE REQUEST: acknowledge the idea, say whether it's feasible, and what the next step is.
- Keep it warm but professional. No fluff, no hedging, no over-explaining.

--- INTERNAL NOTE (BSO LUX) ---
Written BY the PO TO the BSO development team. Natural, clear, practical.

Structure:
1. "Hi team,"
2. Classification + priority on its own line:
   - Bug: "Classification: Bug — Priority: [High/Normal/Low]"
   - Feature: "Classification: Feature Request — Priority: [High/Normal/Low]"
   - Enhancement: "Classification: Enhancement — Priority: [High/Normal/Low]"
   - How-to: "Classification: How-to (no development needed)"
   Classification: {classification_label}. Priority: {priority_label}. Adjust if your analysis warrants it.
3. Your position: agree / partially agree / working as designed / needs investigation
4. The issue in plain language (what's happening, when it happens). NEVER reference code.
5. IF WORKING AS DESIGNED: explain WHY the template behaves this way (the functional/accounting reason).
   Example: "This is intentional — the dividend distribution row follows the legal profit allocation sequence. The client needs to first allocate to legal reserve before distributing."
6. IF BUG WITH WORKAROUND: describe the workaround first ("In the meantime, the client can..."), then the permanent fix.
7. Conditions when the issue occurs (bullet points with *)
8. "Next step:" — ONE language only. Concrete action, no file names.
9. "Proposed new wording" — for text changes: EXACT current + new text (FR + EN). For logic: describe condition + expected result.
10. Edge cases if relevant.
11. "Thanks"
- NEVER list file paths. BSO knows which files to change.
- For QUICK FIXES (typo, label, font, cropped): keep the note SHORT. 3-5 lines. Don't over-explain obvious fixes.

--- BACKLOG TICKET ---
Skip ENTIRELY if it's a simple wording fix.

Template/Note: [human-readable name]
Workflow: [workflow name]
Type: Bug fix / Feature request / Enhancement

Current behaviour: [plain language]
Expected behaviour: [specific — what text, what condition, what display]
Impact: [who is affected]

Reply with ONLY the response text starting with "--- CLIENT RESPONSE ---". No JSON, no markdown blocks."""

    short_thread = truncate_text(compiled_thread, 6000)
    short_kb = truncate_text(kb_context, 5000) if kb_context else ""

    user_text = short_thread
    if short_kb:
        user_text += f"\n\nKNOWLEDGE BASE CONTEXT — YOU MUST READ AND VERIFY AGAINST THIS BEFORE RESPONDING. If you cite a law, regulation, or accounting rule, it MUST match what is in this KB. Never contradict the KB:\n{short_kb}"

    # Load screenshots as vision content blocks if available
    screenshot_blocks = []
    if ticket_id:
        screenshot_blocks = load_screenshots_for_ai(ticket_id)

    if screenshot_blocks:
        # Build multimodal content: screenshots first, then text
        user_content = screenshot_blocks + [{"type": "text", "text": user_text}]
    else:
        user_content = user_text

    resp = call_anthropic_with_retry(
        client,
        model="claude-sonnet-4-5",
        max_tokens=600 if force_simple else 2000,
        system=system,
        messages=[{"role": "user", "content": user_content}],
    )
    return strip_code_from_output(resp.content[0].text.strip())


def generate_decline_response(compiled_thread, anthropic_key, lang="fr", writing_style="customer_support",
                              kb_context="", project_instructions="", analysis="", decline_reason="", code_context="", terminology_context="", ticket_id=None):
    """Generate a professional decline response for a DECLINED ticket.
    Explains to the client why we cannot or will not action this request,
    and provides an internal note summarising the decision."""
    client = Anthropic(api_key=anthropic_key)
    style = WRITING_STYLES.get(writing_style, WRITING_STYLES["customer_support"])

    project_ctx = ""
    if project_instructions:
        project_ctx = f"\nProject context & instructions:\n{truncate_text(project_instructions, 2000)}\n"

    term_ctx = ""
    if terminology_context:
        term_ctx = f"\n{terminology_context}\n"

    code_ctx = ""
    if code_context:
        is_agent_brief = code_context.startswith("CODE AGENT BRIEF") or "TEMPLATE OVERVIEW" in code_context[:200]
        if is_agent_brief:
            code_ctx = f"\nTEMPLATE ANALYSIS (functional description — already in plain language, do NOT add code references):\n{code_context}"
        else:
            code_ctx = f"\nTEMPLATE CODE REFERENCE (for your understanding ONLY — NEVER reference any code in your output):\n{code_context}"

    if lang == "fr":
        lang_instruction = """LANGUAGE RULE — MANDATORY: Write ALL responses in FRENCH. Every word must be in French.
Do NOT copy English phrases from context data below — translate everything to French."""
    else:
        lang_instruction = """LANGUAGE RULE — MANDATORY: Write ALL responses in ENGLISH. Every word must be in English.
Do NOT copy French phrases from context data below — translate everything to English."""

    system = f"""You are the Product Owner for Silverfin's Luxembourg templates team (BSO LUX).
You have DECLINED this ticket. Draft a professional response explaining why, and an internal note documenting the decision.

{lang_instruction}

BEFORE WRITING, THINK THROUGH (do not output this — but your response MUST reflect these checks):
1. KNOWLEDGE BASE CHECK (MANDATORY — DO THIS FIRST):
   Read the KNOWLEDGE BASE CONTEXT below carefully. It contains:
   - Luxembourg commercial law provisions (Loi du 19 décembre 2002, Code de Commerce, etc.)
   - Chart of accounts (plan comptable normalisé) with account ranges
   - Template documentation, reconciliation rules, Liquid limitations
   YOU MUST verify your decline reason against the KB. If the KB says something specific about the topic,
   your response MUST be consistent with it. NEVER contradict the knowledge base.
   If the KB has the correct legal provision and you cite it differently → that is a CRITICAL ERROR.
2. LEGAL CHECK: Is the current behaviour actually correct per Luxembourg law / eCDF / CSSF?
   - If YES: cite the SPECIFIC law, article, or provision (from the KB if available).
   - If NO: the ticket should probably not be declined — flag this.
   - If NOT A LEGAL MATTER: say so. Don't invent legal justifications for preference-based declines.
3. CODE CHECK: If template code/analysis was provided, verify the current behaviour is indeed working as designed.
   Does the template already handle this correctly? Verify against the functional description.
4. REASON: Why exactly is this declined? Be specific — "working as designed", "compliant with regulation X article Y", "client-specific preference vs standard", "low impact", "duplicate of ticket X", "not technically feasible because..."
   CRITICAL: If you cite a law or regulation, make sure it's the CORRECT one. Check the KB first.
5. ALTERNATIVE: Is there a workaround the client can use? A different way to achieve what they want?

{style['instructions']}{project_ctx}{term_ctx}

{code_ctx}

SILVERFIN CONTEXT:
- Cloud accounting platform for Luxembourg firms (BDO, EY, KPMG, Deloitte, JTC, etc.)
- Workflows: "Luxembourg Annual Accounts", "Luxembourg Corporate Tax", "Legal Documents"

LUXEMBOURG LEGAL TERMINOLOGY — EXACT terms:
- SARL 1 manager: "Gérant Unique" / "Sole Manager" (NEVER "Gérants")
- SARL 2+ managers: "Conseil de Gérance" / "Board of Managers"
- SA 1 director: "Administrateur Unique" / "Sole Director"
- SA 2+ directors: "Conseil d'Administration" / "Board of Directors"

Ticket analysis:
{truncate_text(analysis, 1500) if analysis else "No detailed analysis available."}

PO decline reason: {decline_reason if decline_reason else "No specific reason provided."}

COMPLEXITY MATCHING: Keep your response proportional to the issue. Simple decline (out of scope, duplicate) = 3-4 sentences. Complex decline (technically infeasible, regulation conflict) = more detail needed.

ABSOLUTE RULES:
1. ABSOLUTELY ZERO CODE. Never write variable names, file paths, template IDs, code logic, or any programming terminology.
   Use the code to UNDERSTAND, then write in plain human language.
2. NEVER mix languages. French = 100% French. English = 100% English. No exceptions.
3. NO MARKDOWN. No **, no #. Plain text only.
4. Be SPECIFIC about why. Cite regulation, design rationale, or priority reason.
5. Offer a workaround if one exists.

Write a SINGLE professional decline response to the client. Do NOT split into sections or use any headers like "--- CLIENT RESPONSE ---" or "--- INTERNAL NOTE ---".

The response should:
- Sound like a senior PO who deeply understands the issue — not a support agent following a script
- Be direct and confident: "This works as designed because..." not "Unfortunately we are unable to..."
- Explain WHY specifically this is declined (cite regulation, design rationale, or priority reason)
- Suggest a workaround if one exists
- Keep it proportional: simple decline = 3-4 sentences max. Don't over-explain obvious things.

Reply with ONLY the response text. No JSON, no markdown blocks, no section headers."""

    short_thread = truncate_text(compiled_thread, 6000)
    short_kb = truncate_text(kb_context, 5000) if kb_context else ""

    user_text = short_thread
    if short_kb:
        user_text += f"\n\nKNOWLEDGE BASE CONTEXT — YOU MUST READ AND VERIFY AGAINST THIS BEFORE RESPONDING:\n{short_kb}"

    # Load screenshots as vision content blocks if available
    screenshot_blocks = []
    if ticket_id:
        screenshot_blocks = load_screenshots_for_ai(ticket_id)

    if screenshot_blocks:
        user_content = screenshot_blocks + [{"type": "text", "text": user_text}]
    else:
        user_content = user_text

    resp = call_anthropic_with_retry(
        client,
        model="claude-sonnet-4-5",
        max_tokens=1500,
        system=system,
        messages=[{"role": "user", "content": user_content}],
    )
    return strip_code_from_output(resp.content[0].text.strip())


def translate_draft(text, source_lang, target_lang, anthropic_key):
    """Translate a draft response from one language to another using Haiku (fast + cheap).
    Preserves the exact structure, section headers, meaning, and proposed wordings.
    The ONLY thing that changes is the language of the surrounding text."""
    if not text or not text.strip():
        return text

    client = Anthropic(api_key=anthropic_key)

    source = "French" if source_lang == "fr" else "English"
    target = "French" if target_lang == "fr" else "English"

    system = f"""You are a professional translator for Silverfin's Luxembourg templates team.
Translate the following {source} text to {target}.

RULES:
1. Preserve the EXACT same structure, sections, headers (--- CLIENT RESPONSE ---, --- INTERNAL NOTE (BSO LUX) ---, --- BACKLOG TICKET ---).
2. Preserve the EXACT same meaning, solutions, conditions, and technical accuracy.
3. Keep "Proposed new wording" sections as-is — they already contain both FR and EN text.
4. Keep "Hi team," and "Thanks" as-is (they stay in English in both versions).
5. Keep Luxembourg legal terms accurate: Gérant Unique, Conseil de Gérance, etc.
6. NO MARKDOWN, NO CODE, NO VARIABLE NAMES. Plain text only.
7. Do NOT add, remove, or change any content. Only translate the language.
8. The "Next step:" label stays as "Next step:" in both languages — only translate the text after it.

Reply with ONLY the translated text. No explanations."""

    resp = call_anthropic_with_retry(
        client,
        model="claude-haiku-4-5-20251001",
        max_tokens=2500,
        system=system,
        messages=[{"role": "user", "content": text}],
    )
    return strip_code_from_output(resp.content[0].text.strip())


# Luxembourg legal terms that must be exact in translation — never paraphrased.
# Maps FR → EN (authoritative pairs). If the EN translation drifts, we fix it.
LEGAL_TERM_PAIRS = {
    "Gérant Unique": "Sole Manager",
    "Conseil de Gérance": "Board of Managers",
    "Administrateur Unique": "Sole Director",
    "Conseil d'Administration": "Board of Directors",
    "Commissaire aux comptes": "Statutory Auditor",
    "Réviseur d'entreprises agréé": "Approved Statutory Auditor",
    "Plan Comptable Normalisé": "Standardized Chart of Accounts",
    "Registre de Commerce et des Sociétés": "Trade and Companies Register",
}


def validate_translation(source_fr, translated_en, anthropic_key):
    """Validate that the EN translation preserves critical Luxembourg legal terms.
    Does a fast term-level check (no API call needed for most cases).
    Only calls the API for a reverse-translation spot-check if the draft is long.

    Returns the corrected EN translation (or the original if no issues found).
    """
    if not source_fr or not translated_en:
        return translated_en

    corrected = translated_en
    fixes_applied = []

    # Step 1: Fast term-level check — ensure exact legal term pairs
    for fr_term, en_term in LEGAL_TERM_PAIRS.items():
        if fr_term in source_fr:
            # The FR source uses this term — check the EN version has the correct translation
            if en_term not in corrected:
                # Check for common mistranslations
                wrong_variants = {
                    "Sole Manager": ["Solo Manager", "Unique Manager", "Single Manager"],
                    "Board of Managers": ["Management Board", "Gérants", "Board of Gérants", "Manager Board"],
                    "Sole Director": ["Solo Director", "Unique Director", "Single Director"],
                    "Board of Directors": ["Directors Board", "Administration Board"],
                }
                for wrong in wrong_variants.get(en_term, []):
                    if wrong in corrected:
                        corrected = corrected.replace(wrong, en_term)
                        fixes_applied.append(f"Fixed: '{wrong}' → '{en_term}'")
                        break

    # Step 2: For longer drafts (500+ words), do a quick reverse-translation spot-check
    # using Haiku. This catches semantic drift that term-matching can't detect.
    # Skip for short drafts to avoid unnecessary API calls (speed matters).
    if len(source_fr.split()) > 500 and anthropic_key:
        try:
            client = Anthropic(api_key=anthropic_key)
            check_resp = call_anthropic_with_retry(
                client,
                model="claude-haiku-4-5-20251001",
                max_tokens=500,
                system="""You are a translation quality checker for Luxembourg legal/accounting documents.
Compare the French source and English translation below. Check ONLY for:
1. Legal terms mistranslated (e.g., "Conseil de Gérance" must be "Board of Managers", never "Management Board")
2. Numbers or amounts that changed
3. Sections that were omitted or added in translation
4. Meaning that shifted significantly

Reply with ONLY a JSON object:
{"issues_found": true/false, "fixes": [{"original": "wrong text", "corrected": "right text"}]}
If no issues: {"issues_found": false, "fixes": []}""",
                messages=[{"role": "user", "content": f"FRENCH SOURCE:\n{source_fr[:3000]}\n\nENGLISH TRANSLATION:\n{corrected[:3000]}"}],
            )
            check_text = check_resp.content[0].text.strip()
            if check_text.startswith("```"):
                check_text = check_text.split("```")[1]
                if check_text.startswith("json"):
                    check_text = check_text[4:]
                check_text = check_text.strip()
            check_result = json.loads(check_text)
            if check_result.get("issues_found") and check_result.get("fixes"):
                for fix in check_result["fixes"][:5]:
                    if fix.get("original") and fix.get("corrected") and fix["original"] in corrected:
                        corrected = corrected.replace(fix["original"], fix["corrected"], 1)
                        fixes_applied.append(f"AI fix: '{fix['original'][:50]}' → '{fix['corrected'][:50]}'")
        except Exception as e:
            logger.warning(f"Translation validation API check failed: {e}")

    if fixes_applied:
        logger.info(f"Translation validation applied {len(fixes_applied)} fix(es): {fixes_applied}")

    return corrected


# ── Freshdesk-reply learning helper ─────────────────────────────────────────

def _extract_last_public_agent_reply(conversations):
    """From a list of Freshdesk conversations, return the most recent PUBLIC agent
    reply (outgoing, non-private) along with its id and created_at.
    Returns (reply_dict, conv_id, created_at_iso) or (None, None, None)."""
    if not conversations:
        return None, None, None

    # Sort by id (monotonic) to find the latest reply deterministically
    sorted_convs = sorted(
        conversations,
        key=lambda c: (c.get("id") or 0),
        reverse=True
    )
    for c in sorted_convs:
        # incoming=True → from customer; incoming=False → from agent
        if c.get("incoming", True):
            continue
        # Skip private notes (those are internal, not sent to the client)
        if c.get("private", False):
            continue
        body = strip_html(c.get("body", "") or c.get("body_text", "") or "")
        if not body.strip():
            continue
        return c, c.get("id"), c.get("created_at", "")
    return None, None, None


def _maybe_learn_from_freshdesk_reply(ticket_id, conversations, anthropic_key, domain, db=None):
    """If a new public agent reply has been posted directly in Freshdesk (without
    the PO editing the draft in this tool), compare it with the stored draft_response
    and run the Learning Agent. Tracked via `tickets.last_learned_conv_id` to avoid
    re-learning from the same reply."""
    if not anthropic_key or not conversations:
        return

    reply, conv_id, conv_created = _extract_last_public_agent_reply(conversations)
    if not reply or not conv_id:
        return

    # Caller (background thread) passes its own db; fall back to Flask g only in request context
    if db is None:
        try:
            db = get_db()
        except Exception:
            db = get_db_standalone()
    row = db.execute(
        "SELECT subject, draft_response, template_name, workflow_name, last_learned_conv_id "
        "FROM tickets WHERE ticket_id = ?",
        (ticket_id,)
    ).fetchone()
    if not row:
        return

    last_seen = row["last_learned_conv_id"] or 0
    if conv_id <= last_seen:
        return  # already learned from this reply (or an even newer one)

    draft = (row["draft_response"] or "").strip()
    if not draft:
        # Nothing to compare against — just update the watermark so we don't keep checking
        db.execute(
            "UPDATE tickets SET last_learned_conv_id = ? WHERE ticket_id = ?",
            (conv_id, ticket_id)
        )
        db.commit()
        return

    reply_text = strip_html(reply.get("body") or reply.get("body_text") or "").strip()
    if not reply_text:
        return

    # Heuristic: if the reply is almost identical to the draft, skip. Only learn
    # when the agent actually reworded/shortened/corrected the draft.
    if reply_text == draft:
        db.execute(
            "UPDATE tickets SET last_learned_conv_id = ? WHERE ticket_id = ?",
            (conv_id, ticket_id)
        )
        db.commit()
        return

    # Require a minimum reply length to avoid learning noise from one-liners
    if len(reply_text) < 60:
        db.execute(
            "UPDATE tickets SET last_learned_conv_id = ? WHERE ticket_id = ?",
            (conv_id, ticket_id)
        )
        db.commit()
        return

    subject = row["subject"] or ""
    template_name = row["template_name"] or ""
    workflow_name = row["workflow_name"] or ""

    # Capture for the background thread (this DB connection belongs to Flask)
    _api = anthropic_key
    _tid = ticket_id
    _subj = subject
    _tpl = template_name
    _wf = workflow_name
    _old = draft
    _new = reply_text
    _conv_id = conv_id

    def _learn_bg():
        try:
            bg_db = get_db_standalone()
            try:
                orch = AgentOrchestrator(_api, db=bg_db)
                orch.run_learning(
                    _tid, _subj, _tpl, _wf, _old, _new,
                    output_type="draft_response",
                    source="freshdesk_reply",
                )
                # Mark this reply as learned-from so we never re-run on it
                bg_db.execute(
                    "UPDATE tickets SET last_learned_conv_id = ? WHERE ticket_id = ?",
                    (_conv_id, _tid)
                )
                bg_db.commit()
                logger.info(
                    f"Freshdesk-reply learning ran for ticket {_tid} (conv {_conv_id})"
                )
            finally:
                bg_db.close()
        except Exception as e:
            logger.warning(f"Freshdesk-reply learning failed for ticket {_tid}: {e}")

    threading.Thread(target=_learn_bg, daemon=True).start()


# ── Background Job ───────────────────────────────────────────────────────────

def run_analysis_job():
    """Run the full analysis pipeline in a background thread."""
    global job_status
    with _job_lock:
        job_status = {"running": True, "progress": "Starting...", "processed": 0, "total": 0, "errors": []}

    db = get_db_standalone()
    try:
        api_key = get_setting("freshdesk_api_key", db=db)
        domain = get_setting("freshdesk_domain", "silverfin.freshdesk.com", db=db)
        group_id = get_setting("freshdesk_group_id", "101000372179", db=db)
        anthropic_key = get_setting("anthropic_api_key", db=db)
        country = get_setting("freshdesk_country", "", db=db)
        statuses = get_setting("freshdesk_statuses", "2,3,4,20", db=db)

        if not api_key:
            _job_add_error("Freshdesk API key not set. Go to Settings.")
            _update_job(running=False)
            return
        ai_enabled = bool(anthropic_key)
        if not ai_enabled:
            logger.info("Anthropic API key not set — fetching tickets without AI analysis.")

        writing_style = get_setting("writing_style", "customer_support", db=db)
        kb_context = get_knowledge_base_context(db)

        # Search tickets
        _update_job(progress="Searching Freshdesk for tickets...")
        try:
            ticket_list = search_tickets(api_key, domain, group_id, country, statuses)
        except requests.exceptions.HTTPError as e:
            _job_add_error(f"Freshdesk API error: {e}")
            _update_job(running=False)
            return

        _update_job(total=len(ticket_list), progress=f"Found {len(ticket_list)} tickets. Processing...")

        def save_ticket_to_db(tid, ticket_data, conversations, compiled, analysis, draft, domain, db, rice=None, screenshots=None):
            """Helper to upsert a processed ticket into the database."""
            requester = ticket_data.get("requester", {})
            if isinstance(requester, dict):
                req_name = requester.get("name", "Unknown")
                req_email = requester.get("email", "Unknown")
            else:
                req_name = "Unknown"
                req_email = "Unknown"
            cf = ticket_data.get("custom_fields", {}) or {}
            now = datetime.now(timezone.utc).isoformat()
            # RICE values: use provided rice dict, or keep zeros
            r = rice or {}
            r_reach = r.get("rice_reach", 0)
            r_impact = r.get("rice_impact", 0)
            r_conf = r.get("rice_confidence", 0)
            r_effort = r.get("rice_effort", 0)
            r_score = round((r_reach * r_impact * r_conf / r_effort), 1) if r_effort > 0 else 0

            # Extract Freshdesk SLA timestamps
            stats = ticket_data.get("stats", {}) or {}
            resolved_at = stats.get("resolved_at", "")
            first_responded_at = stats.get("first_responded_at", "")

            # Calculate SLA times in hours
            sla_resolution_hours = 0
            sla_first_response_hours = 0
            created_at = ticket_data.get("created_at", "")

            if resolved_at and created_at:
                try:
                    created = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    resolved = datetime.fromisoformat(resolved_at.replace("Z", "+00:00"))
                    sla_resolution_hours = (resolved - created).total_seconds() / 3600
                except (ValueError, TypeError):
                    pass

            if first_responded_at and created_at:
                try:
                    created = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    responded = datetime.fromisoformat(first_responded_at.replace("Z", "+00:00"))
                    sla_first_response_hours = (responded - created).total_seconds() / 3600
                except (ValueError, TypeError):
                    pass

            # Extract template and workflow names from analysis
            template_name = analysis.get("template_name", "")
            workflow_name = analysis.get("workflow_name", "")

            db.execute("""
                INSERT INTO tickets (
                    ticket_id, ticket_url, subject, status, status_code,
                    priority, priority_code, group_name, requester_name, requester_email,
                    country, created_at, updated_at, last_analysis,
                    classification, confidence, needs_review, summary, analysis,
                    draft_response, review_status, processing_date, risk_level,
                    raw_description, raw_conversations, compiled_thread,
                    rice_reach, rice_impact, rice_confidence, rice_effort, rice_score,
                    resolved_at, first_responded_at, template_name, workflow_name,
                    sla_resolution_hours, sla_first_response_hours,
                    screenshots_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(ticket_id) DO UPDATE SET
                    ticket_url=excluded.ticket_url, subject=excluded.subject,
                    status=excluded.status, status_code=excluded.status_code,
                    priority=excluded.priority, priority_code=excluded.priority_code,
                    group_name=excluded.group_name, requester_name=excluded.requester_name,
                    requester_email=excluded.requester_email, country=excluded.country,
                    updated_at=excluded.updated_at,
                    last_analysis=CASE WHEN excluded.classification != 'pending' THEN excluded.last_analysis ELSE tickets.last_analysis END,
                    classification=CASE WHEN excluded.classification != 'pending' THEN excluded.classification ELSE tickets.classification END,
                    confidence=CASE WHEN excluded.classification != 'pending' THEN excluded.confidence ELSE tickets.confidence END,
                    needs_review=CASE WHEN excluded.classification != 'pending' THEN excluded.needs_review ELSE tickets.needs_review END,
                    summary=CASE WHEN excluded.summary != '' AND excluded.summary != 'Fetched — waiting for AI analysis' THEN excluded.summary ELSE tickets.summary END,
                    analysis=CASE WHEN excluded.analysis != '' THEN excluded.analysis ELSE tickets.analysis END,
                    draft_response=CASE WHEN excluded.draft_response != '' THEN excluded.draft_response ELSE tickets.draft_response END,
                    processing_date=excluded.processing_date,
                    risk_level=CASE WHEN excluded.classification != 'pending' THEN excluded.risk_level ELSE tickets.risk_level END,
                    raw_description=excluded.raw_description, raw_conversations=excluded.raw_conversations,
                    compiled_thread=excluded.compiled_thread,
                    resolved_at=excluded.resolved_at, first_responded_at=excluded.first_responded_at,
                    template_name=CASE WHEN excluded.template_name != '' THEN excluded.template_name ELSE tickets.template_name END,
                    workflow_name=CASE WHEN excluded.workflow_name != '' THEN excluded.workflow_name ELSE tickets.workflow_name END,
                    sla_resolution_hours=excluded.sla_resolution_hours, sla_first_response_hours=excluded.sla_first_response_hours,
                    screenshots_json=CASE WHEN excluded.screenshots_json != '[]' THEN excluded.screenshots_json ELSE tickets.screenshots_json END,
                    rice_reach=CASE WHEN excluded.rice_score > 0 THEN excluded.rice_reach ELSE tickets.rice_reach END,
                    rice_impact=CASE WHEN excluded.rice_score > 0 THEN excluded.rice_impact ELSE tickets.rice_impact END,
                    rice_confidence=CASE WHEN excluded.rice_score > 0 THEN excluded.rice_confidence ELSE tickets.rice_confidence END,
                    rice_effort=CASE WHEN excluded.rice_score > 0 THEN excluded.rice_effort ELSE tickets.rice_effort END,
                    rice_score=CASE WHEN excluded.rice_score > 0 THEN excluded.rice_score ELSE tickets.rice_score END
            """, (
                tid,
                f"https://{domain}/a/tickets/{tid}",
                ticket_data.get("subject", ""),
                STATUS_MAP.get(ticket_data.get("status", 0), "Unknown"),
                ticket_data.get("status", 0),
                PRIORITY_MAP.get(ticket_data.get("priority", 0), "Unknown"),
                ticket_data.get("priority", 0),
                str(ticket_data.get("group_id", "")),
                req_name, req_email,
                cf.get("cf_country", ""),
                created_at,
                ticket_data.get("updated_at", ""),
                now,
                analysis.get("classification", "other"),
                analysis.get("confidence", 0),
                "Yes" if analysis.get("needs_review") else "No",
                analysis.get("summary", ""),
                analysis.get("analysis", ""),
                draft, "Pending", now,
                analysis.get("risk_level", "medium"),
                ticket_data.get("description", ""),
                json.dumps(conversations),
                compiled,
                r_reach, r_impact, r_conf, r_effort, r_score,
                resolved_at, first_responded_at, template_name, workflow_name,
                sla_resolution_hours, sla_first_response_hours,
                json.dumps(screenshots or [], ensure_ascii=False),
            ))
            db.commit()

        ai_processed_count = 0  # Track how many AI calls we've made

        for i, t in enumerate(ticket_list):
            tid = t.get("id")
            try:
                _update_job(progress=f"Fetching ticket {tid} ({i+1}/{len(ticket_list)})...")

                # Check if we should skip (already analyzed and unchanged)
                existing = db.execute(
                    "SELECT last_analysis, updated_at, classification FROM tickets WHERE ticket_id = ?", (tid,)
                ).fetchone()

                ticket_updated = t.get("updated_at", "")

                if existing and existing["last_analysis"]:
                    # Skip AI re-analysis if already has a real analysis and ticket hasn't changed.
                    # BUT always re-fetch conversations so inbox detection stays accurate.
                    has_real_analysis = existing["classification"] not in ("", "pending", None)
                    skip_ai = False
                    try:
                        last_a = datetime.fromisoformat(existing["last_analysis"].replace("Z", "+00:00"))
                        tkt_u = datetime.fromisoformat(ticket_updated.replace("Z", "+00:00"))
                        if has_real_analysis and last_a >= tkt_u:
                            skip_ai = True
                    except (ValueError, TypeError):
                        pass

                    if skip_ai:
                        # Still refresh conversations + status for inbox detection
                        try:
                            _update_job(progress=f"Refreshing ticket {tid} ({i+1}/{len(ticket_list)})...")
                            tkt_detail, convs_fresh = get_ticket_details(api_key, domain, tid)
                            db.execute(
                                "UPDATE tickets SET raw_conversations = ?, status = ?, status_code = ?, updated_at = ? WHERE ticket_id = ?",
                                (json.dumps(convs_fresh), STATUS_MAP.get(tkt_detail.get("status", 0), "Unknown"),
                                 tkt_detail.get("status", 0), tkt_detail.get("updated_at", ""), tid)
                            )
                            db.commit()

                            # Also check for new screenshots on unchanged tickets
                            try:
                                existing_ss = db.execute("SELECT screenshots_json FROM tickets WHERE ticket_id = ?", (tid,)).fetchone()
                                has_screenshots = existing_ss and existing_ss["screenshots_json"] and existing_ss["screenshots_json"] != "[]"
                                if not has_screenshots:
                                    new_ss = extract_and_download_screenshots(tkt_detail, convs_fresh, api_key, domain, tid)
                                    if new_ss:
                                        db.execute("UPDATE tickets SET screenshots_json = ? WHERE ticket_id = ?",
                                                   (json.dumps(new_ss, ensure_ascii=False), tid))
                                        db.commit()
                                        logger.info(f"Ticket {tid}: backfilled {len(new_ss)} screenshot(s)")
                            except Exception:
                                pass

                            # Learn from direct Freshdesk replies on skipped tickets too
                            try:
                                _maybe_learn_from_freshdesk_reply(tid, convs_fresh, anthropic_key, domain, db=db)
                            except Exception:
                                pass
                        except Exception as refresh_err:
                            logger.warning(f"Conversation refresh failed for ticket {tid}: {refresh_err}")

                        _update_job(progress=f"Skipping AI for ticket {tid} (unchanged)")
                        _job_increment_processed()
                        continue

                # STEP 1: Always fetch and save the ticket from Freshdesk first
                ticket_data, conversations = get_ticket_details(api_key, domain, tid)
                compiled = compile_ticket_thread(ticket_data, conversations, domain)

                # Extract and download screenshots from the ticket
                try:
                    screenshots = extract_and_download_screenshots(
                        ticket_data, conversations, api_key, domain, tid
                    )
                    if screenshots:
                        logger.info(f"Ticket {tid}: downloaded {len(screenshots)} screenshot(s)")
                except Exception as ss_err:
                    screenshots = []
                    logger.warning(f"Screenshot extraction failed for ticket {tid}: {ss_err}")

                # Save with placeholder analysis (so ticket is always in DB even if AI fails)
                analysis = {"classification": "pending", "confidence": 0, "needs_review": True,
                            "summary": "Fetched — waiting for AI analysis",
                            "analysis": "", "risk_level": "medium"}
                draft = ""
                ai_rice = None

                # STEP 2: Try AI if enabled (may fail due to rate limits / no credits — that's OK)
                if ai_enabled:
                    try:
                        if ai_processed_count > 0:
                            time.sleep(10)
                        _update_job(progress=f"Analyzing ticket {tid} ({i+1}/{len(ticket_list)})...")
                        project_instr = get_setting("claude_project_instructions") or ""
                        term_context = get_terminology_context(db)
                        code_ctx = find_template_code(ticket_data.get("subject", ""), "", db=db)
                        client_ctx = get_setting("client_context") or ""

                        # ── Agent Pipeline: Parallel Prep → Main Analysis → QA with Retry ──
                        orchestrator = AgentOrchestrator(anthropic_key, db=db)

                        # Periodically clean up expired cache entries (every 10th ticket)
                        if i % 10 == 0:
                            orchestrator._cleanup_expired_cache()

                        # Pre-load KB index for batch optimization (reuses across tickets)
                        if i == 0 and kb_context:
                            orchestrator.preload_kb_index(kb_context, term_context)

                        # Search Jira for related issues (if Jira is configured)
                        jira_ctx = search_jira_for_ticket(
                            ticket_data.get("subject", ""),
                            template_name=ticket_data.get("template_name", ""),
                            db=db
                        )

                        # 2a. Run KB + Code + Research agents in PARALLEL
                        _update_job(progress=f"Agents analyzing ticket {tid} ({i+1}/{len(ticket_list)})...")
                        kb_brief, code_brief, research_brief = orchestrator.run_preparation_agents_parallel(
                            tid, ticket_data.get("subject", ""), compiled[:1000],
                            kb_context, code_ctx,
                            terminology_context=term_context,
                            template_name=ticket_data.get("template_name", ""),
                            workflow_name=ticket_data.get("workflow_name", ""),
                            jira_context=jira_ctx,
                        )

                        # 2b. Build enhanced context from all agent briefs
                        enhanced_kb = ""
                        if kb_brief:
                            enhanced_kb += f"\n\nKB AGENT BRIEF (pre-validated, targeted knowledge for this ticket):\n{kb_brief}\n"
                        if research_brief:
                            enhanced_kb += f"\n\nRESEARCH AGENT BRIEF (similar past tickets and lessons learned):\n{research_brief}\n"
                        enhanced_kb += f"\n\nRAW KNOWLEDGE BASE (backup reference):\n{truncate_text(kb_context, 2000)}" if kb_context else ""

                        # 2b-ii. LEARNING LOOP: Inject direct lessons from past PO corrections
                        # AND from direct Freshdesk replies the agent sent without using the tool.
                        try:
                            from agents import _find_relevant_lessons
                            t_name = ticket_data.get("template_name", "") or ""
                            direct_lessons = _find_relevant_lessons(db, ticket_data.get("subject", ""), t_name, limit=60)
                            if direct_lessons:
                                pinned = [l for l in direct_lessons if l.get("pinned")]
                                high = [l for l in direct_lessons
                                        if not l.get("pinned") and l.get("importance") == "high"]
                                rest = [l for l in direct_lessons if l not in pinned and l not in high]
                                lessons_text = (
                                    "\n\nLESSONS LEARNED — apply every one of these (PO corrections + direct Freshdesk replies). "
                                    "Tags: [category×hits|source]  source: PO=edited in tool, FD=direct Freshdesk reply.\n"
                                )
                                def _fmt(l):
                                    hits = l.get("hit_count", 1) or 1
                                    hit_mark = f"×{hits}" if hits > 1 else ""
                                    src = "FD" if l.get("source") == "freshdesk_reply" else "PO"
                                    return f"- [{l.get('category', 'general')}{hit_mark}|{src}] {l['lesson']}"
                                if pinned:
                                    lessons_text += "PINNED (always apply):\n" + "\n".join(_fmt(l) for l in pinned) + "\n"
                                if high:
                                    lessons_text += "HIGH-IMPORTANCE:\n" + "\n".join(_fmt(l) for l in high) + "\n"
                                if rest:
                                    lessons_text += "OTHER RELEVANT:\n" + "\n".join(_fmt(l) for l in rest) + "\n"
                                enhanced_kb += lessons_text
                        except Exception:
                            pass  # Non-critical: lessons are a bonus, not a blocker

                        # Use Code Agent brief instead of raw code for the main agent
                        # Sanitize to remove any code that Haiku may have slipped into the brief
                        effective_code_ctx = strip_code_from_output(code_brief) if code_brief else ""

                        # 2c. Main Analysis Agent (existing function, now with enriched context)
                        _update_job(progress=f"Analyzing ticket {tid} ({i+1}/{len(ticket_list)})...")
                        result = analyze_and_draft_ai(compiled, anthropic_key, writing_style, enhanced_kb, project_instr, term_context, effective_code_ctx, client_ctx)

                        # NOTE: QA is deferred to when the PO saves/reviews the ticket.

                        analysis = {k: result[k] for k in ("classification", "confidence", "needs_review",
                                                            "summary", "analysis", "risk_level", "template_name", "workflow_name")}

                        # If analysis failed (JSON parse error), save raw output for PO debugging
                        if result.get("_raw_output"):
                            try:
                                db.execute("UPDATE tickets SET analysis_raw_output = ? WHERE ticket_id = ?",
                                           (result["_raw_output"][:5000], tid))
                            except Exception:
                                pass  # Column might not exist yet on first run

                        # NOTE: Draft responses are NO LONGER saved during initial analysis.
                        # They are generated separately AFTER PO approval via the /generate-drafts route.
                        draft = ""
                        # Extract AI-suggested RICE values
                        ai_rice = {
                            "rice_reach": float(result.get("rice_reach", 0) or 0),
                            "rice_impact": float(result.get("rice_impact", 0) or 0),
                            "rice_confidence": float(result.get("rice_confidence", 0) or 0),
                            "rice_effort": float(result.get("rice_effort", 0) or 0),
                        }
                        ai_processed_count += 1
                    except Exception as ai_err:
                        logger.warning(f"AI failed for ticket {tid} (ticket still saved): {ai_err}")
                        _job_add_error(f"Ticket {tid}: AI failed — {str(ai_err)[:120]} (ticket data saved)")

                save_ticket_to_db(tid, ticket_data, conversations, compiled, analysis, draft, domain, db,
                                  rice=ai_rice, screenshots=screenshots)

                # ── Learn from direct Freshdesk replies (agent responded without using the tool) ──
                # If there's a new public agent reply since we last learned, compare it to
                # the stored draft_response and run the Learning Agent on it.
                try:
                    _maybe_learn_from_freshdesk_reply(tid, conversations, anthropic_key, domain, db=db)
                except Exception as fd_learn_err:
                    logger.warning(f"Freshdesk-reply learning skipped for ticket {tid}: {fd_learn_err}")

                _job_increment_processed()

            except Exception as e:
                logger.error(f"Failed fetching ticket {tid}: {e}")
                _job_add_error(f"Ticket {tid}: {str(e)[:200]}")

        # ── Inbox detection: group-based ──
        # Simple rule: if a ticket is currently in MY group → "In Inbox" (needs attention).
        # If it's NOT in my group (moved out, resolved, reassigned) → "Responded".
        # The search query already filters by group_id, so every ticket in ticket_list
        # is in our group = In Inbox. Tickets in DB but NOT in ticket_list have left our group.
        _update_job(progress="Updating inbox status (group-based)...")

        fetched_ids = {t.get("id") for t in ticket_list if t.get("id")}
        inbox_count = len(fetched_ids)

        # All tickets returned by our Freshdesk search are in our group → In Inbox
        for tid in fetched_ids:
            db.execute(
                "UPDATE tickets SET responded = 'In Inbox' WHERE ticket_id = ?",
                (tid,)
            )

        # All tickets in DB but NOT returned by search → no longer in our group → Responded
        all_db_tickets = [r["ticket_id"] for r in db.execute("SELECT ticket_id FROM tickets").fetchall()]
        stale_tids = [tid for tid in all_db_tickets if tid not in fetched_ids]
        responded_count = 0
        for stale_tid in stale_tids:
            db.execute("UPDATE tickets SET responded = 'Responded' WHERE ticket_id = ?", (stale_tid,))
            responded_count += 1

        db.commit()
        logger.info(
            f"Inbox detection (group-based): {inbox_count} in inbox, "
            f"{responded_count} responded/moved out."
        )

        snap = _get_job_status()
        _update_job(progress=f"Done! Processed {snap['processed']} of {snap['total']} tickets.")

    except Exception as e:
        logger.error(f"Job failed: {e}", exc_info=True)
        _job_add_error(str(e)[:500])
        _update_job(progress="Job failed. Check errors.")

    finally:
        _update_job(running=False)
        db.close()


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    """Main dashboard showing all tickets."""
    db = get_db()

    # Filters
    status_filter = request.args.get("status", "")
    classification_filter = request.args.get("classification", "")
    review_filter = request.args.get("review", "")
    risk_filter = request.args.get("risk", "")
    client_filter = request.args.get("client", "")
    requester_filter = request.args.get("requester", "")
    po_filter = request.args.get("po_decision", "")
    responded_filter = request.args.get("responded", "")
    search_q = request.args.get("q", "")
    sort_by = request.args.get("sort", "updated_at")
    sort_dir = request.args.get("dir", "DESC")

    if sort_dir not in ("ASC", "DESC"):
        sort_dir = "DESC"
    allowed_sorts = ["ticket_id", "subject", "status", "priority", "classification",
                     "confidence", "risk_level", "rice_score", "review_status", "updated_at",
                     "created_at", "po_decision", "responded"]
    if sort_by not in allowed_sorts:
        sort_by = "updated_at"

    query = "SELECT * FROM tickets WHERE 1=1"
    params = []

    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)
    if classification_filter:
        query += " AND classification = ?"
        params.append(classification_filter)
    if review_filter:
        query += " AND review_status = ?"
        params.append(review_filter)
    if risk_filter:
        query += " AND risk_level = ?"
        params.append(risk_filter)
    if client_filter:
        query += " AND LOWER(requester_email) LIKE ?"
        params.append(f"%@{client_filter.lower()}%")
    if requester_filter:
        query += " AND requester_name = ?"
        params.append(requester_filter)
    if po_filter:
        query += " AND LOWER(po_decision) = ?"
        params.append(po_filter.lower())
    if responded_filter:
        query += " AND responded = ?"
        params.append(responded_filter)
    if search_q:
        query += " AND (subject LIKE ? OR summary LIKE ? OR requester_name LIKE ? OR requester_email LIKE ? OR CAST(ticket_id AS TEXT) LIKE ?)"
        like = f"%{search_q}%"
        params.extend([like, like, like, like, like])

    query += f" ORDER BY {sort_by} {sort_dir}"
    tickets = db.execute(query, params).fetchall()

    # Stats
    total = db.execute("SELECT COUNT(*) as c FROM tickets").fetchone()["c"]
    pending = db.execute("SELECT COUNT(*) as c FROM tickets WHERE review_status = 'Pending'").fetchone()["c"]
    reviewed = db.execute("SELECT COUNT(*) as c FROM tickets WHERE review_status = 'Reviewed'").fetchone()["c"]
    sent = db.execute("SELECT COUNT(*) as c FROM tickets WHERE review_status = 'Sent'").fetchone()["c"]
    high_risk = db.execute("SELECT COUNT(*) as c FROM tickets WHERE risk_level IN ('high', 'critical')").fetchone()["c"]

    # Build distinct lists for client & requester dropdowns
    all_emails = db.execute("SELECT DISTINCT requester_email FROM tickets WHERE requester_email IS NOT NULL AND requester_email != ''").fetchall()
    company_set = {}
    for row in all_emails:
        email = row["requester_email"]
        comp = extract_company(email)
        domain = email.split("@")[1].split(".")[0].lower() if "@" in email else ""
        if comp != "Unknown" and domain:
            company_set[domain] = comp
    companies = sorted(company_set.items(), key=lambda x: x[1])  # list of (domain, display_name)

    all_requesters = db.execute(
        "SELECT DISTINCT requester_name FROM tickets WHERE requester_name IS NOT NULL AND requester_name != '' ORDER BY requester_name"
    ).fetchall()
    requesters = [r["requester_name"] for r in all_requesters]

    # Distinct statuses for filter dropdown
    all_statuses = db.execute(
        "SELECT DISTINCT status FROM tickets WHERE status IS NOT NULL AND status != '' ORDER BY status"
    ).fetchall()
    statuses = [s["status"] for s in all_statuses]

    # Period stats
    this_week_count = db.execute(
        "SELECT COUNT(*) as c FROM tickets WHERE created_at >= date('now', '-7 days')"
    ).fetchone()["c"]
    resolved_this_week_count = db.execute(
        "SELECT COUNT(*) as c FROM tickets WHERE resolved_at != '' AND resolved_at >= date('now', '-7 days')"
    ).fetchone()["c"]

    return render_template("dashboard.html",
        tickets=tickets, total=total, pending=pending, reviewed=reviewed,
        sent=sent, high_risk=high_risk, job_status=_get_job_status(),
        companies=companies, requesters=requesters, statuses=statuses,
        extract_company=extract_company,
        this_week_count=this_week_count,
        resolved_this_week_count=resolved_this_week_count,
        filters={"status": status_filter, "classification": classification_filter,
                 "review": review_filter, "risk": risk_filter, "q": search_q,
                 "client": client_filter, "requester": requester_filter,
                 "po_decision": po_filter, "responded": responded_filter,
                 "sort": sort_by, "dir": sort_dir},
    )


@app.route("/refresh-inbox", methods=["POST"])
def refresh_inbox():
    """Lightweight inbox refresh: re-search Freshdesk for tickets in our group and
    update inbox status. Tickets in our group = In Inbox, tickets not = Responded.
    No AI, no analysis — just a quick group membership check."""
    db = get_db()
    api_key = get_setting("freshdesk_api_key", db=db)
    domain = get_setting("freshdesk_domain", "silverfin.freshdesk.com", db=db)
    group_id = get_setting("freshdesk_group_id", "101000372179", db=db)
    country = get_setting("freshdesk_country", "", db=db)
    statuses = get_setting("freshdesk_statuses", "2,3,4,20", db=db)

    if not api_key:
        flash("Freshdesk API key not configured.", "error")
        return redirect(url_for("dashboard"))

    try:
        # Re-search Freshdesk for tickets currently in our group
        ticket_list = search_tickets(api_key, domain, group_id, country, statuses)
        fetched_ids = {t.get("id") for t in ticket_list if t.get("id")}

        # Build a lookup for ticket data from the API (for inserting new ones)
        ticket_data_map = {t.get("id"): t for t in ticket_list if t.get("id")}

        # Check which tickets already exist in DB
        existing_ids = {r["ticket_id"] for r in db.execute("SELECT ticket_id FROM tickets").fetchall()}

        # Tickets in our group → In Inbox (update existing ones)
        for tid in fetched_ids:
            if tid in existing_ids:
                db.execute("UPDATE tickets SET responded = 'In Inbox' WHERE ticket_id = ?", (tid,))
            else:
                # New ticket not yet in DB — insert a row with basic info from the API
                td = ticket_data_map.get(tid, {})
                requester = td.get("requester", {}) or {}
                if isinstance(requester, dict):
                    req_name = requester.get("name", "")
                    req_email = requester.get("email", "")
                else:
                    req_name = ""
                    req_email = ""
                db.execute("""
                    INSERT OR IGNORE INTO tickets
                    (ticket_id, ticket_url, subject, status, priority, requester_name, requester_email,
                     created_at, updated_at, responded, review_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'In Inbox', 'Pending')
                """, (
                    tid,
                    f"https://{domain}/a/tickets/{tid}",
                    td.get("subject", "(New ticket — run analysis to fetch details)"),
                    STATUS_MAP.get(td.get("status", 0), "Open"),
                    td.get("priority", 1),
                    req_name,
                    req_email,
                    td.get("created_at", ""),
                    td.get("updated_at", ""),
                ))

        # Tickets in DB but NOT in our group anymore → Responded
        stale_count = 0
        for tid in existing_ids:
            if tid not in fetched_ids:
                db.execute("UPDATE tickets SET responded = 'Responded' WHERE ticket_id = ?", (tid,))
                stale_count += 1

        db.commit()

        # Report the ACTUAL DB count so it matches the dashboard filter
        actual_inbox = db.execute("SELECT COUNT(*) as c FROM tickets WHERE responded = 'In Inbox'").fetchone()["c"]
        new_count = len(fetched_ids - existing_ids)
        new_msg = f" ({new_count} new)" if new_count > 0 else ""
        flash(f"Inbox refreshed: {actual_inbox} in inbox{new_msg}, {stale_count} responded/moved out.", "success")
    except Exception as e:
        flash(f"Inbox refresh failed: {str(e)[:200]}", "error")

    return redirect(url_for("dashboard"))


@app.route("/bulk-action", methods=["POST"])
def bulk_action():
    """Handle bulk actions on multiple tickets (approve, decline, mark reviewed)."""
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "No data provided"}), 400

    action = data.get("action", "")
    ticket_ids = data.get("ticket_ids", [])

    if not action or not ticket_ids:
        return jsonify({"success": False, "message": "Missing action or ticket_ids"}), 400

    if action not in ("approve", "decline", "review"):
        return jsonify({"success": False, "message": f"Unknown action: {action}"}), 400

    db = get_db()
    try:
        placeholders = ",".join("?" for _ in ticket_ids)
        if action == "approve":
            db.execute(
                f"UPDATE tickets SET po_decision = 'approved', review_status = 'Reviewed' WHERE ticket_id IN ({placeholders})",
                ticket_ids
            )
        elif action == "decline":
            db.execute(
                f"UPDATE tickets SET po_decision = 'declined', review_status = 'Reviewed' WHERE ticket_id IN ({placeholders})",
                ticket_ids
            )
        elif action == "review":
            db.execute(
                f"UPDATE tickets SET review_status = 'Reviewed' WHERE ticket_id IN ({placeholders})",
                ticket_ids
            )
        db.commit()
        return jsonify({"success": True, "message": f"Updated {len(ticket_ids)} ticket(s)"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/backfill-stats", methods=["POST"])
def backfill_stats():
    """Backfill SLA stats and template/workflow names for existing tickets using Claude AI."""
    db = get_db()
    api_key = get_setting("freshdesk_api_key", db=db)
    domain = get_setting("freshdesk_domain", db=db)
    anthropic_key = get_setting("anthropic_api_key", db=db)

    if not api_key or not domain:
        return jsonify({"success": False, "message": "Freshdesk API key or domain not configured"}), 400

    tickets = db.execute(
        "SELECT ticket_id, subject, summary, analysis, created_at, resolved_at, first_responded_at, template_name FROM tickets"
    ).fetchall()

    updated = 0
    errors = []

    for ticket in tickets:
        tid = ticket["ticket_id"]
        try:
            changes = {}

            # ── 1) Backfill SLA data from Freshdesk stats if missing ──
            if not ticket["resolved_at"] and not ticket["first_responded_at"]:
                try:
                    td = freshdesk_request(
                        "GET", f"/tickets/{tid}", api_key, domain,
                        params={"include": "stats"}
                    )
                    stats = td.get("stats", {}) or {}
                    resolved_at = stats.get("resolved_at", "") or ""
                    first_responded_at = stats.get("first_responded_at", "") or ""
                    created_at = td.get("created_at", ticket["created_at"]) or ""

                    sla_resolution_hours = 0
                    sla_first_response_hours = 0

                    if resolved_at and created_at:
                        try:
                            created = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                            resolved = datetime.fromisoformat(resolved_at.replace("Z", "+00:00"))
                            sla_resolution_hours = (resolved - created).total_seconds() / 3600
                        except (ValueError, TypeError):
                            pass

                    if first_responded_at and created_at:
                        try:
                            created = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                            responded = datetime.fromisoformat(first_responded_at.replace("Z", "+00:00"))
                            sla_first_response_hours = (responded - created).total_seconds() / 3600
                        except (ValueError, TypeError):
                            pass

                    changes["resolved_at"] = resolved_at
                    changes["first_responded_at"] = first_responded_at
                    changes["sla_resolution_hours"] = sla_resolution_hours
                    changes["sla_first_response_hours"] = sla_first_response_hours
                except Exception as e:
                    errors.append(f"#{tid} stats fetch: {str(e)[:80]}")

            # ── 2) Backfill template_name and workflow_name using Claude AI ──
            current_workflow = db.execute("SELECT workflow_name FROM tickets WHERE ticket_id = ?", (tid,)).fetchone()["workflow_name"]
            needs_template = not ticket["template_name"]
            needs_workflow = not current_workflow

            if (needs_template or needs_workflow) and anthropic_key:
                try:
                    context = f"Subject: {ticket['subject'] or ''}\nSummary: {ticket['summary'] or ''}\nAnalysis: {ticket['analysis'] or ''}"
                    prompt = f"""Based on this Silverfin support ticket, identify:
1. template_name: The exact Silverfin template/reconciliation name discussed (e.g. "Capital and Reserves", "Staff Cost", "Tangible Fixed Assets", "Financial Fixed Assets", "Receivables", "Payables", "Provisions", "Cash", "Turnover", "Profit & Loss", "Balance Sheet", "Tax Note", "Off Balance Sheet", "Accounting Policies", "Management Report", "Cover Page", "Intangible Fixed Assets", "Inventories", "Value Adjustments", "Related Parties", "Subsequent Events", "Corporate Tax Return", "Net Wealth Tax", "Municipal Business Tax"). Use the actual template name as known in Silverfin, not a generic description. If a section like dividends is part of "Capital and Reserves", use "Capital and Reserves".
2. workflow_name: Which workflow this belongs to — one of: "Annual Accounts", "Corporate Tax", "Legal Docs". If unclear, use "Annual Accounts" as default.

Reply ONLY with valid JSON: {{"template_name":"...", "workflow_name":"..."}}

Ticket:
{context[:2000]}"""

                    client = Anthropic(api_key=anthropic_key)
                    resp = call_anthropic_with_retry(
                        client, model="claude-sonnet-4-5", max_tokens=200,
                        system="You extract Silverfin template and workflow names from support tickets. Reply only with JSON.",
                        messages=[{"role": "user", "content": prompt}],
                    )
                    text = resp.content[0].text.strip()
                    if text.startswith("```"):
                        text = text.split("```")[1]
                        if text.startswith("json"):
                            text = text[4:]
                        text = text.strip()
                    extracted = json.loads(text)

                    if needs_template and extracted.get("template_name"):
                        changes["template_name"] = extracted["template_name"]
                    if needs_workflow and extracted.get("workflow_name"):
                        changes["workflow_name"] = extracted["workflow_name"]
                except Exception as e:
                    errors.append(f"#{tid} AI extract: {str(e)[:80]}")

            # ── Apply changes ──
            if changes:
                set_clause = ", ".join(f"{k} = ?" for k in changes.keys())
                values = list(changes.values()) + [tid]
                db.execute(f"UPDATE tickets SET {set_clause} WHERE ticket_id = ?", values)
                updated += 1

        except Exception as e:
            errors.append(f"#{tid}: {str(e)[:80]}")

    db.commit()
    return jsonify({
        "success": True,
        "message": f"Backfilled {updated} of {len(tickets)} tickets",
        "updated": updated,
        "total": len(tickets),
        "errors": errors
    })




def _resolve_image_placeholders(ticket_id, body_text, db):
    """Scan body_text for [IMAGE: filename] tokens. Returns:
      (cleaned_text, attachment_paths_list, rendered_html)
    cleaned_text: the body with [IMAGE: ...] tokens removed (visible to client)
    attachment_paths_list: local file paths of matching screenshots (to upload)
    rendered_html: HTML version of the body — tokens replaced by a small placeholder
                   line (e.g. "[Screenshot: filename.png — attached]") so the client
                   knows an image is attached. Freshdesk will show them as attachments.
    """
    import html as html_module
    token_pattern = re.compile(r"\[IMAGE:\s*([^\]]+?)\s*\]", re.IGNORECASE)
    tokens = token_pattern.findall(body_text or "")

    attachment_paths = []
    if tokens:
        try:
            row = db.execute("SELECT screenshots_json FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
            screenshots = json.loads(row["screenshots_json"] or "[]") if row else []
            fname_to_path = {s.get("filename"): s.get("path") for s in screenshots if s.get("filename")}
            for fname in tokens:
                p = fname_to_path.get(fname.strip())
                if p and os.path.exists(p) and p not in attachment_paths:
                    attachment_paths.append(p)
        except Exception as e:
            logger.warning(f"Image placeholder resolution failed on ticket {ticket_id}: {e}")

    # For the HTML body we replace each token with a small visual marker — Freshdesk
    # will append the attachment(s) below the message, but we want the reader to know
    # where each image was intended to appear.
    def _token_to_marker(m):
        fname = html_module.escape(m.group(1).strip())
        return f'<br><em style="color:#64748b;">[Screenshot attached: {fname}]</em><br>'
    body_for_html = token_pattern.sub(_token_to_marker, body_text or "")
    rendered_html = html_module.escape(body_for_html).replace("\n", "<br>")
    # Unescape our own markers so the HTML tags survive
    rendered_html = rendered_html.replace("&lt;br&gt;", "<br>") \
                                 .replace("&lt;em style=&#34;color:#64748b;&#34;&gt;", '<em style="color:#64748b;">') \
                                 .replace("&lt;/em&gt;", "</em>")

    # Strip tokens from the raw text for any caller that wants a clean version
    cleaned_text = token_pattern.sub("", body_text or "").replace("\n\n\n", "\n\n")

    return cleaned_text, attachment_paths, rendered_html


def _freshdesk_multipart_post(endpoint, api_key, domain, fields, files):
    """POST a multipart/form-data request to Freshdesk with attachments.
    fields: dict of str->str form fields
    files: list of (file_path) to upload under the 'attachments[]' key
    Returns the parsed JSON response."""
    url = f"https://{domain}/api/v2{endpoint}"
    auth = HTTPBasicAuth(api_key, "X")
    time.sleep(0.5)  # basic rate limiting
    multipart = []
    for k, v in fields.items():
        multipart.append((k, (None, str(v))))
    for p in files:
        try:
            fh = open(p, "rb")
            multipart.append(("attachments[]", (os.path.basename(p), fh, "application/octet-stream")))
        except Exception as e:
            logger.warning(f"Could not open attachment {p}: {e}")
    resp = requests.post(url, auth=auth, files=multipart)
    # Close file handles
    for _, tup in multipart:
        if isinstance(tup, tuple) and len(tup) >= 2 and hasattr(tup[1], "close"):
            try:
                tup[1].close()
            except Exception:
                pass
    resp.raise_for_status()
    return resp.json()


@app.route("/ticket/<int:ticket_id>/post-note", methods=["POST"])
def post_freshdesk_note(ticket_id):
    """Post a private note to the Freshdesk ticket via API.
    Supports [IMAGE: filename] tokens in the body — those screenshots are uploaded
    as attachments so the note carries the images.
    """
    db = get_db()
    api_key = get_setting("freshdesk_api_key", db=db)
    domain = get_setting("freshdesk_domain", db=db)

    if not api_key or not domain:
        return jsonify({"success": False, "message": "Freshdesk API not configured"}), 400

    data = request.get_json()
    if not data or not data.get("body"):
        return jsonify({"success": False, "message": "No content provided"}), 400

    body_text = data["body"]
    is_private = data.get("private", True)

    _, attachment_paths, body_html = _resolve_image_placeholders(ticket_id, body_text, db)

    try:
        if attachment_paths:
            result = _freshdesk_multipart_post(
                f"/tickets/{ticket_id}/notes", api_key, domain,
                fields={"body": body_html, "private": "true" if is_private else "false"},
                files=attachment_paths,
            )
        else:
            result = freshdesk_request(
                "POST", f"/tickets/{ticket_id}/notes", api_key, domain,
                json={"body": body_html, "private": is_private}
            )

        # After posting, re-check group membership to update inbox status
        try:
            my_group = get_setting("freshdesk_group_id", "101000372179", db=db)
            fresh_ticket, fresh_convs = get_ticket_details(api_key, domain, ticket_id)
            ticket_group = str(fresh_ticket.get("group_id", ""))
            new_status = "In Inbox" if ticket_group == str(my_group) else "Responded"
            db.execute(
                "UPDATE tickets SET responded = ?, raw_conversations = ? WHERE ticket_id = ?",
                (new_status, json.dumps(fresh_convs), ticket_id)
            )
            db.commit()
        except Exception:
            pass

        return jsonify({
            "success": True,
            "message": f"Note posted to ticket #{ticket_id}"
                       + (f" with {len(attachment_paths)} attachment(s)" if attachment_paths else ""),
            "note_id": result.get("id"),
            "attachments": len(attachment_paths),
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/ticket/<int:ticket_id>/reply-ticket", methods=["POST"])
def reply_freshdesk_ticket(ticket_id):
    """Post a public reply to the Freshdesk ticket via API.
    Supports [IMAGE: filename] tokens in the body — those screenshots are uploaded
    as attachments so the reply carries the images.
    """
    db = get_db()
    api_key = get_setting("freshdesk_api_key", db=db)
    domain = get_setting("freshdesk_domain", db=db)

    if not api_key or not domain:
        return jsonify({"success": False, "message": "Freshdesk API not configured"}), 400

    data = request.get_json()
    if not data or not data.get("body"):
        return jsonify({"success": False, "message": "No content provided"}), 400

    body_text = data["body"]

    _, attachment_paths, body_html = _resolve_image_placeholders(ticket_id, body_text, db)

    try:
        if attachment_paths:
            result = _freshdesk_multipart_post(
                f"/tickets/{ticket_id}/reply", api_key, domain,
                fields={"body": body_html},
                files=attachment_paths,
            )
        else:
            result = freshdesk_request(
                "POST", f"/tickets/{ticket_id}/reply", api_key, domain,
                json={"body": body_html}
            )

        # After replying, check if the ticket moved out of our group.
        # If it's still in our group → stays "In Inbox". If it left → "Responded".
        try:
            my_group = get_setting("freshdesk_group_id", "101000372179", db=db)
            fresh_ticket, fresh_convs = get_ticket_details(api_key, domain, ticket_id)
            ticket_group = str(fresh_ticket.get("group_id", ""))
            new_status = "In Inbox" if ticket_group == str(my_group) else "Responded"
            db.execute(
                "UPDATE tickets SET responded = ?, raw_conversations = ?, status = ?, status_code = ? WHERE ticket_id = ?",
                (new_status, json.dumps(fresh_convs),
                 STATUS_MAP.get(fresh_ticket.get("status", 0), "Unknown"),
                 fresh_ticket.get("status", 0), ticket_id)
            )
            db.commit()
        except Exception as conv_err:
            logger.warning(f"Post-reply refresh failed for {ticket_id}: {conv_err}")

        return jsonify({
            "success": True,
            "message": f"Reply posted to ticket #{ticket_id}"
                       + (f" with {len(attachment_paths)} attachment(s)" if attachment_paths else ""),
            "conversation_id": result.get("id"),
            "attachments": len(attachment_paths),
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/ticket/<int:ticket_id>/open-note", methods=["POST"])
def open_freshdesk_note(ticket_id):
    """Copy content to system clipboard, open Freshdesk ticket in the user's default browser,
    and use AppleScript (macOS) to click 'Add Note' and paste.
    Works with the user's existing browser session — no extra installs needed."""
    db = get_db()
    ticket = db.execute("SELECT ticket_url FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        return jsonify({"success": False, "message": "Ticket not found"}), 404

    data = request.get_json()
    body_text = data.get("body", "")
    if not body_text:
        return jsonify({"success": False, "message": "No content provided"}), 400

    ticket_url = ticket["ticket_url"]

    import platform
    import subprocess
    import threading

    system = platform.system()

    def run_automation():
        try:
            if system == "Darwin":
                # ── macOS: Use pbcopy for clipboard + AppleScript for browser control ──

                # 1. Copy content to macOS clipboard via pbcopy
                proc = subprocess.Popen(["pbcopy"], stdin=subprocess.PIPE)
                proc.communicate(body_text.encode("utf-8"))

                # 2. Open URL in default browser (Chrome/Safari/etc)
                subprocess.run(["open", ticket_url])

                # 3. Wait for Freshdesk page to load, then:
                #    - Press "n" (Freshdesk shortcut to open Note editor)
                #    - Wait for editor to appear
                #    - Cmd+V to paste from clipboard
                time.sleep(3)  # Wait for Freshdesk page to load

                applescript = '''
                tell application "Google Chrome"
                    activate
                    delay 0.5
                    tell application "System Events"
                        tell process "Google Chrome"
                            key code 53
                            delay 0.3
                        end tell
                        keystroke "n"
                        delay 1.5
                        keystroke "v" using command down
                    end tell
                end tell
                '''
                subprocess.run(["osascript", "-e", applescript], timeout=20)

            elif system == "Windows":
                # ── Windows: Use clip.exe + PowerShell for browser control ──
                proc = subprocess.Popen(["clip.exe"], stdin=subprocess.PIPE)
                proc.communicate(body_text.encode("utf-8"))
                # Open URL in default browser
                subprocess.run(["start", ticket_url], shell=True)

            else:
                # ── Linux: Use xclip + xdg-open ──
                try:
                    proc = subprocess.Popen(["xclip", "-selection", "clipboard"], stdin=subprocess.PIPE)
                    proc.communicate(body_text.encode("utf-8"))
                except FileNotFoundError:
                    proc = subprocess.Popen(["xsel", "--clipboard", "--input"], stdin=subprocess.PIPE)
                    proc.communicate(body_text.encode("utf-8"))
                subprocess.run(["xdg-open", ticket_url])

        except Exception as e:
            logger.error(f"Browser automation error: {e}")

    thread = threading.Thread(target=run_automation, daemon=True)
    thread.start()

    if system == "Darwin":
        msg = "Opening Freshdesk in Chrome — clicking 'Add Note' and pasting your content. Review, tag people, and send when ready."
    else:
        msg = "Content copied to clipboard and ticket opened. Click 'Add Note' in Freshdesk and paste (Ctrl+V)."

    return jsonify({"success": True, "message": msg})


@app.route("/ticket/<int:ticket_id>")
def ticket_detail(ticket_id):
    """View single ticket details."""
    db = get_db()
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    # Sanitize AI-generated text fields before displaying to strip any leaked code
    ticket_dict = dict(ticket)
    for field in ("analysis", "summary", "draft_response", "draft_response_en"):
        if ticket_dict.get(field):
            ticket_dict[field] = strip_code_from_output(ticket_dict[field])

    # Ensure ticket_url is always populated (backfill if missing)
    if not ticket_dict.get("ticket_url"):
        domain = get_setting("freshdesk_domain", "silverfin.freshdesk.com", db=db)
        ticket_dict["ticket_url"] = f"https://{domain}/a/tickets/{ticket_id}"
        db.execute("UPDATE tickets SET ticket_url = ? WHERE ticket_id = ?",
                   (ticket_dict["ticket_url"], ticket_id))
        db.commit()

    return render_template("ticket.html", ticket=ticket_dict)


@app.route("/ticket/<int:ticket_id>/update", methods=["POST"])
def update_ticket(ticket_id):
    """Update review status, next steps, etc."""
    db = get_db()

    # Fetch current ticket BEFORE update (for Learning Agent comparison)
    ticket_before = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()

    review_status = request.form.get("review_status", "")
    next_steps = request.form.get("next_steps", "")
    draft_response = request.form.get("draft_response", "")
    draft_response_en = request.form.get("draft_response_en", "")

    # RICE scoring
    try:
        rice_reach = float(request.form.get("rice_reach", 0) or 0)
        rice_impact = float(request.form.get("rice_impact", 0) or 0)
        rice_confidence = float(request.form.get("rice_confidence", 0) or 0)
        rice_effort = float(request.form.get("rice_effort", 0) or 0)
        rice_score = (rice_reach * rice_impact * rice_confidence / rice_effort) if rice_effort > 0 else 0
        rice_score = round(rice_score, 1)
    except (ValueError, TypeError, ZeroDivisionError):
        rice_reach = rice_impact = rice_confidence = rice_effort = rice_score = 0

    db.execute(
        """UPDATE tickets SET review_status = ?, next_steps = ?, draft_response = ?, draft_response_en = ?,
           rice_reach = ?, rice_impact = ?, rice_confidence = ?, rice_effort = ?, rice_score = ?
           WHERE ticket_id = ?""",
        (review_status, next_steps, draft_response, draft_response_en,
         rice_reach, rice_impact, rice_confidence, rice_effort, rice_score, ticket_id),
    )
    db.commit()

    # ── Learning Agent: extract lessons from PO's edits (runs in background) ──
    if ticket_before:
        old_draft = ticket_before["draft_response"] or ""
        new_draft = draft_response or ""
        # Only trigger learning if the draft was actually changed (not just status update)
        if old_draft and new_draft and old_draft.strip() != new_draft.strip():
            try:
                anthropic_key = get_setting("anthropic_api_key", db=db)
                if anthropic_key:
                    # Capture values for the background thread (Flask db may close)
                    _api_key = anthropic_key
                    _tid = ticket_id
                    _subj = ticket_before["subject"] or ""
                    _tpl = ticket_before["template_name"] if "template_name" in ticket_before.keys() else ""
                    _wf = ticket_before["workflow_name"] if "workflow_name" in ticket_before.keys() else ""
                    _old = old_draft
                    _new = new_draft
                    # Run in background thread with its own DB connection
                    def _learn():
                        try:
                            bg_db = get_db_standalone()
                            orch = AgentOrchestrator(_api_key, db=bg_db)
                            orch.run_learning(_tid, _subj, _tpl, _wf, _old, _new, "draft_response")
                            bg_db.close()
                        except Exception as le:
                            logger.warning(f"Learning Agent failed for ticket {_tid}: {le}")
                    threading.Thread(target=_learn, daemon=True).start()
                    logger.info(f"Learning Agent triggered for ticket {ticket_id} (PO edited draft)")
            except Exception as e:
                logger.warning(f"Could not trigger Learning Agent: {e}")

    # ── QA Agent: validate the saved draft (runs in background, non-blocking) ──
    # QA only activates when the PO explicitly TRANSITIONS the status to reviewed/approved/final
    # in this save action. It does NOT run on re-saves where the status was already reviewed
    # (to avoid repeated QA alerts on every minor edit after initial review).
    previous_status = ticket_before["review_status"] if ticket_before else ""
    status_just_changed_to_review = (
        review_status in ("reviewed", "approved", "final")
        and previous_status != review_status
    )
    if draft_response and status_just_changed_to_review:
        try:
            anthropic_key = get_setting("anthropic_api_key", db=db)
            if anthropic_key:
                _qa_key = anthropic_key
                _qa_tid = ticket_id
                _qa_draft = draft_response
                _qa_subj = ticket_before["subject"] if ticket_before else ""
                _qa_kb = ""
                def _run_qa():
                    try:
                        bg_db = get_db_standalone()
                        orch = AgentOrchestrator(_qa_key, db=bg_db)
                        qa_result, _ = orch.run_qa_with_retry(
                            _qa_tid, _qa_draft, "draft_response", _qa_subj, _qa_kb
                        )
                        if qa_result.get("critical_issues"):
                            # Append QA alert to the draft so PO sees it on next page load
                            qa_note = "\n\n[QA ALERT: " + " | ".join(qa_result["critical_issues"][:3]) + "]"
                            bg_db.execute(
                                "UPDATE tickets SET draft_response = draft_response || ? WHERE ticket_id = ?",
                                (qa_note, _qa_tid)
                            )
                            bg_db.commit()
                            logger.info(f"QA found issues for ticket {_qa_tid}: {qa_result['critical_issues']}")
                        bg_db.close()
                    except Exception as qe:
                        logger.warning(f"QA Agent failed for ticket {_qa_tid}: {qe}")
                threading.Thread(target=_run_qa, daemon=True).start()
        except Exception:
            pass

    flash("Ticket updated.", "success")
    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/po-decision", methods=["POST"])
def po_decision(ticket_id):
    """Record PO Approve/Decline decision for a ticket."""
    db = get_db()
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    decision = request.form.get("decision", "").strip().lower()
    reason = request.form.get("reason", "").strip()
    now = datetime.now(timezone.utc).isoformat()

    if decision not in ("approved", "declined"):
        flash("Invalid decision.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    # Clear old drafts when the decision changes — prevents stale approved drafts
    # from showing up when the ticket is now declined, and vice versa.
    old_decision = ticket["po_decision"] if ticket["po_decision"] else "pending"
    if decision != old_decision:
        db.execute(
            """UPDATE tickets SET po_decision = ?, po_decision_reason = ?, po_decision_date = ?,
               draft_response = '', draft_response_en = ''
               WHERE ticket_id = ?""",
            (decision, reason, now, ticket_id),
        )
    else:
        db.execute(
            """UPDATE tickets SET po_decision = ?, po_decision_reason = ?, po_decision_date = ?
               WHERE ticket_id = ?""",
            (decision, reason, now, ticket_id),
        )
    db.commit()

    if decision == "approved":
        flash("Ticket approved. You can now generate draft responses.", "success")
    else:
        flash("Ticket declined. You can now generate a decline response to explain the decision to the client.", "info")

    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/generate-drafts", methods=["POST"])
def generate_drafts(ticket_id):
    """Generate draft responses (FR + EN) for an approved ticket using AI."""
    db = get_db()
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    if (ticket["po_decision"] or "").lower() not in ("approved", "declined"):
        flash("Ticket must be approved or declined by PO before generating draft responses.", "warning")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    anthropic_key = get_setting("anthropic_api_key", db=db)
    if not anthropic_key:
        flash("Anthropic API key not configured. Go to Settings.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    compiled = ticket["compiled_thread"] or ""
    if not compiled:
        flash("No ticket thread data available. Run a Freshdesk fetch first.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    writing_style = get_setting("writing_style", "customer_support", db=db)
    kb_context = get_knowledge_base_context(db)
    project_instr = get_setting("claude_project_instructions", db=db) or ""
    term_context = get_terminology_context(db)
    code_ctx = find_template_code(ticket["subject"], ticket["analysis"], db=db)

    try:
        # ── Agent Pipeline: Parallel Prep → Main Draft → QA with Retry ──
        orchestrator = AgentOrchestrator(anthropic_key, db=db)

        # Search Jira for related issues
        jira_ctx = search_jira_for_ticket(ticket["subject"], _row_get(ticket, "template_name", ""), db=db)

        # 1. Run KB + Code + Research agents in PARALLEL
        kb_brief, code_brief, research_brief = orchestrator.run_preparation_agents_parallel(
            ticket_id, ticket["subject"], (ticket["analysis"] or "")[:1000],
            kb_context, code_ctx,
            terminology_context=term_context,
            template_name=_row_get(ticket, "template_name", ""),
            workflow_name=_row_get(ticket, "workflow_name", ""),
            jira_context=jira_ctx,
        )

        # 2. Build enhanced context from all agent briefs
        # NOTE: Agent briefs are in English. The language enforcement in the prompt
        # instructs the model to translate — never copy English verbatim into French output.
        enhanced_kb = "\n[CONTEXT DATA BELOW IS FOR REFERENCE ONLY — DO NOT COPY TEXT FROM IT. TRANSLATE TO THE DRAFT LANGUAGE.]\n"
        if kb_brief:
            enhanced_kb += f"\nKB AGENT BRIEF (pre-validated knowledge for this ticket):\n{kb_brief}\n"
        if research_brief:
            enhanced_kb += f"\nRESEARCH AGENT BRIEF (similar past tickets and lessons):\n{research_brief}\n"
        enhanced_kb += f"\nRAW KNOWLEDGE BASE (backup):\n{truncate_text(kb_context, 2000)}" if kb_context else ""

        # 2b. LEARNING LOOP: Inject top lessons directly from past PO corrections.
        # These are concrete rules the PO established by editing previous drafts.
        # They go DIRECTLY into the prompt so the model follows them without relying
        # on the Research Agent's summary (which may dilute or miss specific lessons).
        try:
            template_name = _row_get(ticket, "template_name", "") or ""
            from agents import _find_relevant_lessons
            # No tight cap — durable rules (pinned + high importance + all for this
            # template) are always injected; the limit only bounds the keyword fill tier.
            direct_lessons = _find_relevant_lessons(db, ticket["subject"], template_name, limit=60)

            if direct_lessons:
                # Split into tiers so the model clearly sees the non-negotiable rules
                pinned = [l for l in direct_lessons if l.get("pinned")]
                high = [l for l in direct_lessons
                        if not l.get("pinned") and l.get("importance") == "high"]
                template_specific = [l for l in direct_lessons
                                     if not l.get("pinned")
                                     and l.get("importance") != "high"
                                     and (l.get("template_name") or "") == template_name
                                     and template_name]
                other = [l for l in direct_lessons
                         if l not in pinned and l not in high and l not in template_specific]

                def fmt_lesson(l):
                    hits = l.get("hit_count", 1) or 1
                    reinforce = f" ×{hits}" if hits > 1 else ""
                    src = l.get("source", "po_edit")
                    src_tag = "FD" if src == "freshdesk_reply" else "PO"
                    return f"- [{l.get('category', 'general')}{reinforce}|{src_tag}] {l['lesson']}"

                blocks = []
                if pinned:
                    blocks.append(
                        "PINNED RULES — ALWAYS APPLY (manually curated):\n"
                        + "\n".join(fmt_lesson(l) for l in pinned)
                    )
                if high:
                    blocks.append(
                        "HIGH-IMPORTANCE LESSONS — VIOLATING THESE IS A CRITICAL ERROR:\n"
                        + "\n".join(fmt_lesson(l) for l in high)
                    )
                if template_specific:
                    blocks.append(
                        f"TEMPLATE-SPECIFIC LESSONS ({template_name}) — follow every one:\n"
                        + "\n".join(fmt_lesson(l) for l in template_specific)
                    )
                if other:
                    blocks.append(
                        "OTHER RELEVANT LESSONS (historical context):\n"
                        + "\n".join(fmt_lesson(l) for l in other)
                    )

                lessons_text = (
                    "\nLESSONS FROM PAST PO REVIEWS + DIRECT FRESHDESK REPLIES\n"
                    "These are corrections made on previous tickets. "
                    "Source tags: PO = edited in the tool, FD = learned from your direct Freshdesk reply. "
                    "Hit counts (×N) show how often the same lesson was reinforced — treat those as rules.\n\n"
                    + "\n\n".join(blocks) + "\n"
                )
                enhanced_kb += lessons_text
                logger.info(
                    f"Ticket {ticket_id}: injected {len(direct_lessons)} lessons "
                    f"(pinned={len(pinned)}, high={len(high)}, "
                    f"template={len(template_specific)}, other={len(other)}) into draft generation"
                )
        except Exception as lesson_err:
            logger.warning(f"Could not inject lessons for ticket {ticket_id}: {lesson_err}")

        # Use Code Agent brief instead of raw code
        # Sanitize to remove any code Haiku may have slipped into its "plain language" brief
        effective_code_ctx = strip_code_from_output(code_brief) if code_brief else ""

        # 3. Generate FR draft (Sonnet — full reasoning)
        result_fr = generate_draft_response(compiled, anthropic_key, "fr", writing_style, enhanced_kb, project_instr,
                                            ticket["analysis"] or "", ticket["po_decision_reason"] or "", effective_code_ctx,
                                            terminology_context=term_context, ticket_id=ticket_id,
                                            classification=_row_get(ticket, "classification", ""),
                                            priority=str(_row_get(ticket, "priority", "")))

        # 4. Translate FR → EN (Haiku — fast, consistent, same content)
        result_en = translate_draft(result_fr, "fr", "en", anthropic_key)

        # Lightweight complexity check: if the analysis says it's a simple fix
        # (translation, wording, label) but the draft is excessively long, regenerate
        # with a stronger simplicity instruction.
        analysis_text = (ticket["analysis"] or "").lower()
        subject_text = (ticket["subject"] or "").lower()
        combined_text = f"{analysis_text} {subject_text}"
        simple_indicators = [
            # Translation / wording / labels
            "translation", "traduction", "traduire", "translate",
            "wording", "formulation", "reformuler", "rephrase",
            "label", "libellé", "libelle", "intitulé", "intitule",
            "typo", "faute de frappe", "spelling", "orthographe",
            "rename", "renommer", "replace", "remplacer",
            "change the text", "changer le texte", "change the label", "changer le libellé",
            "add the word", "ajouter le mot", "missing word", "mot manquant",
            "small fix", "minor fix", "quick fix", "petite correction", "correction mineure",
            "text fix", "correction de texte", "text change", "changement de texte",
            "update wording", "mettre à jour la formulation",
            "update label", "mettre à jour le libellé",
            # Simple display/formatting tweaks often flagged as simple
            "punctuation", "ponctuation", "capitalization", "majuscule",
            "accent", "accentuation",
        ]
        is_likely_simple = any(ind in combined_text for ind in simple_indicators)
        draft_word_count = len(result_fr.split())

        # Also force simple regeneration if the draft is absurdly long for what appears simple,
        # OR if ANY simple indicator is present AND the draft exceeds 250 words.
        if is_likely_simple and draft_word_count > 250:
            # Draft is too long for a simple fix — regenerate with strict brevity
            logger.info(f"Ticket {ticket_id}: simple fix detected but draft is {draft_word_count} words. Regenerating shorter version.")
            result_fr = generate_draft_response(
                compiled, anthropic_key, "fr", writing_style, enhanced_kb, project_instr,
                ticket["analysis"] or "", ticket["po_decision_reason"] or "", effective_code_ctx,
                terminology_context=term_context, ticket_id=ticket_id,
                force_simple=True,
                classification=_row_get(ticket, "classification", ""),
                priority=str(_row_get(ticket, "priority", ""))
            )
            result_en = translate_draft(result_fr, "fr", "en", anthropic_key)

        # 5. Validate translation: reverse-translate EN→FR and check for legal term drift
        result_en = validate_translation(result_fr, result_en, anthropic_key)

        # 6. Mandatory QA on generated draft — runs inline (not deferred to PO review)
        #    Uses a background thread so the user isn't blocked, but results are saved immediately.
        _qa_key = anthropic_key
        _qa_tid = ticket_id
        _qa_draft_fr = result_fr
        _qa_subj = ticket["subject"]
        _qa_kb = kb_brief if kb_brief else ""
        def _run_post_gen_qa():
            try:
                bg_db = get_db_standalone()
                orch = AgentOrchestrator(_qa_key, db=bg_db)
                qa_result, _ = orch.run_qa_with_retry(
                    _qa_tid, _qa_draft_fr, "draft_response", _qa_subj, _qa_kb
                )
                # Store QA result as a flag on the ticket so PO sees it on first load
                qa_status = "passed" if qa_result.get("passed") else "needs_review"
                qa_issues_json = json.dumps(qa_result.get("critical_issues", [])[:5])
                bg_db.execute(
                    "UPDATE tickets SET qa_status = ?, qa_issues = ? WHERE ticket_id = ?",
                    (qa_status, qa_issues_json, _qa_tid)
                )
                bg_db.commit()
                logger.info(f"Post-generation QA for ticket {_qa_tid}: {qa_status} (score={qa_result.get('score', 0)})")
                bg_db.close()
            except Exception as qe:
                logger.warning(f"Post-generation QA failed for ticket {_qa_tid}: {qe}")
        threading.Thread(target=_run_post_gen_qa, daemon=True).start()

        db.execute(
            "UPDATE tickets SET draft_response = ?, draft_response_en = ? WHERE ticket_id = ?",
            (result_fr, result_en, ticket_id),
        )
        db.commit()
        flash("Draft responses generated (FR + EN). QA validation running in background.", "success")

    except Exception as e:
        import traceback
        err_msg = str(e)[:300]
        err_trace = traceback.format_exc()[-500:]
        logger.error(f"Draft generation failed for ticket {ticket_id}: {e}\n{err_trace}")
        flash(f"Draft generation failed: {err_msg}", "error")
        # Also save the error in the draft field so the user can always see it
        # (flash messages can be lost if the page redirect doesn't preserve them)
        try:
            db.execute(
                "UPDATE tickets SET draft_response = ? WHERE ticket_id = ? AND (draft_response IS NULL OR draft_response = '')",
                (f"[GENERATION ERROR] {err_msg}\n\nFull trace:\n{err_trace}", ticket_id)
            )
            db.commit()
        except Exception:
            pass

    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/generate-decline-drafts", methods=["POST"])
def generate_decline_drafts(ticket_id):
    """Generate decline response (FR + EN) for a declined ticket using AI."""
    db = get_db()
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    if (ticket["po_decision"] or "").lower() != "declined":
        flash("Ticket must be declined by PO to generate a decline response.", "warning")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    anthropic_key = get_setting("anthropic_api_key", db=db)
    if not anthropic_key:
        flash("Anthropic API key not configured. Go to Settings.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    compiled = ticket["compiled_thread"] or ""
    if not compiled:
        flash("No ticket thread data available. Run a Freshdesk fetch first.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    writing_style = get_setting("writing_style", "customer_support", db=db)
    kb_context = get_knowledge_base_context(db)
    project_instr = get_setting("claude_project_instructions", db=db) or ""
    term_context = get_terminology_context(db)
    code_ctx = find_template_code(ticket["subject"], ticket["analysis"], db=db)

    # Run agents to get focused KB and code briefs (same as draft generation)
    # This ensures the decline response is based on verified, relevant information
    try:
        from agents import AgentOrchestrator
        orch = AgentOrchestrator(anthropic_key, db=db)

        # Run Code Agent for a functional brief (plain language, no raw code)
        code_brief = ""
        if code_ctx:
            code_brief = orch.get_code_brief(
                ticket_id, ticket["subject"] or "",
                ticket["summary"] or ticket["subject"] or "",
                code_ctx
            )

        # Run KB Agent for focused, relevant KB context
        kb_brief = ""
        if kb_context:
            kb_brief = orch.get_kb_brief(
                ticket_id, ticket["subject"] or "",
                ticket["summary"] or ticket["subject"] or "",
                kb_context, term_context,
                code_brief[:500] if code_brief else ""
            )

        # Build enhanced KB context from agent briefs
        enhanced_kb = ""
        if kb_brief:
            enhanced_kb += f"\nKB AGENT BRIEF (pre-validated, relevant knowledge for THIS ticket):\n{kb_brief}\n"
        enhanced_kb += f"\nRAW KNOWLEDGE BASE (backup — check this if KB Agent missed something):\n{truncate_text(kb_context, 3000)}" if kb_context else ""

        # Use Code Agent brief instead of raw code
        effective_code_ctx = strip_code_from_output(code_brief) if code_brief else code_ctx

        logger.info(f"Decline ticket {ticket_id}: ran agents — KB brief {len(kb_brief)} chars, code brief {len(code_brief)} chars")
    except Exception as agent_err:
        logger.warning(f"Agent pre-processing failed for decline ticket {ticket_id}: {agent_err}")
        enhanced_kb = kb_context
        effective_code_ctx = code_ctx

    try:
        # Generate FR decline response (Sonnet — full reasoning)
        result_fr = generate_decline_response(compiled, anthropic_key, "fr", writing_style, enhanced_kb, project_instr,
                                              ticket["analysis"] or "", ticket["po_decision_reason"] or "", effective_code_ctx,
                                              terminology_context=term_context, ticket_id=ticket_id)

        # Translate FR → EN (Haiku — fast, consistent, same content)
        result_en = translate_draft(result_fr, "fr", "en", anthropic_key)

        db.execute(
            "UPDATE tickets SET draft_response = ?, draft_response_en = ? WHERE ticket_id = ?",
            (result_fr, result_en, ticket_id),
        )
        db.commit()
        flash("Decline responses generated (FR + EN translation).", "success")

    except Exception as e:
        import traceback
        err_msg = str(e)[:300]
        err_trace = traceback.format_exc()[-500:]
        logger.error(f"Decline draft generation failed for ticket {ticket_id}: {e}\n{err_trace}")
        flash(f"Decline draft generation failed: {err_msg}", "error")
        try:
            db.execute(
                "UPDATE tickets SET draft_response = ? WHERE ticket_id = ? AND (draft_response IS NULL OR draft_response = '')",
                (f"[GENERATION ERROR] {err_msg}\n\nFull trace:\n{err_trace}", ticket_id)
            )
            db.commit()
        except Exception:
            pass

    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/create-jira", methods=["POST"])
def create_jira_ticket(ticket_id):
    """Create a Jira ticket from the Freshdesk ticket analysis and link them.
    Accepts JSON body with overrides or form POST for simple creation."""
    db = get_db()
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        if request.is_json:
            return jsonify({"ok": False, "message": "Ticket not found"})
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    existing_key = _row_get(ticket, "jira_ticket_key", "")
    if existing_key:
        if request.is_json:
            return jsonify({"ok": False, "message": f"Already linked to {existing_key}"})
        flash(f"Already linked to Jira {existing_key}. Unlink first to create a new one.", "warning")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    # Parse overrides from JSON body
    overrides = {}
    if request.is_json:
        data = request.get_json(silent=True) or {}
        overrides = {
            "issue_type": data.get("issue_type", ""),
            "priority": data.get("priority", ""),
            "epic_key": data.get("epic_key", ""),
            "parent_key": data.get("parent_key", ""),
            "components": data.get("components", []),
            "labels": data.get("labels", []),
            "summary": data.get("summary", ""),
        }

    try:
        ticket_dict = dict(ticket)
        jira_key, jira_url = create_jira_ticket_from_freshdesk(ticket_dict, db=db, overrides=overrides)
        if request.is_json:
            return jsonify({"ok": True, "key": jira_key, "url": jira_url, "message": f"Created {jira_key}"})
        flash(f"Jira ticket {jira_key} created and linked.", "success")
    except Exception as e:
        logger.error(f"Jira creation failed for ticket {ticket_id}: {e}")
        if request.is_json:
            return jsonify({"ok": False, "message": str(e)[:300]})
        flash(f"Jira ticket creation failed: {str(e)[:200]}", "error")

    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/link-jira", methods=["POST"])
def link_jira_ticket(ticket_id):
    """Manually link an existing Jira ticket to a Freshdesk ticket."""
    db = get_db()
    jira_key = request.form.get("jira_key", "").strip().upper()
    if not jira_key:
        flash("Please enter a Jira ticket key (e.g. LUX-123).", "warning")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    jira_domain = get_setting("jira_domain", "", db=db)
    if not jira_domain:
        flash("Jira domain not configured. Go to Settings.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    jira_url = f"https://{jira_domain}/browse/{jira_key}"

    # Verify the ticket exists by trying to fetch it
    jira_email = get_setting("jira_email", "", db=db)
    jira_token = get_setting("jira_api_token", "", db=db)
    if jira_email and jira_token:
        try:
            _jira_request("GET", f"/issue/{jira_key}", jira_domain, jira_email, jira_token,
                          params={"fields": "summary"})
        except Exception as e:
            flash(f"Could not verify Jira ticket {jira_key}: {str(e)[:100]}", "warning")

        # Add Freshdesk as remote link on Jira
        ticket = db.execute("SELECT ticket_url, subject FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
        if ticket:
            fd_url = ticket["ticket_url"] or f"https://silverfin.freshdesk.com/a/tickets/{ticket_id}"
            try:
                _jira_request(
                    "POST", f"/issue/{jira_key}/remotelink",
                    jira_domain, jira_email, jira_token,
                    json_data={
                        "globalId": f"freshdesk={ticket_id}",
                        "application": {"type": "com.freshdesk", "name": "Freshdesk"},
                        "object": {
                            "url": fd_url,
                            "title": f"Freshdesk #{ticket_id}: {(ticket['subject'] or '')[:100]}",
                        }
                    }
                )
            except Exception:
                pass

    db.execute(
        "UPDATE tickets SET jira_ticket_key = ?, jira_ticket_url = ? WHERE ticket_id = ?",
        (jira_key, jira_url, ticket_id)
    )
    db.commit()
    flash(f"Linked to Jira {jira_key}.", "success")
    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/ai-chat", methods=["POST"])
def ai_chat(ticket_id):
    """AI chat endpoint: user can ask questions or request draft changes.
    Changes are always applied directly to the draft sections — never returned as chat text."""
    db = get_db()
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        return jsonify({"error": "Ticket not found."}), 404

    anthropic_key = get_setting("anthropic_api_key", db=db)
    if not anthropic_key:
        return jsonify({"error": "Anthropic API key not configured. Go to Settings."}), 400

    data = request.get_json()
    if not data or not data.get("message"):
        return jsonify({"error": "No message provided."}), 400

    user_message = (data.get("message") or "").strip()
    current_draft = data.get("current_draft", "")
    lang = data.get("lang", "fr")
    pinned_screenshots = data.get("pinned_screenshots") or []  # list of filenames
    attached_images = data.get("attached_images") or []  # [{name, data_b64, media_type}]

    # Allow image-only messages (user pinned/attached images but typed nothing)
    if not user_message and not pinned_screenshots and not attached_images:
        return jsonify({"error": "No message or images provided."}), 400
    if not user_message:
        user_message = "Use the attached images above to inform your next edit of the draft."

    lang_instruction = (
        "LANGUAGE RULE — MANDATORY: Respond in FRENCH. Every word must be in French. Do NOT copy English from context data — translate everything."
        if lang == "fr"
        else "LANGUAGE RULE — MANDATORY: Respond in ENGLISH. Every word must be in English. Do NOT copy French from context data — translate everything."
    )

    # If there is a current draft, almost always treat it as a rewrite request
    # Only pure questions (starting with "what", "why", "how", "is", "does", "qu'est", "pourquoi", "comment", "est-ce")
    # without change-intent words are treated as questions
    question_only = False
    msg_lower = user_message.lower().strip()
    question_starters = ["what ", "why ", "how ", "is ", "does ", "can ", "where ",
                         "qu'est", "pourquoi ", "comment ", "est-ce ", "où ", "quel "]
    change_words = ["change", "rewrite", "reword", "modify", "update", "replace", "adjust", "shorten",
                    "make", "add", "remove", "less", "more", "exact", "wording", "specific", "vague",
                    "provide", "include", "should", "need", "must", "too", "not enough", "missing",
                    "changer", "réécrire", "modifier", "raccourcir", "ajouter", "supprimer",
                    "reformuler", "préciser", "manque", "trop", "pas assez", "exact", "devrait",
                    "the internal", "la note", "bso", "client response", "backlog"]
    if any(msg_lower.startswith(q) for q in question_starters) and not any(cw in msg_lower for cw in change_words):
        question_only = True

    compiled = ticket["compiled_thread"] or ""
    analysis = ticket["analysis"] or ""
    code_ctx = find_template_code(ticket["subject"], analysis, db=db)
    kb_context = get_knowledge_base_context(db)

    code_section = ""
    if code_ctx:
        code_section = f"""
TEMPLATE CODE REFERENCE (for your understanding ONLY — NEVER reference code, variable names, file paths, or template IDs in your output):
{code_ctx}

USE THIS INTERNALLY (never expose in output):
1. Verify what the template actually does vs what the client reports
2. Find REFERENCE PATTERNS in other sections
3. Propose solutions CONSISTENT with existing patterns
WARNING: Writing ANY code reference in your output is a CRITICAL ERROR.
"""

    # ── Agent Pipeline for Chat: KB Agent + Code Agent ──
    kb_section = ""
    try:
        chat_orchestrator = AgentOrchestrator(anthropic_key, db=db)

        # KB Agent
        if kb_context:
            kb_brief = chat_orchestrator.get_kb_brief(
                ticket_id, ticket["subject"],
                f"{user_message}\n\n{analysis[:500]}",
                kb_context, "",
                code_ctx[:300] if code_ctx else ""
            )
            if kb_brief:
                kb_section = f"\nKB AGENT BRIEF (pre-validated knowledge):\n{kb_brief}\n"
                kb_section += f"\nRAW KB (backup):\n{truncate_text(kb_context, 2000)}"
            else:
                kb_section = f"\n{truncate_text(kb_context, 5000)}"

        # Code Agent: replace raw code_section with functional brief
        if code_ctx:
            code_brief = chat_orchestrator.get_code_brief(
                ticket_id, ticket["subject"],
                f"{user_message}\n\n{analysis[:500]}",
                code_ctx, kb_brief if kb_context else ""
            )
            if code_brief:
                code_section = f"\nCODE AGENT BRIEF (functional analysis of template code — NEVER reference code in your output):\n{code_brief}\n"
    except Exception as agent_err:
        logger.warning(f"Agent pipeline failed for chat on ticket {ticket_id}: {agent_err}")
        if not kb_section and kb_context:
            kb_section = f"\n{truncate_text(kb_context, 5000)}"

    # Load screenshots for vision API so AI can see Freshdesk attachments.
    # Priority order:
    #   1. User-attached / pasted images (always included, up to 5)
    #   2. User-pinned Freshdesk screenshots (subset of ticket attachments)
    #   3. Fallback: all ticket attachments (original behaviour)
    screenshot_blocks = []
    try:
        # 1. Pasted/uploaded images
        import base64 as _b64
        for img in (attached_images or [])[:5]:
            b64 = (img.get("data_b64") or "").strip()
            if not b64:
                continue
            media_type = img.get("media_type", "image/png")
            name = img.get("name", "pasted_image.png")
            screenshot_blocks.append({"type": "text", "text": f"[User-attached image: {name}]"})
            screenshot_blocks.append({
                "type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": b64},
            })

        # 2/3. Ticket screenshots (pinned subset OR all)
        row = db.execute("SELECT screenshots_json FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
        if row and row["screenshots_json"]:
            all_ss = json.loads(row["screenshots_json"] or "[]")
            if pinned_screenshots:
                wanted = {fn for fn in pinned_screenshots}
                chosen = [s for s in all_ss if s.get("filename") in wanted]
            else:
                chosen = all_ss
            if chosen:
                screenshot_blocks += load_screenshots_for_ai(chosen, max_images=5)
    except Exception as img_err:
        logger.warning(f"ai_chat: failed to assemble image blocks: {img_err}")

    try:
        client = Anthropic(api_key=anthropic_key)

        if current_draft and not question_only:
            # REWRITE MODE: Apply the change directly to the draft
            system_prompt = f"""You are the Product Owner for Silverfin's Luxembourg templates team (BSO LUX).
The user wants you to modify their draft response for a Freshdesk ticket based on their instruction.

{lang_instruction}

IMPORTANT — KNOWLEDGE BASE VALIDATION:
Before applying any changes, check the knowledge base below for relevant information:
- Chart of accounts / account ranges: verify the accounts mentioned are correct
- Reconciliation rules: verify the request is accounting-correct
- Liquid limitations: verify the proposed change is technically feasible
If something is accounting-incorrect or technically infeasible in Liquid, WARN in your response.
Do not blindly implement changes that are wrong from an accounting or technical perspective.

TICKET SUBJECT: {ticket['subject']}
TICKET ANALYSIS: {truncate_text(analysis, 2000)}
{code_section}
{kb_section}

CURRENT DRAFT:
{current_draft}

CRITICAL INSTRUCTIONS:
1. Apply the user's requested change DIRECTLY to the draft.
2. Return the COMPLETE updated draft — all 3 sections (or 2 for decline) with their exact headers:
   --- CLIENT RESPONSE ---
   --- INTERNAL NOTE (BSO LUX) ---
   --- BACKLOG TICKET ---
3. Only modify the section(s) affected by the user's request. Keep other sections unchanged.
4. For the INTERNAL NOTE, follow this ORDER: Hi team → position → issue explanation → conditions (bullet points with *) → "Next step:" → "Proposed new wording" (with FR+EN pair) → edge cases → Thanks.
   - "Next step:" ALWAYS comes BEFORE "Proposed new wording". This order is mandatory.
   - CHECK THE ACTUAL TEMPLATE CODE to find the REAL current wording. Do NOT invent or guess.
   - Provide EXACT current text and EXACT new text.
   - Provide BOTH FR and EN for the proposed new wording (BSO needs both to implement).
   - Everything else in the note is in ONE language only (the language of this draft).

ABSOLUTE RULES (violation = critical error):
5. ZERO CODE. Never write variable names (employees_cy, hide_breakdown), file paths (main.liquid, text_parts/...),
   template IDs (lux_aa_an_staff_cost), code logic (if X == 0), or section names from code (INTRO SECTION).
   You USE the code to understand. You WRITE in plain human language.
   BAD: "La variable employees_cy = 0" → GOOD: "L'effectif de l'exercice en cours est de 0"
   BAD: "Fichiers concernés : lu_market-main/..." → NEVER list file paths.
6. NO markdown (no **, no #). Plain text only.
7. "Next step:" in ONE language only (matching this draft's language). NEVER duplicate in both FR and EN.
8. Keep everything in {lang.upper()} except the FR+EN pair for proposed wording changes.
9. Return ONLY the updated full draft text. No preamble, no explanation.

LUXEMBOURG LEGAL TERMINOLOGY:
- SARL 1 manager: "Gérant Unique" / "Sole Manager"
- SARL 2+ managers: "Conseil de Gérance" / "Board of Managers" (NEVER "Gérants")
- SA 1 director: "Administrateur Unique" / "Sole Director"
- SA 2+ directors: "Conseil d'Administration" / "Board of Directors" """

            # Build user content with optional screenshot vision blocks
            if screenshot_blocks:
                user_content = screenshot_blocks + [{"type": "text", "text": user_message}]
            else:
                user_content = user_message

            response = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=2500,
                system=system_prompt,
                messages=[{"role": "user", "content": user_content}],
            )
            updated_draft = strip_code_from_output(response.content[0].text.strip())

            # Update BOTH language versions — not just the one being edited.
            # The edited language gets the AI rewrite; the other language gets a fresh translation.
            # This ensures both FR and EN drafts stay in sync and neither is stale.
            if lang == "fr":
                updated_fr = updated_draft
                updated_en = translate_draft(updated_draft, "fr", "en", anthropic_key)
            else:
                updated_en = updated_draft
                updated_fr = translate_draft(updated_draft, "en", "fr", anthropic_key)

            db.execute("UPDATE tickets SET draft_response = ?, draft_response_en = ? WHERE ticket_id = ?",
                       (updated_fr, updated_en, ticket_id))
            db.commit()

            # Learning Agent: track AI-assisted edits for future improvement
            # (the user's instruction to the AI is a form of PO guidance)
            if current_draft and current_draft.strip() != updated_draft.strip():
                try:
                    _api_k = anthropic_key
                    _t_id = ticket_id
                    _t_subj = ticket["subject"]
                    _t_tpl = _row_get(ticket, "template_name", "") or ""
                    _t_wf = _row_get(ticket, "workflow_name", "") or ""
                    _old_d = current_draft
                    _new_d = updated_draft
                    def _learn_chat():
                        try:
                            bg_db = get_db_standalone()
                            orch = AgentOrchestrator(_api_k, db=bg_db)
                            orch.run_learning(_t_id, _t_subj, _t_tpl, _t_wf, _old_d, _new_d, "ai_chat_rewrite")
                            bg_db.close()
                        except Exception:
                            pass
                    threading.Thread(target=_learn_chat, daemon=True).start()
                except Exception:
                    pass

            return jsonify({
                "reply": "Draft updated — both FR and EN versions have been synced.",
                "updated_draft": updated_draft,
                "updated_draft_fr": updated_fr,
                "updated_draft_en": updated_en,
            })

        else:
            # QUESTION MODE: Answer about the ticket/template
            system_prompt = f"""You are the Product Owner for Silverfin's Luxembourg templates team (BSO LUX).
The user is asking a question about a Freshdesk ticket. Answer based on your knowledge AND the actual template code if provided.

{lang_instruction}

TICKET SUBJECT: {ticket['subject']}
TICKET ANALYSIS: {truncate_text(analysis, 1500)}
TICKET THREAD (summary): {truncate_text(compiled, 2000)}
{code_section}
{kb_section}

RULES:
- Answer directly and specifically. If the user asks about current wording, CHECK THE CODE and quote the actual text.
- CHECK THE KNOWLEDGE BASE for relevant chart of accounts info, reconciliation rules, account ranges,
  and Liquid template limitations before answering. If the KB has relevant data, USE IT in your answer.
- If the client's request seems accounting-incorrect or technically infeasible in Liquid, say so clearly.
- NO markdown (no **, no #). Plain text only.
- NEVER reference code variables or programming terms. Describe in functional language.
- Keep your answer focused and concise.

LUXEMBOURG LEGAL TERMINOLOGY:
- SARL 1 manager: "Gérant Unique" / SARL 2+ managers: "Conseil de Gérance"
- SA 1 director: "Administrateur Unique" / SA 2+ directors: "Conseil d'Administration" """

            # Build user content with optional screenshot vision blocks
            if screenshot_blocks:
                q_content = screenshot_blocks + [{"type": "text", "text": user_message}]
            else:
                q_content = user_message

            response = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=800,
                system=system_prompt,
                messages=[{"role": "user", "content": q_content}],
            )
            reply = strip_code_from_output(response.content[0].text.strip())
            return jsonify({"reply": reply})

    except Exception as e:
        logger.error(f"AI chat failed for ticket {ticket_id}: {e}")
        return jsonify({"error": f"AI request failed: {str(e)[:200]}"}), 500


def generate_prd_analysis(compiled_thread, anthropic_key, lang, existing_analysis="", project_instructions="", kb_context="", code_context="", terminology_context="", po_draft_response="", ticket_id=None):
    """Generate deep functional analysis content for the PRD document using AI.
    Uses a comprehensive prompt that produces developer-ready functional specs.
    If a PO draft response is provided (edited via AI assistant), use it as the PRIMARY source of truth.
    The AI also analyses template code to propose solutions based on existing patterns in the same template."""
    client = Anthropic(api_key=anthropic_key)

    project_section = ("Project context & instructions:\n" + truncate_text(project_instructions, 2000)) if project_instructions else ""
    analysis_section = ("Previous analysis summary:\n" + truncate_text(existing_analysis, 1500)) if existing_analysis else ""
    term_section = f"\n{terminology_context}\n" if terminology_context else ""

    code_section = ""
    if code_context:
        is_agent_brief = code_context.startswith("CODE AGENT BRIEF") or "TEMPLATE OVERVIEW" in code_context[:200]
        if is_agent_brief:
            code_section = f"""
TEMPLATE ANALYSIS (functional description — already in plain language):
{code_context}
USE THIS to understand current behaviour and find reference patterns. Do NOT add code references to your output.
"""
        else:
            code_section = f"""
TEMPLATE CODE REFERENCE (for your understanding ONLY — NEVER reference code, variable names, or file paths in your output):
{code_context}
Use this to understand current behaviour and find reference patterns. Write in PLAIN HUMAN LANGUAGE only.
"""

    po_draft_section = ""
    if po_draft_response:
        po_draft_section = f"""
PO'S VALIDATED INTERNAL NOTE — PRIMARY SOURCE OF TRUTH:
The Product Owner has written and refined this BSO internal note. It contains their final validated
analysis: position, problem description, requested functional change, functional logic, visibility rules,
edge cases, and proposed wording.

YOUR JOB: Transform this into a STRUCTURED DEVELOPER SPEC following the exact document format below.
- Do NOT simplify — EXPAND into developer-ready detail with all combinations, tables, and rules.
- Do NOT add "Next step" — that belongs only in the BSO note, not in the PRD.
- Extract ALL functional logic (dropdowns, conditions, visibility rules, wording options).
- If the PO provides exact wording in FR and EN — copy EXACTLY.
- If the PO describes a pattern to follow — look in the template code to understand that pattern
  and describe what the developer needs to replicate.

PO'S NOTE:
{truncate_text(po_draft_response, 5000)}
"""

    if lang == "fr":
        lang_instruction = """LANGUAGE RULE — MANDATORY: Write ALL content in FRENCH. Every word must be in French.
Do NOT copy English from context data — translate everything. The ONLY exception: proposed template wording
must include BOTH FR and EN versions for developer reference."""
    else:
        lang_instruction = """LANGUAGE RULE — MANDATORY: Write ALL content in ENGLISH. Every word must be in English.
Do NOT copy French from context data — translate everything. The ONLY exception: proposed template wording
must include BOTH FR and EN versions for developer reference."""

    system = f"""You are a product analyst for the Silverfin Luxembourg templates team. You write analysis documents
(PRDs) for bug fixes and small features. These documents are the SINGLE SOURCE OF TRUTH between product,
development, and QA.

{lang_instruction}{term_section}

{po_draft_section}

THE NON-NEGOTIABLE CONTENT RULE:
You write only what the user/PO told you. You never invent business rules, field names, wording,
account ranges, dropdown options, logic, or edge cases. Missing detail → write "<<<ADD>>>" and continue.

KNOWLEDGE BASE VALIDATION — CRITICAL:
The knowledge base context below contains chart of accounts, account ranges, reconciliation rules,
template documentation, and Liquid template limitations. You MUST check it before writing your analysis:
- Verify account ranges and classes mentioned in the ticket are correct per the chart of accounts
- Verify the request is accounting-correct (proper account mappings, valid debit/credit logic)
- Verify the request is technically feasible in Silverfin Liquid templates
- If something is accounting-incorrect → state clearly in the context section
- If something is technically infeasible in Liquid → state clearly and flag with <<<INFEASIBLE>>>
Do NOT blindly transform the PO's note into a spec if the underlying request is wrong.

WHEN PROPOSING A SOLUTION:
If the PO says "follow the same pattern as section X" or "restructure like section Y":
1. Look at the template code provided to find that reference section
2. Describe functionally what that section does (dropdowns, conditions, text fields, visibility)
3. Apply that SAME pattern to the section being fixed
4. NEVER propose a simpler solution that ignores the reference pattern
If the PO describes specific dropdowns, conditions, or functional logic → use EXACTLY that.
If no PO draft is provided, analyse the code to find similar patterns in the same template and propose
a solution consistent with those existing patterns.

DOCUMENT STRUCTURE — Section 1 (Bug fix / small feature):
The JSON you produce maps directly to these document sections in this exact order:
1. Intro sentence (italic) — one sentence: template, section, workflow, what the fix is about
2. Quick links — Freshdesk URL, Jira <<<ADD>>>
3. Info impacted template(s) — name, workflow
4. Bullet items to check — period logic (Yes/No), linked templates (Yes/No + which)
5. Current issue / feature:
   - Context (Heading 2) — 1-2 paragraphs for a first-time reader, neutral, no judgement
   - Problem statement (Heading 2) — what is wrong, bullet list of issues if multiple
   - Current behaviour (Heading 2) — what happens now, FR+EN sentences, consequence
   - New behaviour after fix (Heading 2) — the CORE section:
     - For simple fixes: one sentence + new FR/EN wording
     - For complex fixes with dropdowns/infobox: Heading 3 subsections for each component:
       * Infobox structure (sequential steps)
       * Infobox base sentence (FR+EN)
       * Each dropdown (options table: Option | FR | EN, mandatory/conditional rules)
       * Editable text field (default text per combination table)
       * Complete sentence construction (published output table per combination)
   - Visibility rules summary (Heading 2) — ALWAYS a table: Condition | Infobox view | Published output
   - Reference implementation (Heading 2) — which section to examine + bullet list of what "same pattern" means
6. Test plan — ALWAYS a table: # | Scenario | Input | Expected result | Status

FORMATTING RULES (critical for document quality):
- FR/EN bilingual text: two consecutive lines with bold labels "FR:" and "EN:", NEVER as bullet points
- Dropdown options: ALWAYS a table (Option | FR | EN), never bullets
- Visibility rules: ALWAYS a table, never prose only
- Test plan: ALWAYS a table, Status always = "To test"
- Bold only for: labels, step names, language prefixes (FR:/EN:), key terms. Never bold entire sentences.
- Use * bullet points for: listing issues, sequential steps, reference pattern items
- Use numbered lists (1. 2. 3.) for functional logic steps in the new behaviour section
- Put constraints (mandatory, conditional visibility) in Heading 3 titles AND in prose after tables
- Every table needs a lead-in sentence before it

TEST SCENARIOS — how many and what to cover:
- Simple fix (label/date change): 3-5 scenarios
- Fix with one dropdown: 4-6 scenarios
- Fix with conditional dropdowns + editable field: 8-10 scenarios
Always include: section inactive, active but no input, each dropdown path, custom text edit,
text field emptied, validation failure per required dropdown, original client use case.

{code_section}

ABSOLUTE RULES:
1. ZERO CODE in output. No variable names, file paths, template IDs, code logic. Plain human language.
2. French prose = 100% French. English prose = 100% English. Exception: proposed wording always has BOTH.
3. NO "Next step" — BSO note concept only, never in the PRD.
4. NO markdown formatting (**, #, ```). Plain text with * for bullets.
5. If PO draft has functional logic with dropdowns/conditions/wordings → copy ALL of it, expanded.
6. Match depth to complexity. Wording fix = concise. Infobox restructuring = very detailed with all tables.
7. <<<ADD>>> for any missing information. Never invent.

{project_section}
{analysis_section}

Reply with ONLY valid JSON (no markdown, no code blocks):

{{
  "template_name": "exact human-readable name",
  "workflow": "Luxembourg Annual Accounts" or "Luxembourg Corporate Tax" or "Legal Documents",
  "period_logic": "No" or "Yes – [why]" or "To be determined",
  "linked_templates": "No" or "Yes – [list names + brief explanation]",
  "intro_sentence": "This analysis covers [specific change] for the [template name] section in the [workflow] workflow.",
  "context": "1-2 paragraphs. What this template does, what the section does, what user interaction looks like. Neutral, no judgement. For a first-time reader.",
  "problem_statement": "What specifically is wrong or missing. If multiple issues: 'This creates N critical issues:\\n* issue 1\\n* issue 2'. Include the specific client scenario if known.",
  "current_behaviour": "What happens now when the section is active. Include the exact FR and EN sentences that currently print (prefixed with FR: and EN:). End with a consequence sentence (what the user cannot do / what the workaround is).",
  "new_behaviour_summary": "One sentence: what must change and what reference pattern to follow (if any).",
  "new_behaviour_subsections": [
    {{
      "heading": "Infobox structure",
      "content": "Description of the sequential interaction flow. Use bullet points with * for each step: '* Step 1 — First mandatory dropdown appears when...\\n* Step 2 — Second mandatory dropdown appears only if...\\n* Step 3 — Editable text field appears below...'"
    }},
    {{
      "heading": "Infobox base sentence",
      "content": "Lead-in sentence + FR: [base sentence fragment]\\nEN: [base sentence fragment]"
    }},
    {{
      "heading": "First dropdown — [constraint] (mandatory)",
      "content": "Lead-in + table description. Use this exact format for the table:\\nDROPDOWN_TABLE:\\nOption|FR text|EN text\\nA|[FR option A]|[EN option A]\\nB|[FR option B]|[EN option B]\\n\\nConstraint: This dropdown must be marked as required..."
    }},
    {{
      "heading": "Second dropdown — [when visible] (mandatory when visible)",
      "content": "Same format as first dropdown with DROPDOWN_TABLE"
    }},
    {{
      "heading": "Editable text field — default text per dropdown combination",
      "content": "Lead-in + COMBINATION_TABLE:\\nCombination|Default FR|Default EN|Editable?\\n[row per combination]"
    }},
    {{
      "heading": "Complete sentence construction — published output examples",
      "content": "Lead-in + OUTPUT_TABLE:\\nCombination|Published FR|Published EN|Editable field\\n[row per combination]"
    }}
  ],
  "visibility_rules": "VISIBILITY_TABLE:\\nCondition|Infobox (template view)|Published output|Second dropdown\\n[one row per distinct state combination]",
  "reference_implementation": "Which section to examine + bullet list of what 'same pattern' means. If no reference: 'N/A'",
  "proposed_wording_current_fr": "Exact current French text. If N/A: 'N/A'",
  "proposed_wording_current_en": "Exact current English text. If N/A: 'N/A'",
  "proposed_wording_new_fr": "Exact new French text with [placeholders]. If N/A: 'N/A'",
  "proposed_wording_new_en": "Exact new English text with [placeholders]. If N/A: 'N/A'",
  "legal_reference": "N/A or relevant law",
  "complexity_assessment": "Simple fix (wording/config)" or "Medium (logic/condition)" or "Complex (infobox restructuring/multi-template)",
  "test_scenarios": [
    {{"scenario": "short label 3-6 words", "input": "exact state: checkbox value, dropdown value, field state", "expected": "what user sees + what prints. For bilingual: FR: '[sentence]' EN: '[sentence]'. Include secondary elements."}}
  ],
  "po_checklist_extra": ["Specific checklist items for this fix beyond the 3 defaults"]
}}

IMPORTANT — the new_behaviour_subsections array:
- For SIMPLE fixes (wording change, single condition): use 1-2 subsections or leave empty array [].
- For COMPLEX fixes (dropdown logic, infobox restructuring): use 4-7 subsections as shown above.
- Each subsection has "heading" (the Heading 3 title with constraint) and "content" (the body).
- Put the KEY CONSTRAINT directly in the heading title: "First dropdown — completes the sentence (mandatory)"
- Tables inside content use a special format: TABLE_TYPE:\\nCol1|Col2|Col3\\nRow1col1|Row1col2|Row1col3
  Types: DROPDOWN_TABLE, COMBINATION_TABLE, OUTPUT_TABLE, VISIBILITY_TABLE"""

    # Build user message with all context
    short_thread = truncate_text(compiled_thread, 8000)
    short_kb = truncate_text(kb_context, 5000) if kb_context else ""
    user_text = short_thread
    if short_kb:
        user_text += f"\n\nKNOWLEDGE BASE — CHECK THIS BEFORE WRITING (chart of accounts, account ranges, reconciliation rules, Liquid limitations):\n{short_kb}"

    # Load screenshots as vision content blocks if available
    screenshot_blocks = []
    if ticket_id:
        screenshot_blocks = load_screenshots_for_ai(ticket_id)

    if screenshot_blocks:
        user_content = screenshot_blocks + [{"type": "text", "text": user_text}]
    else:
        user_content = user_text

    resp = call_anthropic_with_retry(
        client,
        model="claude-sonnet-4-5",
        max_tokens=10000,
        system=system,
        messages=[{"role": "user", "content": user_content}],
    )
    text = resp.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.strip()
    try:
        prd = json.loads(text)
        # Strip code that leaked into text fields
        for field in ("context", "problem_statement", "current_behaviour", "new_behaviour_summary",
                      "visibility_rules", "reference_implementation", "intro_sentence"):
            if prd.get(field):
                prd[field] = strip_code_from_output(prd[field])
        # Also clean subsections
        for sub in prd.get("new_behaviour_subsections", []):
            if sub.get("content"):
                sub["content"] = strip_code_from_output(sub["content"])
        return prd
    except json.JSONDecodeError:
        logger.error(f"PRD analysis JSON parse failed: {text[:300]}")
        return {
            "template_name": "Could not determine",
            "workflow": "To be determined",
            "period_logic": "To be determined",
            "linked_templates": "To be determined",
            "intro_sentence": "This analysis covers the issue reported in the Freshdesk ticket.",
            "context": "<<<ADD>>> — Describe what this template/note does and its role in the workflow.",
            "problem_statement": text[:800] if text else "<<<ADD>>> — Analysis generation failed.",
            "current_behaviour": "<<<ADD>>>",
            "new_behaviour_summary": "<<<ADD>>>",
            "new_behaviour_subsections": [],
            "visibility_rules": "",
            "reference_implementation": "N/A",
            "test_scenarios": [],
            "po_checklist_extra": [],
        }


@app.route("/ticket/<int:ticket_id>/refresh-screenshots", methods=["POST"])
def refresh_screenshots(ticket_id):
    """On-demand: re-fetch ticket from Freshdesk and extract/download any screenshots.
    Useful for old tickets that were saved before the screenshot feature existed."""
    db = get_db()
    api_key = get_setting("freshdesk_api_key", db=db)
    domain = get_setting("freshdesk_domain", "silverfin.freshdesk.com", db=db)
    if not api_key:
        return jsonify({"success": False, "message": "Freshdesk API key not configured."}), 400
    try:
        ticket_data, conversations = get_ticket_details(api_key, domain, ticket_id)
        screenshots = extract_and_download_screenshots(ticket_data, conversations, api_key, domain, ticket_id)

        # Also update raw_conversations (keeps inbox detection fresh)
        db.execute(
            "UPDATE tickets SET screenshots_json = ?, raw_conversations = ?, status = ?, status_code = ? WHERE ticket_id = ?",
            (json.dumps(screenshots or [], ensure_ascii=False), json.dumps(conversations),
             STATUS_MAP.get(ticket_data.get("status", 0), "Unknown"), ticket_data.get("status", 0),
             ticket_id)
        )
        db.commit()
        return jsonify({"success": True, "count": len(screenshots),
                        "message": f"Found {len(screenshots)} image(s)." if screenshots else "No images found in this ticket."})
    except Exception as e:
        logger.error(f"Screenshot refresh failed for ticket {ticket_id}: {e}")
        return jsonify({"success": False, "message": str(e)[:300]}), 500


@app.route("/ticket/<int:ticket_id>/upload-screenshot", methods=["POST"])
def upload_screenshot(ticket_id):
    """Upload an image file (from paste, drag-drop, or file picker in the rich editor).
    Saves it to screenshots/<ticket_id>/ and updates screenshots_json in the DB.
    Returns JSON: {success, filename, url}
    """
    db = get_db()
    ticket = db.execute("SELECT ticket_id, screenshots_json FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        return jsonify({"success": False, "message": "Ticket not found"}), 404

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"success": False, "message": "No file provided"}), 400

    # Validate it's an image
    allowed_ext = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_ext:
        return jsonify({"success": False, "message": f"Invalid file type: {ext}"}), 400

    screenshots_dir = os.path.join(_DATA_DIR, "screenshots", str(ticket_id))
    os.makedirs(screenshots_dir, exist_ok=True)

    safe_name = secure_filename(file.filename) or f"uploaded_{int(time.time())}.png"

    # Avoid overwriting — add counter suffix if file exists
    local_path = os.path.join(screenshots_dir, safe_name)
    counter = 1
    base, ext_part = os.path.splitext(safe_name)
    while os.path.exists(local_path):
        safe_name = f"{base}_{counter}{ext_part}"
        local_path = os.path.join(screenshots_dir, safe_name)
        counter += 1

    file.save(local_path)
    file_size = os.path.getsize(local_path)

    # Update screenshots_json in DB
    try:
        existing = json.loads(ticket["screenshots_json"] or "[]")
    except (json.JSONDecodeError, TypeError):
        existing = []

    existing.append({
        "filename": safe_name,
        "path": local_path,
        "source": "User upload (editor)",
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "content_type": file.content_type or "image/png",
        "size": file_size,
    })
    db.execute("UPDATE tickets SET screenshots_json = ? WHERE ticket_id = ?",
               (json.dumps(existing, ensure_ascii=False), ticket_id))
    db.commit()

    logger.info(f"Uploaded screenshot '{safe_name}' for ticket {ticket_id} ({file_size} bytes)")
    return jsonify({
        "success": True,
        "filename": safe_name,
        "url": f"/screenshots/{ticket_id}/{safe_name}",
    })


@app.route("/screenshots/<int:ticket_id>/<path:filename>")
def serve_screenshot(ticket_id, filename):
    """Serve a screenshot file for a ticket."""
    screenshots_dir = os.path.join(_DATA_DIR, "screenshots", str(ticket_id))
    safe_name = secure_filename(filename)
    file_path = os.path.join(screenshots_dir, safe_name)
    if os.path.exists(file_path):
        return send_file(file_path)
    return "Screenshot not found", 404


@app.route("/ticket/<int:ticket_id>/prepare-analysis", methods=["POST"])
def prepare_analysis(ticket_id):
    """Pre-generate deep PRD analysis content via AI and store it in the DB.
    This is a separate step from downloading the doc, so the user gets feedback."""
    db = get_db()
    lang = request.form.get("lang", "fr")
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    anthropic_key = get_setting("anthropic_api_key", db=db)
    if not anthropic_key:
        flash("Anthropic API key not configured. Go to Settings.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    if not ticket["compiled_thread"]:
        flash("No ticket data available. Run a Freshdesk fetch first.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    try:
        project_instr = get_setting("claude_project_instructions", db=db) or ""
        kb_context = get_knowledge_base_context(db)
        term_context = get_terminology_context(db)
        code_ctx = find_template_code(ticket["subject"], ticket["analysis"], db=db)

        # Include the user's edited draft response — this contains the refined BSO note,
        # conditions, next steps, and proposed wordings that the PO has validated.
        draft_fr = ticket["draft_response"] or ""
        draft_en = ticket["draft_response_en"] if "draft_response_en" in ticket.keys() else ""
        current_draft = draft_fr if lang == "fr" else (draft_en or draft_fr)

        # ── Agent Pipeline: Parallel Prep → Main PRD → QA with Retry ──
        orchestrator = AgentOrchestrator(anthropic_key, db=db)

        # Search Jira for related issues
        jira_ctx = search_jira_for_ticket(ticket["subject"], _row_get(ticket, "template_name", ""), db=db)

        # 1. Run KB + Code + Research agents in PARALLEL
        kb_brief, code_brief, research_brief = orchestrator.run_preparation_agents_parallel(
            ticket_id, ticket["subject"], (ticket["analysis"] or "")[:1000],
            kb_context, code_ctx,
            terminology_context=term_context,
            template_name=_row_get(ticket, "template_name", ""),
            workflow_name=_row_get(ticket, "workflow_name", ""),
            jira_context=jira_ctx,
        )

        # 2. Build enhanced context from all agent briefs
        enhanced_kb = "\n[CONTEXT DATA BELOW IS FOR REFERENCE ONLY — DO NOT COPY TEXT FROM IT. TRANSLATE TO THE OUTPUT LANGUAGE.]\n"
        if kb_brief:
            enhanced_kb += f"\nKB AGENT BRIEF (pre-validated knowledge for this ticket):\n{kb_brief}\n"
        if research_brief:
            enhanced_kb += f"\nRESEARCH AGENT BRIEF (similar past tickets and lessons):\n{research_brief}\n"
        enhanced_kb += f"\nRAW KNOWLEDGE BASE (backup):\n{truncate_text(kb_context, 2000)}" if kb_context else ""

        # Use Code Agent brief instead of raw code
        # Sanitize to remove any code Haiku may have slipped into its brief
        effective_code_ctx = strip_code_from_output(code_brief) if code_brief else ""

        # 3. Main PRD Agent
        prd_result = generate_prd_analysis(
            ticket["compiled_thread"], anthropic_key, lang,
            existing_analysis=ticket["analysis"] or "",
            project_instructions=project_instr,
            kb_context=enhanced_kb,
            code_context=effective_code_ctx,
            terminology_context=term_context,
            po_draft_response=current_draft,
            ticket_id=ticket_id,
        )

        # 4. QA Agent with retry: validate the PRD output
        qa_result, _ = orchestrator.run_qa_with_retry(
            ticket_id, json.dumps(prd_result, ensure_ascii=False), "prd_analysis",
            ticket["subject"], kb_brief
        )
        if qa_result.get("critical_issues"):
            logger.warning(f"QA issues for PRD of ticket {ticket_id}: {qa_result['critical_issues']}")
            # Add QA notes to the PRD context field so PO sees them
            qa_note = "[QA ALERT: " + " | ".join(qa_result["critical_issues"][:3]) + "]"
            prd_result["context"] = prd_result.get("context", "") + f"\n\n{qa_note}"

        # Store as JSON in the prd_content column
        db.execute(
            "UPDATE tickets SET prd_content = ? WHERE ticket_id = ?",
            (json.dumps(prd_result, ensure_ascii=False), ticket_id),
        )
        db.commit()

        # Extract and save template_name and workflow_name if present
        if prd_result.get("template_name"):
            db.execute("UPDATE tickets SET template_name = ?, workflow_name = ? WHERE ticket_id = ?",
                       (prd_result.get("template_name", ""), prd_result.get("workflow_name", ""), ticket_id))
            db.commit()
        flash("Deep analysis prepared successfully (agent pipeline) — you can now download the document.", "success")
    except Exception as e:
        logger.error(f"PRD analysis generation failed for ticket {ticket_id}: {e}")
        flash(f"Analysis generation failed: {str(e)[:200]}", "error")

    return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/ticket/<int:ticket_id>/generate-doc")
def generate_doc(ticket_id):
    """Generate and download a Word document analysis report for a ticket.
    Uses pre-generated PRD content from the DB (instant — no AI call here)."""
    db = get_db()
    lang = request.args.get("lang", "fr")
    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    # Load pre-generated PRD analysis from DB (if available)
    prd_analysis = {}
    prd_raw = ticket["prd_content"] if "prd_content" in ticket.keys() else None
    if prd_raw:
        try:
            prd_analysis = json.loads(prd_raw)
        except (json.JSONDecodeError, TypeError):
            pass

    # Load screenshots metadata
    screenshots = []
    screenshots_raw = ticket["screenshots_json"] if "screenshots_json" in ticket.keys() else "[]"
    try:
        screenshots = json.loads(screenshots_raw or "[]")
    except (json.JSONDecodeError, TypeError):
        screenshots = []

    # Get selected screenshots from query params (if user selected specific ones)
    selected = request.args.get("screenshots", "")
    if selected:
        selected_files = [s.strip() for s in selected.split(",") if s.strip()]
        if selected_files:
            screenshots = [s for s in screenshots if s.get("filename") in selected_files]

    # Prepare payload for the Node.js generator
    ticket_data = {k: ticket[k] for k in ticket.keys()}
    payload = {
        "language": lang,
        "ticket": ticket_data,
        "prd_analysis": prd_analysis,
        "screenshots": screenshots,
    }

    input_path = os.path.join(_DATA_DIR, f"_tmp_ticket_{ticket_id}.json")
    output_dir = os.path.join(_DATA_DIR, "generated_docs")
    os.makedirs(output_dir, exist_ok=True)

    safe_subject = re.sub(r'[^\w\s-]', '', (ticket["subject"] or "ticket")[:40]).strip().replace(" ", "_")
    filename = f"Analysis_{ticket_id}_{safe_subject}_{lang.upper()}.docx"
    output_path = os.path.join(output_dir, filename)

    try:
        with open(input_path, "w") as f:
            json.dump(payload, f, default=str)

        script_path = os.path.join(app_dir, "generate_analysis.js")
        node_modules_path = os.path.join(app_dir, "node_modules")
        env = os.environ.copy()
        env["NODE_PATH"] = node_modules_path
        result = subprocess.run(
            ["node", script_path, input_path, output_path],
            capture_output=True, text=True, timeout=30, cwd=app_dir, env=env
        )

        if result.returncode != 0:
            logger.error(f"Doc generation failed: {result.stderr}")
            flash(f"Document generation failed: {result.stderr[:200]}", "error")
            return redirect(url_for("ticket_detail", ticket_id=ticket_id))

        return send_file(output_path, as_attachment=True, download_name=filename)

    except Exception as e:
        logger.error(f"Doc generation error: {e}")
        flash(f"Document generation error: {str(e)[:200]}", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))
    finally:
        if os.path.exists(input_path):
            os.remove(input_path)


@app.route("/settings", methods=["GET", "POST"])
def settings():
    """Settings page for API keys and configuration."""
    db = get_db()

    if request.method == "POST":
        section = request.form.get("section", "")

        # Claude Project settings (separate form)
        if section == "claude_project":
            for cp_field in ["claude_project_name", "claude_project_id", "claude_project_instructions"]:
                val = request.form.get(cp_field, "").strip()
                set_setting(cp_field, val, db=db)
            flash("Claude Project settings saved.", "success")
            return redirect(url_for("settings") + "#knowledge-base")

        if section == "client_context":
            val = request.form.get("client_context", "").strip()
            set_setting("client_context", val, db=db)
            flash("Client data context saved.", "success")
            return redirect(url_for("settings") + "#knowledge-base")

        if section == "google_drive":
            sa_json = request.form.get("google_sa_json", "").strip()
            export_folder = request.form.get("google_export_folder", "").strip()
            kb_folder = request.form.get("google_kb_folder", "").strip()
            set_setting("google_sa_json", sa_json, db=db)
            set_setting("google_export_folder", export_folder, db=db)
            set_setting("google_kb_folder", kb_folder, db=db)
            flash("Google Drive settings saved.", "success")
            return redirect(url_for("settings") + "#knowledge-base")

        if section == "notion":
            notion_token = request.form.get("notion_token", "").strip()
            notion_page_id = request.form.get("notion_page_id", "").strip()
            set_setting("notion_token", notion_token, db=db)
            set_setting("notion_page_id", notion_page_id, db=db)
            flash("Notion settings saved.", "success")
            return redirect(url_for("settings") + "#knowledge-base")

        if section == "jira":
            for jf in ["jira_domain", "jira_email", "jira_api_token", "jira_project"]:
                val = request.form.get(jf, "").strip()
                set_setting(jf, val, db=db)
            flash("Jira settings saved.", "success")
            return redirect(url_for("settings"))

        fields = [
            "freshdesk_api_key", "freshdesk_domain", "freshdesk_group_id",
            "anthropic_api_key",
        ]
        for field in fields:
            val = request.form.get(field, "").strip()
            if val:  # Only update non-empty values
                set_setting(field, val, db=db)

        # LLM provider settings
        llm_provider = request.form.get("llm_provider", "").strip()
        if llm_provider:
            set_setting("llm_provider", llm_provider, db=db)
        llm_api_key = request.form.get("llm_api_key", "").strip()
        if llm_api_key:
            set_setting("llm_api_key", llm_api_key, db=db)
            if llm_provider == "anthropic" or not llm_provider:
                set_setting("anthropic_api_key", llm_api_key, db=db)
        llm_base_url = request.form.get("llm_base_url", "").strip()
        set_setting("llm_base_url", llm_base_url, db=db)
        llm_fast_model = request.form.get("llm_fast_model", "").strip()
        if llm_fast_model:
            set_setting("llm_fast_model", llm_fast_model, db=db)
        llm_main_model = request.form.get("llm_main_model", "").strip()
        if llm_main_model:
            set_setting("llm_main_model", llm_main_model, db=db)

        # Jira settings (allow saving empty to clear)
        for jf in ["jira_domain", "jira_email", "jira_api_token", "jira_project"]:
            val = request.form.get(jf, "").strip()
            set_setting(jf, val, db=db)

        # Template code path
        code_path = request.form.get("template_code_path", "").strip()
        set_setting("template_code_path", code_path, db=db)

        # Country filter
        country = request.form.get("freshdesk_country", "").strip()
        if country == "other":
            country = request.form.get("freshdesk_country_custom", "").strip()
        set_setting("freshdesk_country", country, db=db)

        # Writing style
        writing_style = request.form.get("writing_style", "customer_support").strip()
        if writing_style in WRITING_STYLES:
            set_setting("writing_style", writing_style, db=db)

        # Statuses (checkboxes)
        statuses = request.form.getlist("freshdesk_statuses")
        set_setting("freshdesk_statuses", ",".join(statuses) if statuses else "2,3,4,20", db=db)

        flash("Settings saved.", "success")
        return redirect(url_for("settings"))

    current = {}
    for key in ["freshdesk_api_key", "freshdesk_domain", "freshdesk_group_id",
                "anthropic_api_key", "freshdesk_country", "freshdesk_country_custom",
                "freshdesk_statuses", "writing_style", "template_code_path",
                "jira_domain", "jira_email", "jira_api_token", "jira_project",
                "llm_provider", "llm_api_key", "llm_base_url",
                "llm_fast_model", "llm_main_model"]:
        current[key] = get_setting(key, db=db)

    # Set defaults
    if not current["freshdesk_domain"]:
        set_setting("freshdesk_domain", "silverfin.freshdesk.com", db=db)
        current["freshdesk_domain"] = "silverfin.freshdesk.com"
    if not current["freshdesk_group_id"]:
        set_setting("freshdesk_group_id", "101000372179", db=db)
        current["freshdesk_group_id"] = "101000372179"
    if not current["writing_style"]:
        current["writing_style"] = "customer_support"

    # Knowledge base entries
    kb_entries = db.execute("SELECT * FROM knowledge_base ORDER BY category, title").fetchall()

    # Claude Project settings
    claude_project_name = get_setting("claude_project_name", db=db) or ""
    claude_project_id = get_setting("claude_project_id", db=db) or ""
    claude_project_instructions = get_setting("claude_project_instructions", db=db) or ""
    client_context = get_setting("client_context", db=db) or ""

    # Check if template code path resolves
    template_path_resolved = resolve_template_path(current.get("template_code_path", ""))
    template_path_status = ""
    if current.get("template_code_path"):
        if template_path_resolved:
            import glob as globmod
            liquid_count = len(globmod.glob(os.path.join(template_path_resolved, "**", "*.liquid"), recursive=True))
            template_path_status = f"ok:{liquid_count}"
        else:
            template_path_status = "error"

    # Google Drive settings
    google_sa_json = get_setting("google_sa_json", db=db) or ""
    google_export_folder = get_setting("google_export_folder", db=db) or ""
    google_kb_folder = get_setting("google_kb_folder", db=db) or ""

    # Notion settings
    notion_token = get_setting("notion_token", db=db) or ""
    notion_page_id = get_setting("notion_page_id", db=db) or ""

    return render_template("settings.html", settings=current,
                           writing_styles=WRITING_STYLES,
                           knowledge_categories=KNOWLEDGE_CATEGORIES,
                           kb_entries=kb_entries,
                           claude_project_name=claude_project_name,
                           claude_project_id=claude_project_id,
                           claude_project_instructions=claude_project_instructions,
                           client_context=client_context,
                           template_path_status=template_path_status,
                           google_sa_json=google_sa_json,
                           google_export_folder=google_export_folder,
                           google_kb_folder=google_kb_folder,
                           notion_token=notion_token,
                           notion_page_id=notion_page_id)


@app.route("/knowledge-base/add", methods=["POST"])
def kb_add():
    """Add a knowledge base entry (text, URL, or file upload)."""
    db = get_db()
    category = request.form.get("category", "").strip()
    title = request.form.get("title", "").strip()
    entry_type = request.form.get("entry_type", "text").strip()

    if not category or not title:
        flash("Category and title are required.", "error")
        return redirect(url_for("settings") + "#knowledge-base")

    if category not in KNOWLEDGE_CATEGORIES:
        flash("Invalid category.", "error")
        return redirect(url_for("settings") + "#knowledge-base")

    content = ""
    file_path_saved = ""
    file_name_saved = ""
    url_saved = ""

    if entry_type == "file":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Please select a file to upload.", "error")
            return redirect(url_for("settings") + "#knowledge-base")
        if not allowed_file(file.filename):
            flash(f"File type not supported. Allowed: {', '.join(ALLOWED_EXTENSIONS)}", "error")
            return redirect(url_for("settings") + "#knowledge-base")

        filename = secure_filename(file.filename)
        # Add timestamp to avoid collisions
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{ts}_{filename}"
        save_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(save_path)
        file_path_saved = save_path
        file_name_saved = file.filename  # Original name for display
        content = request.form.get("content", "").strip()  # Optional notes

    elif entry_type == "url":
        url_saved = request.form.get("url", "").strip()
        if not url_saved:
            flash("Please enter a URL.", "error")
            return redirect(url_for("settings") + "#knowledge-base")
        content = request.form.get("content", "").strip()  # Notes about the URL

    else:  # text
        content = request.form.get("content", "").strip()
        if not content:
            flash("Content is required for text entries.", "error")
            return redirect(url_for("settings") + "#knowledge-base")

    db.execute(
        """INSERT INTO knowledge_base (category, title, content, entry_type, file_path, file_name, url)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (category, title, content, entry_type, file_path_saved, file_name_saved, url_saved),
    )
    db.commit()
    flash(f"Knowledge base entry '{title}' added.", "success")
    return redirect(url_for("settings") + "#knowledge-base")


@app.route("/knowledge-base/<int:entry_id>/edit", methods=["POST"])
def kb_edit(entry_id):
    """Edit a knowledge base entry."""
    db = get_db()
    category = request.form.get("category", "").strip()
    title = request.form.get("title", "").strip()
    content = request.form.get("content", "").strip()
    url = request.form.get("url", "").strip()

    if not category or not title:
        flash("Category and title are required.", "error")
        return redirect(url_for("settings") + "#knowledge-base")

    # Handle file replacement if a new file is uploaded
    new_file = request.files.get("file")
    if new_file and new_file.filename and allowed_file(new_file.filename):
        filename = secure_filename(new_file.filename)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{ts}_{filename}"
        save_path = os.path.join(UPLOAD_FOLDER, filename)
        new_file.save(save_path)
        # Delete old file
        old = db.execute("SELECT file_path FROM knowledge_base WHERE id = ?", (entry_id,)).fetchone()
        if old and old["file_path"] and os.path.exists(old["file_path"]):
            try:
                os.remove(old["file_path"])
            except OSError:
                pass
        db.execute(
            """UPDATE knowledge_base SET category=?, title=?, content=?, url=?, file_path=?, file_name=?,
               updated_at=? WHERE id=?""",
            (category, title, content, url, save_path, new_file.filename,
             datetime.now(timezone.utc).isoformat(), entry_id),
        )
    else:
        db.execute(
            """UPDATE knowledge_base SET category=?, title=?, content=?, url=?,
               updated_at=? WHERE id=?""",
            (category, title, content, url, datetime.now(timezone.utc).isoformat(), entry_id),
        )

    db.commit()
    flash("Knowledge base entry updated.", "success")
    return redirect(url_for("settings") + "#knowledge-base")


@app.route("/knowledge-base/<int:entry_id>/delete", methods=["POST"])
def kb_delete(entry_id):
    """Delete a knowledge base entry and its uploaded file."""
    db = get_db()
    entry = db.execute("SELECT file_path FROM knowledge_base WHERE id = ?", (entry_id,)).fetchone()
    if entry and entry["file_path"] and os.path.exists(entry["file_path"]):
        try:
            os.remove(entry["file_path"])
        except OSError:
            pass
    db.execute("DELETE FROM knowledge_base WHERE id = ?", (entry_id,))
    db.commit()
    flash("Knowledge base entry deleted.", "success")
    return redirect(url_for("settings") + "#knowledge-base")


@app.route("/run", methods=["POST"])
def run_analysis():
    """Start the analysis job."""
    if _get_job_status()["running"]:
        flash("Analysis is already running.", "warning")
        return redirect(url_for("dashboard"))

    thread = threading.Thread(target=run_analysis_job, daemon=True)
    thread.start()
    flash("Analysis started! This page will update automatically.", "success")
    return redirect(url_for("dashboard"))


@app.route("/api/status")
def api_status():
    """Get current job status (for auto-refresh)."""
    return jsonify(_get_job_status())


@app.route("/api/test-freshdesk", methods=["POST"])
def test_freshdesk():
    """Test Freshdesk connection. Accepts key from JSON body or reads from DB."""
    db = get_db()
    data = request.get_json(silent=True) or {}
    api_key = data.get("api_key") or get_setting("freshdesk_api_key", db=db)
    domain = data.get("domain") or get_setting("freshdesk_domain", "silverfin.freshdesk.com", db=db)

    # Save the key if provided so it persists
    if data.get("api_key"):
        set_setting("freshdesk_api_key", data["api_key"], db=db)
    if data.get("domain"):
        set_setting("freshdesk_domain", data["domain"], db=db)

    if not api_key:
        return jsonify({"ok": False, "message": "API key not set"})

    try:
        resp = requests.get(
            f"https://{domain}/api/v2/tickets?per_page=1",
            auth=HTTPBasicAuth(api_key, "X"),
            timeout=10,
        )
        resp.raise_for_status()
        return jsonify({"ok": True, "message": f"Connected to {domain}"})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:200]})


@app.route("/api/test-anthropic", methods=["POST"])
def test_anthropic():
    """Backward-compat wrapper — delegates to test_llm."""
    return test_llm()


@app.route("/api/test-llm", methods=["POST"])
def test_llm():
    """Test LLM provider connection. Reads config from DB or request body."""
    db = get_db()
    data = request.get_json(silent=True) or {}
    provider = data.get("provider") or get_setting("llm_provider", "anthropic", db=db)
    api_key = data.get("api_key") or get_setting(
        "llm_api_key", get_setting("anthropic_api_key", "", db=db), db=db
    )
    base_url = data.get("base_url") or get_setting("llm_base_url", "", db=db)

    # Save the key if provided
    if data.get("api_key"):
        set_setting("llm_api_key", data["api_key"], db=db)
        if provider == "anthropic":
            set_setting("anthropic_api_key", data["api_key"], db=db)

    if not api_key:
        return jsonify({"ok": False, "message": "API key not set"})

    if LLMGateway is not None:
        try:
            gw = LLMGateway(provider_name=provider, api_key=api_key,
                            base_url=base_url or None)
            ok, msg = gw.test_connection()
            return jsonify({"ok": ok, "message": msg})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)[:200]})
    else:
        # Fallback: direct Anthropic test
        try:
            client = Anthropic(api_key=api_key)
            client.messages.create(
                model="claude-sonnet-4-5", max_tokens=10,
                messages=[{"role": "user", "content": "Say OK"}],
            )
            return jsonify({"ok": True, "message": "Connected to Claude AI"})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)[:200]})


@app.route("/api/agents/model-config", methods=["GET"])
def api_agent_model_configs():
    """List all agent model configs."""
    db = get_db()
    return jsonify({"ok": True, "configs": list_agent_model_configs(db)})


@app.route("/api/agents/model-config/<agent_name>", methods=["POST"])
def api_update_agent_model_config(agent_name):
    """Update a single agent's model config."""
    db = get_db()
    payload = request.get_json(silent=True) or {}
    try:
        update_agent_model_config(db, agent_name, payload)
        return jsonify({"ok": True, "message": f"Config for {agent_name} updated"})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:200]}), 500


@app.route("/api/agents/test-model/<agent_name>", methods=["POST"])
def api_test_agent_model(agent_name):
    """Test the configured model for a specific agent."""
    db = get_db()
    cfg = get_agent_model_config(db, agent_name)
    if not cfg:
        return jsonify({"ok": False, "message": f"No config found for {agent_name}"}), 404
    provider = cfg.get("provider", "anthropic")
    api_key = get_setting("llm_api_key", get_setting("anthropic_api_key", "", db=db), db=db)
    base_url = get_setting("llm_base_url", "", db=db)
    if not api_key:
        return jsonify({"ok": False, "message": "API key not configured"})
    if LLMGateway is not None:
        try:
            gw = LLMGateway(provider_name=provider, api_key=api_key,
                            base_url=base_url or None)
            ok, msg = gw.test_connection()
            return jsonify({"ok": ok, "message": msg, "model": cfg.get("model")})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)[:200]})
    return jsonify({"ok": False, "message": "LLMGateway not available"})


@app.route("/api/test-jira", methods=["POST"])
def test_jira():
    """Test Jira connection."""
    data = request.get_json(silent=True) or {}
    domain = data.get("domain", "").strip()
    email = data.get("email", "").strip()
    token = data.get("token", "").strip()
    project = data.get("project", "").strip()

    if not all([domain, email, token]):
        return jsonify({"ok": False, "message": "Domain, email, and API token are required"})

    try:
        import base64
        auth = base64.b64encode(f"{email}:{token}".encode()).decode()
        url = f"https://{domain}/rest/api/3/myself"
        resp = requests.get(url, headers={"Authorization": f"Basic {auth}", "Accept": "application/json"}, timeout=10)
        if resp.status_code == 200:
            user = resp.json()
            msg = f"Connected as {user.get('displayName', email)}"
            if project:
                proj_resp = requests.get(
                    f"https://{domain}/rest/api/3/project/{project}",
                    headers={"Authorization": f"Basic {auth}", "Accept": "application/json"}, timeout=10
                )
                if proj_resp.status_code == 200:
                    proj = proj_resp.json()
                    msg += f" — Project: {proj.get('name', project)} ({proj.get('key', project)})"
                else:
                    msg += f" — Warning: project '{project}' not found"
            return jsonify({"ok": True, "message": msg})
        elif resp.status_code == 401:
            return jsonify({"ok": False, "message": "Authentication failed. Check email and API token."})
        else:
            return jsonify({"ok": False, "message": f"HTTP {resp.status_code}: {resp.text[:200]}"})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:200]})


@app.route("/api/jira/options")
def jira_options():
    """Fetch available issue types, epics, priorities, and components from the Jira project."""
    db = get_db()
    jira_domain = get_setting("jira_domain", "", db=db)
    jira_email = get_setting("jira_email", "", db=db)
    jira_token = get_setting("jira_api_token", "", db=db)
    jira_project = get_setting("jira_project", "LUX", db=db)

    if not jira_domain or not jira_email or not jira_token:
        return jsonify({"ok": False, "message": "Jira not configured. Go to Settings."})

    debug_log = []  # Collect debug info to help diagnose issues

    try:
        # 1. Issue types — try multiple approaches
        issue_types = []

        # Approach A: /project/{key} (works for company-managed projects)
        try:
            project_data = _jira_request("GET", f"/project/{jira_project}",
                                         jira_domain, jira_email, jira_token)
            raw_types = project_data.get("issueTypes", [])
            debug_log.append(f"project endpoint: {len(raw_types)} issue types")
            for it in raw_types:
                issue_types.append({
                    "id": it["id"],
                    "name": it["name"],
                    "description": it.get("description", ""),
                    "subtask": bool(it.get("subtask")),
                    "hierarchyLevel": it.get("hierarchyLevel", 0),
                })
        except Exception as e:
            debug_log.append(f"project endpoint failed: {e}")

        # Approach B: If no types from project endpoint, try /issuetype/project?projectId=
        if not issue_types:
            try:
                # First get the project ID
                proj_info = _jira_request("GET", f"/project/{jira_project}",
                                          jira_domain, jira_email, jira_token)
                project_id = proj_info.get("id", "")
                if project_id:
                    type_data = _jira_request("GET", "/issuetype/project",
                                              jira_domain, jira_email, jira_token,
                                              params={"projectId": project_id})
                    debug_log.append(f"issuetype/project endpoint: {len(type_data)} issue types")
                    for it in type_data:
                        issue_types.append({
                            "id": it["id"],
                            "name": it["name"],
                            "description": it.get("description", ""),
                            "subtask": bool(it.get("subtask", False)),
                            "hierarchyLevel": it.get("hierarchyLevel", it.get("scope", {}).get("type", "") == "PROJECT" and 0 or 0),
                        })
            except Exception as e:
                debug_log.append(f"issuetype/project failed: {e}")

        # Approach C: global /issuetype as last resort
        if not issue_types:
            try:
                all_types = _jira_request("GET", "/issuetype",
                                          jira_domain, jira_email, jira_token)
                debug_log.append(f"global issuetype endpoint: {len(all_types)} types")
                for it in all_types:
                    issue_types.append({
                        "id": it["id"],
                        "name": it["name"],
                        "description": it.get("description", ""),
                        "subtask": bool(it.get("subtask", False)),
                        "hierarchyLevel": it.get("hierarchyLevel", 0),
                    })
            except Exception as e:
                debug_log.append(f"global issuetype failed: {e}")

        logger.info(f"Jira issue types for {jira_project}: {[t['name'] + '(h=' + str(t['hierarchyLevel']) + ',sub=' + str(t['subtask']) + ')' for t in issue_types]}")

        # 2. Priorities
        priorities = []
        try:
            priorities_data = _jira_request("GET", "/priority",
                                            jira_domain, jira_email, jira_token)
            priorities = [{"id": p["id"], "name": p["name"]} for p in priorities_data]
        except Exception as e:
            debug_log.append(f"priorities failed: {e}")

        # 3. Epics — try multiple JQL queries
        epics = []
        epic_type_names = []
        for it in issue_types:
            name_lower = it["name"].lower()
            if name_lower == "epic" or it.get("hierarchyLevel", 0) >= 1:
                epic_type_names.append(it["name"])
        if not epic_type_names:
            epic_type_names = ["Epic"]  # fallback

        debug_log.append(f"epic type names: {epic_type_names}")

        for epic_type_name in epic_type_names:
            # Try 1: exclude Done epics
            try:
                jql = f'project = {jira_project} AND issuetype = "{epic_type_name}" AND statusCategory != Done ORDER BY summary ASC'
                logger.info(f"Fetching epics with JQL: {jql}")
                epic_data = _jira_search_jql(jql, ["summary", "status"],
                                              jira_domain, jira_email, jira_token, max_results=100)
                debug_log.append(f"epic search ({epic_type_name}, not Done): {epic_data.get('total', 0)} results")
                for issue in epic_data.get("issues", []):
                    epics.append({
                        "key": issue["key"],
                        "summary": issue["fields"]["summary"],
                        "status": issue["fields"]["status"]["name"] if issue["fields"].get("status") else "",
                    })
                if epics:
                    break
            except Exception as e:
                debug_log.append(f"epic search ({epic_type_name}, not Done) failed: {e}")

            # Try 2: include ALL epics (maybe they're all Done or status category doesn't work)
            if not epics:
                try:
                    jql2 = f'project = {jira_project} AND issuetype = "{epic_type_name}" ORDER BY summary ASC'
                    logger.info(f"Fetching ALL epics with JQL: {jql2}")
                    epic_data2 = _jira_search_jql(jql2, ["summary", "status"],
                                                   jira_domain, jira_email, jira_token, max_results=100)
                    debug_log.append(f"epic search ({epic_type_name}, ALL): {epic_data2.get('total', 0)} results")
                    for issue in epic_data2.get("issues", []):
                        epics.append({
                            "key": issue["key"],
                            "summary": issue["fields"]["summary"],
                            "status": issue["fields"]["status"]["name"] if issue["fields"].get("status") else "",
                        })
                    if epics:
                        break
                except Exception as e:
                    debug_log.append(f"epic search ({epic_type_name}, ALL) failed: {e}")

        logger.info(f"Found {len(epics)} epics for {jira_project}")

        # 4. Components
        components = []
        try:
            comp_data = _jira_request("GET", f"/project/{jira_project}/components",
                                      jira_domain, jira_email, jira_token)
            for c in comp_data:
                components.append({"id": c["id"], "name": c["name"]})
        except Exception as e:
            debug_log.append(f"components failed: {e}")

        logger.info(f"Jira options for {jira_project}: {len(issue_types)} types, {len(epics)} epics, {len(priorities)} priorities, {len(components)} components")
        logger.info(f"Jira options debug: {' | '.join(debug_log)}")

        return jsonify({
            "ok": True,
            "issue_types": issue_types,
            "priorities": priorities,
            "epics": epics,
            "components": components,
            "project_key": jira_project,
            "debug": " | ".join(debug_log),
        })
    except Exception as e:
        logger.error(f"Jira options error: {e}")
        return jsonify({"ok": False, "message": str(e)[:300]})


@app.route("/api/jira/search-issues")
def jira_search_issues():
    """Search Jira issues by key or summary for parent ticket selection.

    Query params:
        q: search text (min 2 chars)
        parent_types: comma-separated issue type names to filter (e.g. "Epic" or "Story,Task,Bug")
    """
    db = get_db()
    jira_domain = get_setting("jira_domain", "", db=db)
    jira_email = get_setting("jira_email", "", db=db)
    jira_token = get_setting("jira_api_token", "", db=db)
    jira_project = get_setting("jira_project", "LUX", db=db)

    q = request.args.get("q", "").strip()
    parent_types_raw = request.args.get("parent_types", "").strip()

    if not q or len(q) < 2:
        return jsonify({"ok": True, "issues": []})

    if not jira_domain or not jira_email or not jira_token:
        return jsonify({"ok": False, "message": "Jira not configured"})

    try:
        # Build JQL with optional issue type filter
        type_filter = ""
        if parent_types_raw:
            type_names = [t.strip() for t in parent_types_raw.split(",") if t.strip()]
            if type_names:
                quoted = ", ".join(f'"{t}"' for t in type_names)
                type_filter = f" AND issuetype IN ({quoted})"

        if q.upper().startswith(jira_project.upper() + "-"):
            jql = f'project = {jira_project} AND key = "{q.upper()}"{type_filter} ORDER BY updated DESC'
        else:
            safe_q = q.replace('"', '\\"')
            jql = f'project = {jira_project} AND summary ~ "{safe_q}"{type_filter} AND statusCategory != Done ORDER BY updated DESC'

        logger.info(f"Jira search: q={q}, JQL={jql}")

        data = _jira_search_jql(jql, ["summary", "issuetype", "status"],
                                jira_domain, jira_email, jira_token, max_results=15)

        issues = []
        for issue in data.get("issues", []):
            it = issue["fields"].get("issuetype", {})
            issues.append({
                "key": issue["key"],
                "summary": issue["fields"]["summary"],
                "type": it.get("name", ""),
                "type_icon": it.get("iconUrl", ""),
                "subtask": bool(it.get("subtask")),
                "status": issue["fields"].get("status", {}).get("name", ""),
            })

        return jsonify({"ok": True, "issues": issues})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:200], "issues": []})


# ── Export to Excel ───────────────────────────────────────────────────────────

@app.route("/export-excel")
def export_excel():
    """Export the ticket dashboard to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    db = get_db()
    tickets = db.execute("SELECT * FROM tickets ORDER BY updated_at DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ticket Dashboard"

    # Header style
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="5046E5", end_color="5046E5", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    headers = ["#", "Subject", "Status", "Requester", "Client", "Type", "Risk",
               "RICE", "PO Decision", "Inbox", "Summary", "Review", "Created", "Updated"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Data rows
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(vertical="center", wrap_text=False)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for row_idx, t in enumerate(tickets, 2):
        company = extract_company(t["requester_email"]) if t["requester_email"] else ""
        row_data = [
            t["ticket_id"],
            t["subject"] or "",
            t["status"] or "",
            t["requester_name"] or "",
            company,
            t["classification"] or "",
            t["risk_level"] or "",
            round(t["rice_score"], 1) if t["rice_score"] else 0,
            (t["po_decision"] or "pending").title(),
            t["responded"] or "In Inbox",
            t["summary"] or "",
            t["review_status"] or "",
            (t["created_at"] or "")[:10],
            (t["updated_at"] or "")[:10],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    widths = [10, 40, 14, 18, 14, 12, 8, 8, 12, 10, 50, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:N{len(tickets) + 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"ticket_dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )


# ── BSO Kanban Board ─────────────────────────────────────────────────────────

BSO_COLUMNS = ["backlog", "to_do", "in_progress", "done"]
BSO_COLUMN_LABELS = {"backlog": "Backlog", "to_do": "To Do", "in_progress": "In Progress", "done": "Done"}

@app.route("/bso")
def bso_board():
    """BSO kanban board — tickets grouped by work status with Jira ship actions."""
    db = get_db()

    # Only show tickets that have been approved by the PO (or all if no filter)
    show_all = request.args.get("show_all", "")
    classification_filter = request.args.get("classification", "")
    client_filter = request.args.get("client", "")

    query = "SELECT * FROM tickets WHERE 1=1"
    params = []

    if not show_all:
        query += " AND LOWER(COALESCE(po_decision, 'pending')) = 'approved'"

    if classification_filter:
        query += " AND classification = ?"
        params.append(classification_filter)

    if client_filter:
        query += " AND LOWER(requester_email) LIKE ?"
        params.append(f"%@{client_filter.lower()}%")

    query += " ORDER BY rice_score DESC, updated_at DESC"
    tickets = [dict(r) for r in db.execute(query, params).fetchall()]

    # Auto-assign bso_status for tickets that don't have one
    for t in tickets:
        if not t.get("bso_status"):
            # Infer from Jira / ticket state
            if t.get("status") in ("Resolved", "Closed"):
                t["bso_status"] = "done"
            elif t.get("jira_ticket_key"):
                t["bso_status"] = "in_progress"
            else:
                t["bso_status"] = "backlog"

    # Group into columns
    columns = {col: [] for col in BSO_COLUMNS}
    for t in tickets:
        col = t.get("bso_status", "backlog")
        if col not in columns:
            col = "backlog"
        columns[col].append(t)

    # Get unique clients for filter
    all_clients = db.execute(
        "SELECT DISTINCT requester_email FROM tickets WHERE requester_email IS NOT NULL AND requester_email != ''"
    ).fetchall()
    companies = sorted(set(extract_company(r["requester_email"]) for r in all_clients if r["requester_email"]))

    return render_template("bso.html",
                           columns=columns,
                           column_labels=BSO_COLUMN_LABELS,
                           column_order=BSO_COLUMNS,
                           total_tickets=len(tickets),
                           companies=companies,
                           filters={"show_all": show_all, "classification": classification_filter, "client": client_filter},
                           extract_company=extract_company)


@app.route("/api/bso/move", methods=["POST"])
def bso_move_ticket():
    """Move a ticket to a different BSO column."""
    data = request.get_json(silent=True) or {}
    ticket_id = data.get("ticket_id")
    new_status = data.get("status", "").strip()

    if not ticket_id or new_status not in BSO_COLUMNS:
        return jsonify({"ok": False, "message": "Invalid ticket_id or status"})

    db = get_db()
    db.execute("UPDATE tickets SET bso_status = ? WHERE ticket_id = ?", (new_status, ticket_id))
    db.commit()
    return jsonify({"ok": True, "message": f"Ticket {ticket_id} moved to {BSO_COLUMN_LABELS.get(new_status, new_status)}"})


@app.route("/api/bso/bulk-ship-jira", methods=["POST"])
def bso_bulk_ship_jira():
    """Ship one or more tickets to Jira. Accepts overrides for Jira fields."""
    data = request.get_json(silent=True) or {}
    ticket_ids = data.get("ticket_ids", [])

    if not ticket_ids:
        return jsonify({"ok": False, "message": "No tickets provided"})

    # Overrides apply to all tickets in the batch
    overrides = {
        "issue_type": data.get("issue_type", ""),
        "priority": data.get("priority", ""),
        "epic_key": data.get("epic_key", ""),
        "parent_key": data.get("parent_key", ""),
        "components": data.get("components", []),
        "labels": data.get("labels", []),
        "summary": data.get("summary", ""),
    }

    db = get_db()
    results = []
    for tid in ticket_ids:
        ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (tid,)).fetchone()
        if not ticket:
            results.append({"ticket_id": tid, "ok": False, "message": "Not found"})
            continue

        existing_key = _row_get(ticket, "jira_ticket_key", "")
        if existing_key:
            results.append({"ticket_id": tid, "ok": True, "message": f"Already linked: {existing_key}", "key": existing_key})
            continue

        try:
            ticket_dict = dict(ticket)
            jira_key, jira_url = create_jira_ticket_from_freshdesk(ticket_dict, db=db, overrides=overrides)
            db.execute("UPDATE tickets SET bso_status = 'in_progress' WHERE ticket_id = ?", (tid,))
            db.commit()
            results.append({"ticket_id": tid, "ok": True, "message": f"Created {jira_key}", "key": jira_key, "url": jira_url})
        except Exception as e:
            results.append({"ticket_id": tid, "ok": False, "message": str(e)[:200]})

    success_count = sum(1 for r in results if r["ok"])
    return jsonify({"ok": True, "message": f"{success_count}/{len(ticket_ids)} shipped to Jira", "results": results})


# ── Reporting Page ────────────────────────────────────────────────────────────

@app.route("/reporting")
def reporting():
    """Reporting dashboard with charts and metrics for stakeholders."""
    db = get_db()

    # Filters
    client_filter = request.args.get("client", "")
    requester_filter = request.args.get("requester", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    where = "WHERE 1=1"
    params = []
    if client_filter:
        where += " AND LOWER(requester_email) LIKE ?"
        params.append(f"%@{client_filter.lower()}%")
    if requester_filter:
        where += " AND requester_name = ?"
        params.append(requester_filter)
    if date_from:
        where += " AND created_at >= ?"
        params.append(date_from)
    if date_to:
        where += " AND created_at <= ?"
        params.append(date_to + "T23:59:59")

    # ── Key Metrics ──
    total = db.execute(f"SELECT COUNT(*) as c FROM tickets {where}", params).fetchone()["c"]
    open_count = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Open', 'In Progress', 'Pending Approval')", params).fetchone()["c"]
    resolved_count = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed')", params).fetchone()["c"]
    avg_rice = db.execute(f"SELECT AVG(rice_score) as avg FROM tickets {where} AND rice_score > 0", params).fetchone()["avg"] or 0

    # ── Status Distribution ──
    status_dist = db.execute(
        f"SELECT status, COUNT(*) as count FROM tickets {where} GROUP BY status ORDER BY count DESC", params
    ).fetchall()

    # ── Classification Distribution ──
    class_dist = db.execute(
        f"SELECT classification, COUNT(*) as count FROM tickets {where} AND classification IS NOT NULL AND classification != '' GROUP BY classification ORDER BY count DESC", params
    ).fetchall()

    # ── Risk Distribution ──
    risk_dist = db.execute(
        f"SELECT risk_level, COUNT(*) as count FROM tickets {where} AND risk_level IS NOT NULL GROUP BY risk_level ORDER BY count DESC", params
    ).fetchall()

    # ── PO Decision Distribution ──
    po_dist = db.execute(
        f"SELECT po_decision, COUNT(*) as count FROM tickets {where} GROUP BY po_decision ORDER BY count DESC", params
    ).fetchall()

    # ── Client Breakdown ──
    all_tickets_emails = db.execute(
        f"SELECT requester_email FROM tickets {where}", params
    ).fetchall()
    client_counts = {}
    for row in all_tickets_emails:
        comp = extract_company(row["requester_email"])
        client_counts[comp] = client_counts.get(comp, 0) + 1
    client_breakdown = sorted(client_counts.items(), key=lambda x: x[1], reverse=True)[:15]

    # ── Resolution Time (for resolved/closed tickets) ──
    resolved_tickets = db.execute(
        f"SELECT created_at, resolved_at, sla_resolution_hours FROM tickets {where} AND status IN ('Resolved', 'Closed') AND created_at IS NOT NULL", params
    ).fetchall()
    resolution_times = []
    for t in resolved_tickets:
        try:
            if t["sla_resolution_hours"] and t["sla_resolution_hours"] > 0:
                resolution_times.append(t["sla_resolution_hours"])
            elif t["resolved_at"]:
                created = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
                resolved = datetime.fromisoformat(t["resolved_at"].replace("Z", "+00:00"))
                delta_hours = (resolved - created).total_seconds() / 3600
                resolution_times.append(delta_hours)
        except (ValueError, TypeError, AttributeError):
            pass

    avg_resolution_hours = sum(resolution_times) / len(resolution_times) if resolution_times else 0
    avg_resolution_days = avg_resolution_hours / 24

    # Resolution time buckets
    res_buckets = {"< 1 day": 0, "1-3 days": 0, "3-7 days": 0, "1-2 weeks": 0, "2-4 weeks": 0, "> 4 weeks": 0}
    for h in resolution_times:
        days = h / 24
        if days < 1:
            res_buckets["< 1 day"] += 1
        elif days < 3:
            res_buckets["1-3 days"] += 1
        elif days < 7:
            res_buckets["3-7 days"] += 1
        elif days < 14:
            res_buckets["1-2 weeks"] += 1
        elif days < 28:
            res_buckets["2-4 weeks"] += 1
        else:
            res_buckets["> 4 weeks"] += 1

    # ── First Response Time (SLA) ──
    first_response_tickets = db.execute(
        f"SELECT sla_first_response_hours FROM tickets {where} AND sla_first_response_hours > 0", params
    ).fetchall()
    first_response_times = [t["sla_first_response_hours"] for t in first_response_tickets]
    avg_first_response_hours = sum(first_response_times) / len(first_response_times) if first_response_times else 0

    # First response time buckets
    first_response_buckets = {"< 1 hour": 0, "1-4 hours": 0, "4-24 hours": 0, "1-3 days": 0, "> 3 days": 0}
    for h in first_response_times:
        if h < 1:
            first_response_buckets["< 1 hour"] += 1
        elif h < 4:
            first_response_buckets["1-4 hours"] += 1
        elif h < 24:
            first_response_buckets["4-24 hours"] += 1
        elif h < 72:
            first_response_buckets["1-3 days"] += 1
        else:
            first_response_buckets["> 3 days"] += 1

    # ── Template Breakdown ──
    template_breakdown = db.execute(
        f"SELECT template_name, COUNT(*) as count FROM tickets {where} AND template_name != '' GROUP BY template_name ORDER BY count DESC LIMIT 10", params
    ).fetchall()

    # ── Workflow Breakdown ──
    workflow_breakdown = db.execute(
        f"SELECT workflow_name, COUNT(*) as count FROM tickets {where} AND workflow_name != '' GROUP BY workflow_name ORDER BY count DESC", params
    ).fetchall()

    # ── SLA Compliance ──
    resolved_for_sla = db.execute(
        f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed') AND created_at IS NOT NULL", params
    ).fetchone()["c"]
    resolved_within_7d = db.execute(
        f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed') AND sla_resolution_hours > 0 AND sla_resolution_hours <= 168", params
    ).fetchone()["c"]
    sla_resolution_pct = round(100 * resolved_within_7d / resolved_for_sla, 1) if resolved_for_sla > 0 else 0

    first_response_for_sla = db.execute(
        f"SELECT COUNT(*) as c FROM tickets {where} AND sla_first_response_hours > 0", params
    ).fetchone()["c"]
    first_response_within_24h = db.execute(
        f"SELECT COUNT(*) as c FROM tickets {where} AND sla_first_response_hours > 0 AND sla_first_response_hours <= 24", params
    ).fetchone()["c"]
    sla_response_pct = round(100 * first_response_within_24h / first_response_for_sla, 1) if first_response_for_sla > 0 else 0

    # ── Period Stats (this week, last 7 days) ──
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    tickets_this_week = db.execute(
        f"SELECT COUNT(*) as c FROM tickets {where} AND created_at > ?", params + [week_ago]
    ).fetchone()["c"]
    resolved_this_week = db.execute(
        f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed') AND resolved_at > ?", params + [week_ago]
    ).fetchone()["c"]

    # ── RICE Score Distribution ──
    rice_tickets = db.execute(
        f"SELECT rice_score FROM tickets {where} AND rice_score > 0", params
    ).fetchall()
    rice_buckets = {"Low (0-5)": 0, "Medium (5-15)": 0, "High (15-30)": 0, "Very High (30+)": 0}
    for t in rice_tickets:
        s = t["rice_score"]
        if s < 5:
            rice_buckets["Low (0-5)"] += 1
        elif s < 15:
            rice_buckets["Medium (5-15)"] += 1
        elif s < 30:
            rice_buckets["High (15-30)"] += 1
        else:
            rice_buckets["Very High (30+)"] += 1

    # ── Tickets over time (by month) ──
    monthly = db.execute(
        f"SELECT SUBSTR(created_at, 1, 7) as month, COUNT(*) as count FROM tickets {where} AND created_at IS NOT NULL GROUP BY month ORDER BY month", params
    ).fetchall()

    # ── Raw dates for client-side period grouping ──
    ticket_dates = db.execute(
        f"SELECT created_at FROM tickets {where} AND created_at IS NOT NULL ORDER BY created_at", params
    ).fetchall()
    ticket_dates_list = [r["created_at"] for r in ticket_dates]

    # ── Review Status ──
    review_dist = db.execute(
        f"SELECT review_status, COUNT(*) as count FROM tickets {where} GROUP BY review_status ORDER BY count DESC", params
    ).fetchall()

    # Build filter dropdown data
    all_emails = db.execute("SELECT DISTINCT requester_email FROM tickets WHERE requester_email IS NOT NULL AND requester_email != ''").fetchall()
    company_set = {}
    for row in all_emails:
        email = row["requester_email"]
        comp = extract_company(email)
        domain = email.split("@")[1].split(".")[0].lower() if "@" in email else ""
        if comp != "Unknown" and domain:
            company_set[domain] = comp
    companies = sorted(company_set.items(), key=lambda x: x[1])

    all_requesters = db.execute(
        "SELECT DISTINCT requester_name FROM tickets WHERE requester_name IS NOT NULL AND requester_name != '' ORDER BY requester_name"
    ).fetchall()
    requesters_list = [r["requester_name"] for r in all_requesters]

    return render_template("reporting.html",
        total=total, open_count=open_count, resolved_count=resolved_count,
        avg_rice=round(avg_rice, 1),
        avg_resolution_days=round(avg_resolution_days, 1),
        avg_resolution_hours=round(avg_resolution_hours, 1),
        avg_first_response_hours=round(avg_first_response_hours, 1),
        status_dist=[dict(r) for r in status_dist],
        class_dist=[dict(r) for r in class_dist],
        risk_dist=[dict(r) for r in risk_dist],
        po_dist=[dict(r) for r in po_dist],
        client_breakdown=client_breakdown,
        res_buckets=res_buckets,
        first_response_buckets=first_response_buckets,
        rice_buckets=rice_buckets,
        template_breakdown=[[r["template_name"], r["count"]] for r in template_breakdown],
        workflow_breakdown=[[r["workflow_name"], r["count"]] for r in workflow_breakdown],
        sla_resolution_pct=sla_resolution_pct,
        sla_response_pct=sla_response_pct,
        tickets_this_week=tickets_this_week,
        resolved_this_week=resolved_this_week,
        monthly=[dict(r) for r in monthly],
        ticket_dates=ticket_dates_list,
        review_dist=[dict(r) for r in review_dist],
        companies=companies,
        requesters=requesters_list,
        filters={"client": client_filter, "requester": requester_filter,
                 "date_from": date_from, "date_to": date_to},
    )


# ── Google Workspace Integration ─────────────────────────────────────────────

try:
    from google_integration import (
        test_connection as google_test_conn,
        export_report_to_sheets, export_tickets_to_sheets,
        export_analysis_to_docs, export_report_to_slides,
        list_drive_kb_files, get_kb_context_from_drive,
    )
    GOOGLE_AVAILABLE = True
except ImportError:
    GOOGLE_AVAILABLE = False
    logging.warning("Google API libraries not installed. Google Workspace features disabled. Run: pip3 install google-api-python-client google-auth google-auth-httplib2")


@app.route("/google-test-connection", methods=["POST"])
def google_test_connection():
    """Test Google Drive service account connection."""
    if not GOOGLE_AVAILABLE:
        return jsonify({"ok": False, "error": "Google API libraries not installed. Run: pip3 install google-api-python-client google-auth google-auth-httplib2"})
    db = get_db()
    sa_json = get_setting("google_sa_json", db=db)
    if not sa_json:
        return jsonify({"ok": False, "error": "No service account key configured. Add it in Settings first."})

    export_folder = get_setting("google_export_folder", db=db) or ""
    kb_folder = get_setting("google_kb_folder", db=db) or ""

    result = google_test_conn(sa_json, export_folder)
    if not result["ok"]:
        return jsonify(result)

    resp = {"ok": True, "email": result["email"], "export_folder": result.get("folder_name", ""), "kb_folder": "", "kb_files": 0}

    if kb_folder:
        kb_result = google_test_conn(sa_json, kb_folder)
        if kb_result["ok"]:
            resp["kb_folder"] = kb_result.get("folder_name", "")
            try:
                files = list_drive_kb_files(sa_json, kb_folder, max_results=50)
                resp["kb_files"] = len(files)
            except Exception:
                resp["kb_files"] = 0
        else:
            resp["ok"] = False
            resp["error"] = f"Export folder OK but KB folder error: {kb_result['error']}"

    return jsonify(resp)


@app.route("/google-export-report-sheets")
def google_export_report_sheets():
    """Export reporting data to a new Google Sheet."""
    if not GOOGLE_AVAILABLE:
        flash("Google API libraries not installed. Restart the app to install them.", "error")
        return redirect(url_for("reporting"))
    db = get_db()
    sa_json = get_setting("google_sa_json", db=db)
    folder_id = get_setting("google_export_folder", db=db) or ""
    if not sa_json:
        flash("Google Drive not configured. Set up in Settings first.", "error")
        return redirect(url_for("reporting"))

    sections = [s.strip() for s in request.args.get("sections", "").split(",") if s.strip()]
    if not sections:
        sections = ["kpi_cards", "status_chart", "classification_chart", "risk_chart",
                     "po_decisions", "resolution_time", "first_response", "rice_chart",
                     "templates", "workflows", "timeline", "client_breakdown", "sla_compliance"]

    # Reuse the same data-gathering logic from export_report
    data = _gather_report_data(db, request.args)

    try:
        url = export_report_to_sheets(sa_json, folder_id, data, sections)
        flash(f"Report exported to Google Sheets!", "success")
        return redirect(url)
    except Exception as e:
        flash(f"Google Sheets export failed: {e}", "error")
        return redirect(url_for("reporting"))


@app.route("/google-export-tickets-sheets")
def google_export_tickets_sheets():
    """Export all ticket data to a new Google Sheet."""
    if not GOOGLE_AVAILABLE:
        flash("Google API libraries not installed. Restart the app to install them.", "error")
        return redirect(url_for("reporting"))
    db = get_db()
    sa_json = get_setting("google_sa_json", db=db)
    folder_id = get_setting("google_export_folder", db=db) or ""
    if not sa_json:
        flash("Google Drive not configured. Set up in Settings first.", "error")
        return redirect(url_for("reporting"))

    # Gather tickets with same filters
    client_filter = request.args.get("client", "")
    requester_filter = request.args.get("requester", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    where = "WHERE 1=1"
    params = []
    if client_filter:
        where += " AND LOWER(requester_email) LIKE ?"
        params.append(f"%@{client_filter.lower()}%")
    if requester_filter:
        where += " AND requester_name = ?"
        params.append(requester_filter)
    if date_from:
        where += " AND created_at >= ?"
        params.append(date_from)
    if date_to:
        where += " AND created_at <= ?"
        params.append(date_to + "T23:59:59")

    tickets = [dict(r) for r in db.execute(
        f"SELECT * FROM tickets {where} ORDER BY created_at DESC", params
    ).fetchall()]

    try:
        url = export_tickets_to_sheets(sa_json, folder_id, tickets)
        flash(f"Tickets exported to Google Sheets!", "success")
        return redirect(url)
    except Exception as e:
        flash(f"Google Sheets export failed: {e}", "error")
        return redirect(url_for("reporting"))


@app.route("/google-export-analysis-doc/<int:ticket_id>")
def google_export_analysis_doc(ticket_id):
    """Export a single ticket's full analysis to Google Docs."""
    if not GOOGLE_AVAILABLE:
        flash("Google API libraries not installed. Restart the app to install them.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))
    db = get_db()
    sa_json = get_setting("google_sa_json", db=db)
    folder_id = get_setting("google_export_folder", db=db) or ""
    if not sa_json:
        flash("Google Drive not configured. Set up in Settings first.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    try:
        url = export_analysis_to_docs(sa_json, folder_id, dict(ticket))
        flash(f"Analysis exported to Google Docs!", "success")
        return redirect(url)
    except Exception as e:
        flash(f"Google Docs export failed: {e}", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))


@app.route("/google-export-report-slides")
def google_export_report_slides():
    """Export reporting data to a new Google Slides presentation."""
    if not GOOGLE_AVAILABLE:
        flash("Google API libraries not installed. Restart the app to install them.", "error")
        return redirect(url_for("reporting"))
    db = get_db()
    sa_json = get_setting("google_sa_json", db=db)
    folder_id = get_setting("google_export_folder", db=db) or ""
    if not sa_json:
        flash("Google Drive not configured. Set up in Settings first.", "error")
        return redirect(url_for("reporting"))

    sections = [s.strip() for s in request.args.get("sections", "").split(",") if s.strip()]
    if not sections:
        sections = ["kpi_cards", "status_chart", "classification_chart", "risk_chart",
                     "po_decisions", "resolution_time", "first_response", "rice_chart",
                     "templates", "workflows", "timeline", "client_breakdown", "sla_compliance"]

    data = _gather_report_data(db, request.args)

    try:
        url = export_report_to_slides(sa_json, folder_id, data, sections)
        flash(f"Report exported to Google Slides!", "success")
        return redirect(url)
    except Exception as e:
        flash(f"Google Slides export failed: {e}", "error")
        return redirect(url_for("reporting"))


# ── Notion Integration ───────────────────────────────────────────────────────

from notion_integration import (
    test_notion_connection as notion_test_conn,
    export_analysis_to_notion,
)


@app.route("/notion-test-connection", methods=["POST"])
def notion_test_connection():
    """Test Notion integration token and page access."""
    db = get_db()
    token = get_setting("notion_token", db=db)
    if not token:
        return jsonify({"ok": False, "error": "No Notion token configured. Add it in Settings first."})
    page_id = get_setting("notion_page_id", db=db) or ""
    result = notion_test_conn(token, page_id)
    return jsonify(result)


@app.route("/notion-export-analysis/<int:ticket_id>")
def notion_export_analysis(ticket_id):
    """Export a ticket's full analysis to Notion as a rich page."""
    db = get_db()
    token = get_setting("notion_token", db=db)
    page_id = get_setting("notion_page_id", db=db)
    if not token or not page_id:
        flash("Notion not configured. Set up token and page ID in Settings first.", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))

    ticket = db.execute("SELECT * FROM tickets WHERE ticket_id = ?", (ticket_id,)).fetchone()
    if not ticket:
        flash("Ticket not found.", "error")
        return redirect(url_for("dashboard"))

    try:
        url = export_analysis_to_notion(token, page_id, dict(ticket))
        flash("Analysis exported to Notion!", "success")
        return redirect(url)
    except Exception as e:
        flash(f"Notion export failed: {e}", "error")
        return redirect(url_for("ticket_detail", ticket_id=ticket_id))


def _gather_report_data(db, args):
    """Gather all reporting data into a dict (shared by export routes)."""
    client_filter = args.get("client", "")
    requester_filter = args.get("requester", "")
    date_from = args.get("date_from", "")
    date_to = args.get("date_to", "")

    where = "WHERE 1=1"
    params = []
    if client_filter:
        where += " AND LOWER(requester_email) LIKE ?"
        params.append(f"%@{client_filter.lower()}%")
    if requester_filter:
        where += " AND requester_name = ?"
        params.append(requester_filter)
    if date_from:
        where += " AND created_at >= ?"
        params.append(date_from)
    if date_to:
        where += " AND created_at <= ?"
        params.append(date_to + "T23:59:59")

    total = db.execute(f"SELECT COUNT(*) as c FROM tickets {where}", params).fetchone()["c"]
    open_count = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Open', 'In Progress', 'Pending Approval')", params).fetchone()["c"]
    resolved_count = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed')", params).fetchone()["c"]
    avg_rice = db.execute(f"SELECT AVG(rice_score) as avg FROM tickets {where} AND rice_score > 0", params).fetchone()["avg"] or 0

    status_dist = [dict(r) for r in db.execute(f"SELECT status, COUNT(*) as count FROM tickets {where} GROUP BY status ORDER BY count DESC", params).fetchall()]
    class_dist = [dict(r) for r in db.execute(f"SELECT classification, COUNT(*) as count FROM tickets {where} AND classification IS NOT NULL AND classification != '' GROUP BY classification ORDER BY count DESC", params).fetchall()]
    risk_dist = [dict(r) for r in db.execute(f"SELECT risk_level, COUNT(*) as count FROM tickets {where} AND risk_level IS NOT NULL GROUP BY risk_level ORDER BY count DESC", params).fetchall()]
    po_dist = [dict(r) for r in db.execute(f"SELECT po_decision, COUNT(*) as count FROM tickets {where} GROUP BY po_decision ORDER BY count DESC", params).fetchall()]

    resolved_tickets = db.execute(f"SELECT created_at, resolved_at, sla_resolution_hours FROM tickets {where} AND status IN ('Resolved', 'Closed') AND created_at IS NOT NULL", params).fetchall()
    resolution_times = []
    for t in resolved_tickets:
        try:
            if t["sla_resolution_hours"] and t["sla_resolution_hours"] > 0:
                resolution_times.append(t["sla_resolution_hours"])
            elif t["resolved_at"]:
                created = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
                resolved = datetime.fromisoformat(t["resolved_at"].replace("Z", "+00:00"))
                resolution_times.append((resolved - created).total_seconds() / 3600)
        except (ValueError, TypeError, AttributeError):
            pass
    avg_resolution_hours = sum(resolution_times) / len(resolution_times) if resolution_times else 0
    avg_resolution_days = avg_resolution_hours / 24

    res_buckets = {"< 1 day": 0, "1-3 days": 0, "3-7 days": 0, "1-2 weeks": 0, "2-4 weeks": 0, "> 4 weeks": 0}
    for h in resolution_times:
        days = h / 24
        if days < 1: res_buckets["< 1 day"] += 1
        elif days < 3: res_buckets["1-3 days"] += 1
        elif days < 7: res_buckets["3-7 days"] += 1
        elif days < 14: res_buckets["1-2 weeks"] += 1
        elif days < 28: res_buckets["2-4 weeks"] += 1
        else: res_buckets["> 4 weeks"] += 1

    first_response_tickets = db.execute(f"SELECT sla_first_response_hours FROM tickets {where} AND sla_first_response_hours > 0", params).fetchall()
    first_response_times = [t["sla_first_response_hours"] for t in first_response_tickets]
    avg_first_response_hours = sum(first_response_times) / len(first_response_times) if first_response_times else 0

    first_response_buckets = {"< 1 hour": 0, "1-4 hours": 0, "4-24 hours": 0, "1-3 days": 0, "> 3 days": 0}
    for h in first_response_times:
        if h < 1: first_response_buckets["< 1 hour"] += 1
        elif h < 4: first_response_buckets["1-4 hours"] += 1
        elif h < 24: first_response_buckets["4-24 hours"] += 1
        elif h < 72: first_response_buckets["1-3 days"] += 1
        else: first_response_buckets["> 3 days"] += 1

    template_breakdown = [[r["template_name"], r["count"]] for r in db.execute(f"SELECT template_name, COUNT(*) as count FROM tickets {where} AND template_name != '' GROUP BY template_name ORDER BY count DESC LIMIT 10", params).fetchall()]
    workflow_breakdown = [[r["workflow_name"], r["count"]] for r in db.execute(f"SELECT workflow_name, COUNT(*) as count FROM tickets {where} AND workflow_name != '' GROUP BY workflow_name ORDER BY count DESC", params).fetchall()]

    resolved_for_sla = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed') AND created_at IS NOT NULL", params).fetchone()["c"]
    resolved_within_7d = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND status IN ('Resolved', 'Closed') AND sla_resolution_hours > 0 AND sla_resolution_hours <= 168", params).fetchone()["c"]
    sla_resolution_pct = round(100 * resolved_within_7d / resolved_for_sla, 1) if resolved_for_sla > 0 else 0

    first_response_for_sla = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND sla_first_response_hours > 0", params).fetchone()["c"]
    first_response_within_24h = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND sla_first_response_hours > 0 AND sla_first_response_hours <= 24", params).fetchone()["c"]
    sla_response_pct = round(100 * first_response_within_24h / first_response_for_sla, 1) if first_response_for_sla > 0 else 0

    rice_tickets = db.execute(f"SELECT rice_score FROM tickets {where} AND rice_score > 0", params).fetchall()
    rice_buckets = {"Low (0-5)": 0, "Medium (5-15)": 0, "High (15-30)": 0, "Very High (30+)": 0}
    for t in rice_tickets:
        s = t["rice_score"]
        if s < 5: rice_buckets["Low (0-5)"] += 1
        elif s < 15: rice_buckets["Medium (5-15)"] += 1
        elif s < 30: rice_buckets["High (15-30)"] += 1
        else: rice_buckets["Very High (30+)"] += 1

    monthly = [dict(r) for r in db.execute(f"SELECT SUBSTR(created_at, 1, 7) as month, COUNT(*) as count FROM tickets {where} AND created_at IS NOT NULL GROUP BY month ORDER BY month", params).fetchall()]

    all_tickets_emails = db.execute(f"SELECT requester_email FROM tickets {where}", params).fetchall()
    client_counts = {}
    for row in all_tickets_emails:
        comp = extract_company(row["requester_email"])
        client_counts[comp] = client_counts.get(comp, 0) + 1
    client_breakdown = sorted(client_counts.items(), key=lambda x: x[1], reverse=True)[:15]

    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    tickets_this_week = db.execute(f"SELECT COUNT(*) as c FROM tickets {where} AND created_at > ?", params + [week_ago]).fetchone()["c"]

    return {
        "total": total, "open_count": open_count, "resolved_count": resolved_count,
        "avg_rice": round(avg_rice, 1), "avg_resolution_days": round(avg_resolution_days, 1),
        "avg_resolution_hours": round(avg_resolution_hours, 1),
        "avg_first_response_hours": round(avg_first_response_hours, 1),
        "sla_resolution_pct": sla_resolution_pct, "sla_response_pct": sla_response_pct,
        "tickets_this_week": tickets_this_week,
        "status_dist": status_dist, "class_dist": class_dist, "risk_dist": risk_dist, "po_dist": po_dist,
        "res_buckets": res_buckets, "first_response_buckets": first_response_buckets,
        "rice_buckets": rice_buckets, "template_breakdown": template_breakdown,
        "workflow_breakdown": workflow_breakdown, "monthly": monthly,
        "client_breakdown": client_breakdown,
        "filter_client": client_filter, "filter_requester": requester_filter,
        "filter_date_from": date_from, "filter_date_to": date_to,
    }


# ── Export Report ────────────────────────────────────────────────────────────

@app.route("/export-report")
def export_report():
    """Generate and download PDF or PPTX report with selected sections."""
    db = get_db()
    fmt = request.args.get("format", "pdf").lower()
    sections = [s.strip() for s in request.args.get("sections", "").split(",") if s.strip()]
    if not sections:
        sections = ["kpi_cards", "status_chart", "classification_chart", "risk_chart",
                     "po_decisions", "resolution_time", "first_response", "rice_chart",
                     "templates", "workflows", "timeline", "client_breakdown", "sla_compliance"]

    data = _gather_report_data(db, request.args)

    today = datetime.now().strftime("%Y-%m-%d")
    if fmt == "pptx":
        buf = generate_pptx(data, sections)
        filename = f"BSO_LUX_Report_{today}.pptx"
        mime = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    else:
        buf = generate_pdf(data, sections)
        filename = f"BSO_LUX_Report_{today}.pdf"
        mime = "application/pdf"

    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype=mime)


# ── Agent Dashboard ──────────────────────────────────────────────────────────

@app.route("/agents")
def agent_dashboard():
    """Agent monitoring dashboard: logs, costs, lessons learned."""
    db = get_db()
    days = int(request.args.get("days", 7))

    orchestrator = AgentOrchestrator(get_setting("anthropic_api_key", db=db) or "dummy", db=db)

    cost_summary = orchestrator.get_cost_summary(days=days)
    lessons_rows = orchestrator.get_lessons(active_only=True, limit=100)
    logs = orchestrator.get_agent_logs(limit=100)

    # Convert sqlite3.Row to dicts for template
    cost_summary = [dict(r) if hasattr(r, "keys") else r for r in cost_summary]
    lessons = [dict(r) if hasattr(r, "keys") else r for r in lessons_rows]
    logs = [dict(r) if hasattr(r, "keys") else r for r in logs]

    # Compute aggregate stats
    total_cost = sum(r.get("total_cost", 0) or 0 for r in cost_summary)
    total_calls = sum(r.get("calls", 0) or 0 for r in cost_summary)
    total_tokens = sum((r.get("total_input_tokens", 0) or 0) + (r.get("total_output_tokens", 0) or 0) for r in cost_summary)
    total_successes = sum(r.get("successes", 0) or 0 for r in cost_summary)
    success_rate = (total_successes / total_calls * 100) if total_calls > 0 else 100.0

    model_configs = list_agent_model_configs(db)

    return render_template("agents.html",
        cost_summary=cost_summary,
        lessons=lessons,
        logs=logs,
        days=days,
        total_cost=total_cost,
        total_calls=total_calls,
        total_tokens=total_tokens,
        success_rate=success_rate,
        model_configs=model_configs,
    )


@app.route("/agents/lesson/<int:lesson_id>/rate", methods=["POST"])
def rate_lesson(lesson_id):
    """Rate a lesson up or down."""
    db = get_db()
    data = request.get_json() or {}
    try:
        delta = int(data.get("delta", 0))
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "delta must be a valid integer"}), 400
    if delta not in (-1, 1):
        return jsonify({"ok": False, "error": "delta must be 1 or -1"}), 400

    try:
        db.execute("UPDATE agent_lessons SET rating = rating + ? WHERE id = ?", (delta, lesson_id))
        db.commit()
        row = db.execute("SELECT rating FROM agent_lessons WHERE id = ?", (lesson_id,)).fetchone()
        return jsonify({"ok": True, "rating": row["rating"] if row else 0})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]}), 500


@app.route("/agents/lesson/<int:lesson_id>/toggle", methods=["POST"])
def toggle_lesson(lesson_id):
    """Toggle a lesson active/inactive."""
    db = get_db()
    try:
        row = db.execute("SELECT active FROM agent_lessons WHERE id = ?", (lesson_id,)).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Lesson not found"}), 404
        new_active = 0 if row["active"] else 1
        db.execute("UPDATE agent_lessons SET active = ? WHERE id = ?", (new_active, lesson_id))
        db.commit()
        return jsonify({"ok": True, "active": new_active})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]}), 500


# ── Init & Run ───────────────────────────────────────────────────────────────

init_db()

# Initialize agent tables (lessons, logs)
try:
    _agent_db = sqlite3.connect(DB_PATH)
    init_agent_tables(_agent_db)
    _agent_db.close()
except Exception as _e:
    logger.warning(f"Agent table init: {_e}")

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=os.getenv("FLASK_DEBUG", "false").lower() == "true")
